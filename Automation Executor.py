#System variable and io handling
import sys
import os
import multiprocessing
from multiprocessing import Process , Queue, Manager
import queue 
import subprocess
#Get timestamp
import time
from datetime import datetime
import configparser
import random
import csv
import json
#GUI
from tkinter import IntVar, StringVar
from tkinter import Menu, filedialog, messagebox, Radiobutton
from tkinter import Text
from tkinter import colorchooser
from tkinter import scrolledtext 
from tkinter import Button
from tkinter import Toplevel , Tk, Frame
# sticky state
from tkinter import W, E, S, N, END, X, Y, BOTH, TOP, BOTTOM, CENTER, LEFT, NO
# Config state
from tkinter import DISABLED, NORMAL, WORD


from tkinter.ttk import Entry, Label, Style, Treeview, Scrollbar
from tkinter.ttk import Checkbutton, OptionMenu, Notebook
from xml.dom.expatbuilder import parseFragmentString

import importlib.machinery
import importlib.util

#from tkinter import style

from openpyxl import load_workbook


#from libs.general_function import *
# Function use for the action builder here
from libs.automation_driver import Automation as Tester
from libs.automation_driver import *

from ppadb.client import Client as AdbClient

from libs.configmanager import ConfigLoader
from libs.version import get_version
from libs.tkinter_extension import AutocompleteCombobox, AutomationTool_BottomPanel, CreateToolTip
import cv2
import pytesseract


CWD = os.path.abspath(os.path.dirname(sys.argv[0]))
ADBPATH = '\"' + CWD + '\\adb\\adb.exe' + '\"'
#MyTranslatorAgent = 'google'
TOOL = "Auto Tester"
VERNUM = ' 1.0.1h'
VERSION = TOOL  + " " +  VERNUM
DELAY1 = 20
DELAY2 = 100
DELAY3 = 200
IDLING_CHECK = 0

#**********************************************************************************
# UI handle ***********************************************************************
#**********************************************************************************

class Automation_Execuser(Frame):
	def __init__(self, Root, Queue = None, Manager = None,):
		
		Frame.__init__(self, Root) 
		self.pack(side=TOP, expand=Y, fill=X)
		#super().__init__()
		self.parent = Root 
		self.parent.protocol("WM_DELETE_WINDOW", self.on_closing)
		# Queue
		self.Process_Queue = Queue['Process_Queue']
		self.Result_Queue = Queue['Result_Queue']
		self.Status_Queue = Queue['Status_Queue']
		self.Debug_Queue = Queue['Debug_Queue']

		self.Manager = Manager['Default_Manager']

		self.Options = {}

		# UI Variable
		self.Button_Width_Full = 20
		self.Button_Width_Half = 15
		
		self.PadX_Half = 5
		self.PadX_Full = 10
		self.PadY_Half = 5
		self.PadY_Full = 10
		self.StatusLength = 120
		self.AppLanguage = 'en'

		self.DB_Path = ""
		self.Test_Case_Path = ""
		self.TestCase = None


		self.App_LanguagePack = {}

		self.init_App_Setting()

		if self.AppLanguage != 'kr':
			from libs.languagepack import LanguagePackEN as LanguagePack
		else:
			from libs.languagepack import LanguagePackKR as LanguagePack

		self.LanguagePack = LanguagePack
		
		#**************New row#**************#
		self.Notice = StringVar()
		self.Debug = StringVar()
		self.Progress = StringVar()
	
		self.AutoTester = Tester(self.Status_Queue)

		#Generate UI
		self.bottom_panel = AutomationTool_BottomPanel(self)
		self.init_UI()
		self.init_UI_Data()

		try:
			print('Start server')
			os.popen( 'adb start-server')
		except Exception as e:
			print('Error:', e)	
		'''
		try:
			print('Start server')
			os.popen( ADBPATH + ' start-server')
			#app_name = 'android_touch'
			#print('Get CPU profile')
			#self.CPU = os.popen(ADBPATH + ' shell getprop ro.product.cpu.abi').read()
			#process = subprocess.Popen(ADBPATH + ' shell getprop ro.product.cpu.abi', stdout=subprocess.PIPE, stderr=None, shell=True)
			#CPU = process.communicate()[0].decode("utf-8") 
			#CPUFamily = CPU.replace('\r\n', "")
			#print('CPU family: ', CPUFamily)
			self.port = 0
			#if CPUFamily != "":
			#print('Push touch to device')
			#if (os.path.isfile(CWD + '\\libs\\arm64-v8a\\touch')):
			#	print('file available')
			#str_command = '%s push \"%s\\libs\\%s\\touch\" /data/local/tmp' % (ADBPATH, CWD, CPUFamily)
			#print('Command:', str_command)
			#os.system(str_command)
			print('Launch touch on device')
			os.system(ADBPATH + ' shell chmod 755 /data/local/tmp/touch') 
			os.system(ADBPATH + ' shell /data/local/tmp/touch')
			print('Looking for port')
			for port in range(50000,65535):
				self.port = port
				process = subprocess.Popen(ADBPATH + ' forward tcp:9889 tcp:' + str(self.port), stdout=subprocess.PIPE, stderr=None, shell=True)
				return_message = process.communicate()
				for message in return_message:
					if message != None:
						str_message = message.decode("utf-8") 
						if str_message == self.port:
							break
		
				#print('str_message', str_message)
				return_port = str_message.replace('\r\n', "")
				#print('return_port', return_port)
				if return_port == self.port:
					break
			print('Current port:', self.port)

		except Exception as e:
			print('Error:', e)
		'''
		#print('Get device serial')
		#self.Get_Serial()
		self.parent.bind('<Control-l>', self.Btn_Browse_Test_Case_File)
		self.parent.bind('<Control-s>', self.Btn_Save_Test_Case_File) 

	
		self.parent.bind('<Control-x>', self.cut_treeview) 
		self.parent.bind('<Control-c>', self.copy_treeview)
		self.parent.bind('<Control-v>', self.paste_treeview) 
		self.after(DELAY2, self.status_listening)

	# UI init
	def init_UI(self):

		self.parent.resizable(False, False)
		self.parent.title(VERSION)

		self.Generate_Menu_UI()
		self.Generate_Tab_UI()
		
		self.Generate_Automation_Execution_UI(self.Main)
		self.Generate_OCR_Setting_UI(self.OCR_SETTING)
		self.Generate_Debugger_UI(self.Controller)
		#self.Generate_Folder_Comparision_UI(self.FolderComparison)
		#self.Generate_Optimizer_UI(self.Optimizer)
		#self.Generate_Debugger_UI(self.Process)
		
		# Debugger

	#UI Function

	def Generate_Menu_UI(self):
		menubar = Menu(self.parent) 
		# Adding File Menu and commands 
		file = Menu(menubar, tearoff = 0)
		menubar.add_cascade(label =  self.LanguagePack.Menu['File'], menu = file) 
		
		file.add_command(label =  self.LanguagePack.Menu['Load'], command = self.Btn_Browse_Test_Case_File) 
		file.add_command(label =  self.LanguagePack.Menu['Save'], command = self.Btn_Save_Test_Case_File) 
		file.add_separator()
		file.add_command(label =  self.LanguagePack.Menu['CreateTCTemplate'], command = self.Menu_Function_Save_TC_Template) 
		file.add_command(label =  self.LanguagePack.Menu['CreateDBTemplate'], command = self.Menu_Function_Save_DB_Template) 
		file.add_command(label =  self.LanguagePack.Menu['CreateExecuteListTemplate'], command = self.Menu_Function_Save_Execute_List_Template) 
		file.add_separator()
		file.add_command(label =  self.LanguagePack.Menu['Exit'], command = self.on_closing) 
		
		
		# Adding Help Menu
		help_ = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Help'], menu = help_) 
		help_.add_command(label =  self.LanguagePack.Menu['GuideLine'], command = self.Menu_Function_Open_Main_Guideline) 
		help_.add_separator()
		help_.add_command(label =  self.LanguagePack.Menu['About'], command = self.Menu_Function_About) 
		self.parent.config(menu = menubar)

		# Adding Help Menu
		language = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Language'], menu = language) 
		language.add_command(label =  self.LanguagePack.Menu['Hangul'], command = self.SetLanguageKorean) 
		language.add_command(label =  self.LanguagePack.Menu['English'], command = self.SetLanguageEnglish) 
		self.parent.config(menu = menubar) 	

	def Generate_Tab_UI(self):
		
		MainPanel = Frame(self, name='mainpanel')
		MainPanel.pack(side=TOP, fill=BOTH, expand=Y)
		TAB_CONTROL = Notebook(MainPanel, name='notebook')
		TAB_CONTROL.enable_traversal()

		self.Main = Frame(TAB_CONTROL)
		TAB_CONTROL.add(self.Main, text= self.LanguagePack.Tab['Main'])
		
		self.OCR_SETTING = Frame(TAB_CONTROL)
		TAB_CONTROL.add(self.OCR_SETTING, text= self.LanguagePack.Tab['Setting'])
		#Tab
		self.Controller = Frame(TAB_CONTROL)
		TAB_CONTROL.add(self.Controller, text= self.LanguagePack.Tab['TestTool'])
	
		TAB_CONTROL.pack(side=TOP, fill=BOTH, expand=Y)

		#self.TAB_CONTROL.pack(expand=1, fill="both")
		

	def Generate_Automation_Execution_UI(self, Tab):
		Row = 1
		#self.Str_Test_Case_Path.set( CWD + '\\Testcase\\Sample_Automation_Testcase.xlsx')
		#Label(Tab, text=  self.LanguagePack.Label['TestCaseList']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		#self.Entry_New_File_Path = Entry(Tab,width = 80, state="readonly", textvariable=self.Str_Test_Case_Path)
		#self.Entry_New_File_Path.grid(row=Row, column=3, columnspan=7, padx=4, pady=5, sticky=W+E)
		#Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_Test_Case_File).grid(row=Row, column=10, padx=5, pady=0, sticky=W)
		#Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectFontColor'], command= self.Btn_Select_Font_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		
	
		Row += 1
		ExecuteListLabel = Label(Tab, text= self.LanguagePack.Label['ExecuteList'])
		ExecuteListLabel.grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky=W)
		CreateToolTip(ExecuteListLabel, self.LanguagePack.ToolTips['ExecuteList'])
		self.ExecuteList = AutocompleteCombobox(Tab)
		#self.TestProject = AutocompleteCombobox(Tab)
		self.ExecuteList.set_completion_list([])
		self.ExecuteList.Set_Entry_Width(30)
		self.ExecuteList.grid(row=Row, column=3, padx=5, pady=5, sticky=W)
		Button(Tab, width = self.Button_Width_Half, text= self.LanguagePack.Button['Browse'], command= self.Btn_Browse_Execution_List).grid(row=Row, column=4, columnspan=2, padx=5, pady=0, sticky=W)
		


		Label(Tab, text=self.LanguagePack.Label['Serial']).grid(row=Row, column=7, padx=5, pady=5, sticky=W)			
		self.TextSerial = AutocompleteCombobox(Tab)
		self.TextSerial.Set_Entry_Width(20)
		self.TextSerial.bind("<<ComboboxSelected>>", self.ADB_Connect)

		self.TextSerial.grid(row=Row, column=8, columnspan=2,padx=5, pady=5, sticky=W+E)
		self.TextSerial.set_completion_list([])
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Refresh'], command= self.Get_Serial).grid(row=Row, column=10,padx=5, pady=0, sticky=W)
		
		'''
		Row += 1
		Label(Tab, text=self.LanguagePack.Label['TestFeature']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.TestFeature = AutocompleteCombobox(Tab)
		self.TestFeature.set_completion_list(['Gacha'])
		self.TestFeature.Set_Entry_Width(30)
		self.TestFeature.grid(row=Row, column=3, padx=5, pady=5, sticky=W)
		'''
		
		#Button(Tab, width = self.Button_Width_Half, text=  'Get Test info', command= self.Btn_Generate_TestCase).grid(row=Row, column=10,padx=0, pady=0, sticky=W)

		Row += 1
	
		ExecuteTypeLabel = Label(Tab, text= self.LanguagePack.Label['ExecuteType'])
		ExecuteTypeLabel.grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky=W)
		CreateToolTip(ExecuteTypeLabel, self.LanguagePack.ToolTips['ExecuteList'])
		self.ExecuteType = AutocompleteCombobox(Tab)
		#self.TestProject = AutocompleteCombobox(Tab)
		self.ExecuteType.set_completion_list(['', 'Current Value', 'List'])
		self.ExecuteType.Set_Entry_Width(30)
		self.ExecuteType.current(0)
		self.ExecuteType.grid(row=Row, column=3, padx=5, pady=5, sticky=W)
		self.ExecuteType.configure(state=DISABLED)

		Label(Tab, text= self.LanguagePack.Label['WorkingRes']).grid(row=Row, column=7, rowspan=2, padx=5, pady=0, sticky=W)
		Radiobutton(Tab, width= 10, text=  '720p', value=1, variable=self.Resolution, command= self.Auto_Setting_Set_Working_Resolution).grid(row=Row, column=8, padx=5, pady=5, sticky=W)
		#Label(Tab, text= "Device IP").grid(row=Row, column=7, padx=5, pady=5, sticky=W)	
		#self.Device_IP = Text(Tab, width=30, height=1, undo=True, wrap=WORD)
		#self.Device_IP.grid(row=Row, column=8, columnspan=2, padx=5, pady=5, sticky=E)
		#self.Device_IP.insert("end", "0.0.0.0")
		#Button(Tab, width = self.Button_Width_Half, text=  "Wireless Connect", command= self.Connect_Device).grid(row=Row, column=10, padx=5, pady=0, sticky=W)
		
		Row +=1
		TessLanguageLabel = Label(Tab, text= self.LanguagePack.Label['WorkingLang'])
		TessLanguageLabel.grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky=W)
		CreateToolTip(TessLanguageLabel, self.LanguagePack.ToolTips['TessLanguage'])
		self.option_working_language = AutocompleteCombobox(Tab)
		self.option_working_language.Set_Entry_Width(30)

		self.option_working_language.grid(row=Row, column=3, padx=5, pady=5, sticky=W)
		Btn_Refresh_OCR_Lang = Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Refresh'], command= self.Btn_OCR_Update_Working_Language)
		Btn_Refresh_OCR_Lang.grid(row=Row, column= 4, columnspan=2, padx=5, pady=0, sticky=W)	
		CreateToolTip(Btn_Refresh_OCR_Lang, 'Load all supported language from TessData folder.')

		
		
		Radiobutton(Tab, width= 10, text=  '1080p', value=2, variable=self.Resolution, command= self.Auto_Setting_Set_Working_Resolution).grid(row=Row, column=8, padx=5, pady=5, sticky=W)
	
		Row+=1
		Label(Tab, text= self.LanguagePack.Label['ActionType'], width= 20).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky=W)
		action_type_list = ['', 'Loop', 'Condition', 'Get_Result', 'Action', 'Update_Variable', 'Comment']
		action_type_list.sort()
		self.current_action_type.set(action_type_list[0])
		# creating widget
		self.action_type = OptionMenu(Tab, self.current_action_type,	*action_type_list, command=self.Update_Action_Name)
		self.action_type.config(width=30)
		#self.action_type.Set_Entry_Width(20)
		#self.action_type.set_completion_list(['Loop', 'Condition', 'Get_Result', 'Action', 'Update_Variable',])
		self.action_type.grid(row=Row, column=3, columnspan=2, padx=5, pady=5, sticky=W)


		Label(Tab, text= self.LanguagePack.Label['ActionName'], width= 20).grid(row=Row, column=5, padx=5, pady=5, sticky=W)
		
		# creating widget
		action_name_list = ['']
		self.current_action_name.set(action_name_list[0])
		self.action_name = OptionMenu(Tab, self.current_action_name, *action_name_list, command=self.Update_Action_Arg)
		self.action_name.config(width=30)
		#self.action_type.set_completion_list(['Loop', 'Condition', 'Get_Result', 'Action', 'Update_Variable',])
		self.action_name.grid(row=Row, column=6, columnspan=2, padx=5, pady=5, sticky=W)
	
		self.btn_input_data = Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['InputData'], command= self.Generate_Input_Window)
		self.btn_input_data.grid(row=Row, column=10, padx=5, pady=0, sticky=W+E)

		Row+=1
		Treeview_Row = 20
		self.Generate_Treeview_Advanced_UI(Tab, Row, Treeview_Row)

		#Row+=1
		#self.btn_add_step_above = Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['AddOnTop'], command= self.insert_treeview_above)
		#self.btn_add_step_above.grid(row=Row, column=10,padx=5, pady=0, sticky=W)

		Row+=1
		self.btn_add_step = Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['AddStep'], command= self.add_treeview_row)
		self.btn_add_step.grid(row=Row, column=10,padx=5, pady=0, sticky=W)
		#Row+=1
		#Button(Tab, width = self.Button_Width_Half, text=  'Browse Module', command= self.Browse_Module).grid(row=Row, column=10,padx=5, pady=0, sticky=W)

		#Row+=1
		#Button(Tab, width = self.Button_Width_Half, text=  "Move Up", command= self.move_treeview_up).grid(row=Row, column=10,padx=5, pady=0, sticky=W)
		#Row+=1
		#Button(Tab, width = self.Button_Width_Half, text=  "Move Down", command= self.move_treeview_down).grid(row=Row, column=10,padx=5, pady=0, sticky=W)
		#Row+=1
		#Button(Tab, width = self.Button_Width_Half, text=  "Load TC", command= self.Btn_Browse_Test_Case_File).grid(row=Row, column=10,padx=5, pady=0, sticky=W)
		#Row+=1
		#Button(Tab, width = self.Button_Width_Half, text=  "Save TC", command= self.Btn_Save_Test_Case_File).grid(row=Row, column=10,padx=5, pady=0, sticky=W)	
		#Row+=1
		


		Row += Treeview_Row

		
	
		Row+=1

		self.Debugger = scrolledtext.ScrolledText(Tab, width=100, height=10, undo=True, wrap=WORD, )
		self.Debugger.grid(row=Row, column=1, columnspan=9, rowspan=10, padx=5, pady=5, sticky=W+E+N+S)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['ClearLog'], command= self.ClearLog).grid(row=Row, column=10,padx=5, pady=0, sticky=W)
		Row+=1
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Stop'], command= self.Stop).grid(row=Row, column=10,padx=5, pady=0, sticky=W)	
		Row+=1
		self.Btn_Pause = Button(Tab, width = self.Button_Width_Half, text= self.LanguagePack.Button['Pause'], command= self.Pause, fg='red',font='sans 9 bold')
		self.Btn_Pause.grid(row=Row, column=10,padx=5, pady=0, sticky=W)	
		Row+=1
		self.Btn_Execute = Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.Btn_Execute_Script)
		self.Btn_Execute.grid(row=Row, column=10,padx=5, pady=0, sticky=W)
		#ScrollBar = Scrollbar(Tab, bg="green")
		#ScrollBar.pack( side = RIGHT, fill = Y )	

	def Generate_Treeview_Advanced_UI(self, Tab, start_row, row_span):
		TreeView_Row = start_row
		TreeView_Size = row_span
		Treeview_Col = 9
		self.Treeview = Treeview(Tab)
		self.Focused_Item = None
		self.Treeview.grid(row=TreeView_Row, column=1, columnspan=Treeview_Col, rowspan=20, padx=5, pady=5, sticky = W+E)
		verscrlbar = Scrollbar(Tab, orient ="vertical", command = self.Treeview.yview)
		self.Treeview.configure( yscrollcommand=verscrlbar.set)
	
		self.Treeview.Scrollable = True
		self.Treeview['columns'] = ('Type', 'Action', 'Arg1', 'Arg2', 'Arg3', 'Arg4', 'Arg5', 'Arg6')

		self.Treeview.column('#0', width=0, stretch=NO)
		self.Treeview.heading('#0', text='', anchor=CENTER)

		for column in self.Treeview['columns']:
			if column == 'Action':
				self.Treeview.column(column, anchor=CENTER, width=200)
				self.Treeview.heading(column, text=column, anchor=CENTER)
			else:	
				self.Treeview.column(column, anchor=CENTER, width=100)
				self.Treeview.heading(column, text=column, anchor=CENTER)

		verscrlbar.grid(row=TreeView_Row, column=Treeview_Col, rowspan=TreeView_Size, padx=5, pady=5,  sticky = N+S+E)
		Tab.grid_columnconfigure(TreeView_Size, weight=0, pad=5)
		styles = Style()
		styles.configure('Treeview',rowheight=row_span)

		self.Treeview.bind("<Delete>", self.delete_treeview_line)	
		self.Treeview.bind("<Button-1>", self.Treeview_Deselect_Row)
		self.Treeview.bind("<Double-1>", self.Treeview_Edit_Row)
		self.Treeview.bind("<Button-3>", self.Treeview_Show_Right_Menu)
		self.createWindow()

		self.Treeview.tag_configure('comment', foreground="#006C78")
		self.Treeview.tag_configure('loop', foreground="#B8516A")
		self.Treeview.tag_configure('condition', foreground="#FF7722")
		self.Treeview.tag_configure('action', foreground="#3B4248")

	def Generate_OCR_Setting_UI(self, Tab):
		''''
		Create Setting Tab
		'''
		Row = 1
		Label(Tab, text= self.LanguagePack.Label['TesseractPath']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Text_TesseractPath = Entry(Tab,width = 100, state="readonly", textvariable=self.TesseractPath)
		self.Text_TesseractPath.grid(row=Row, column=3, columnspan=5, padx=5, pady=5, sticky=E+W)
		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Select_Tesseract_Path).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
		Row += 1
		Label(Tab, text= self.LanguagePack.Label['TesseractDataPath']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Text_TesseractDataPath = Entry(Tab,width = 100, state="readonly", textvariable=self.TesseractDataPath)
		self.Text_TesseractDataPath.grid(row=Row, column=3, columnspan=5, padx=5, pady=5, sticky=E+W)
		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Select_Tesseract_Data_Path).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
		Row += 1
		Label(Tab, text= self.LanguagePack.Label['DBPath']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Entry_DB_Path = Entry(Tab,width = 100, state="readonly", textvariable=self.Text_DB_Path)
		self.Entry_DB_Path.grid(row=Row, column=3, columnspan=5, padx=5, pady=5, sticky=E+W)
		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_DB_File).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
		Row += 1
		Label(Tab, text= 'Custom Action').grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Entry_CustomAction_Path = Entry(Tab,width = 100, state="readonly", textvariable=self.Text_CustomAction_Path)
		self.Entry_CustomAction_Path.grid(row=Row, column=3, columnspan=5, padx=5, pady=5, sticky=E+W)
		Button(Tab, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_DB_File).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		
	def Generate_Debugger_UI(self,Tab):
		Row = 1
		Button(Tab, width = self.Button_Width_Half, text=  "TAB", command= self.Btn_Send_Tab).grid(row=Row, column=1,padx=5, pady=5, sticky=W)
		Button(Tab, width = self.Button_Width_Half, text=  "Enter", command= self.Btn_Send_Enter).grid(row=Row, column=2,padx=5, pady=5, sticky=W)
		Button(Tab, width = self.Button_Width_Half, text=  "4 touch", command= self.Btn_Send_4_Touch).grid(row=Row, column=3,padx=5, pady=5, sticky=W)
		Row += 1
		Button(Tab, width = self.Button_Width_Half, text=  "Home", command= self.Btn_Send_Home).grid(row=Row, column=1,padx=5, pady=5, sticky=W)
		Button(Tab, width = self.Button_Width_Half, text=  "Backkey", command= self.Btn_Send_Backkey).grid(row=Row, column=2,padx=5, pady=5, sticky=W)

		Row += 1
		self.Str_Template_Path = StringVar()
		Label(Tab, text= 'Template path').grid(row=Row, column=1, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = 80, state="readonly", textvariable=self.Str_Template_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=2, columnspan=5, padx=5, pady=5, sticky=W)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_Template_File).grid(row=Row, column=8, padx=5, pady=5, sticky=W)
		Button(Tab, width = self.Button_Width_Half, text= 'Tap', command= self.Btn_Tap_Template).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		
###########################################################################################
# OCR Setting
###########################################################################################

	def Btn_Select_Tesseract_Path(self):
		filename = filedialog.askopenfilename(title =  'Please select Tesseract.exe file',filetypes = (("Executable files","*.exe" ), ), )	
		if os.path.isfile(filename):
			_tess_path = self.CorrectPath(filename)
			self.AppConfig.Save_Config(self.AppConfig.Auto_Tool_Config_Path, 'AUTO_TOOL', 'tess_path', _tess_path, True)
			pytesseract.pytesseract.tesseract_cmd = _tess_path
			self.TesseractPath.set(_tess_path)
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['TessNotSelect'])

	def Btn_Select_Tesseract_Data_Path(self):
		folder_name = filedialog.askdirectory(title =  'Please select Tesseract data folder',)	
		if os.path.isdir(folder_name):
			folder_name = self.CorrectPath(folder_name)
			self.TesseractDataPath.set(folder_name)

			self.AppConfig.Save_Config(self.AppConfig.Auto_Tool_Config_Path, 'AUTO_TOOL', 'tess_data', folder_name, True)

			self.Write_Debug(self.LanguagePack.ToolTips['TessDataUpdate'] + folder_name)
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])

	def Btn_Browse_DB_File(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'], filetypes = (("Workbook files", "*.xlsx *.xlsm, *.csv"), ), multiple = False)	
		if filename != "":
			self.DB_Path = self.CorrectPath(filename)
			self.DB_Folder = os.path.dirname(self.DB_Path)
			Init_Folder(self.DB_Folder)
			self.Text_DB_Path.set(self.DB_Path)
			
			self.AppConfig.Save_Config(self.AppConfig.Auto_Tool_Config_Path, 'AUTO_TOOL', 'db_path', self.DB_Path, True)
			self.AutoTester.Update_DB_Path(self.DB_Path)
			self.Write_Debug(self.LanguagePack.ToolTips['DataSelected'] + ": " + self.DB_Path)

	
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Auto_Setting_Set_Working_Resolution(self):
		_resolution_index = self.Resolution.get()
		if _resolution_index == 1:
			self.WorkingResolution = 720
		else:
			self.WorkingResolution = 1080
		
		self.AppConfig.Save_Config(self.AppConfig.Auto_Tool_Config_Path, 'AUTO_TOOL', 'resolution', _resolution_index)
		self.AutoTester.Update_Resolution(self.WorkingResolution)
		self.Write_Debug(self.LanguagePack.ToolTips['SetResolution'] + str(self.WorkingResolution) + 'p')

	def Auto_Setting_Set_Working_Language(self, select_value):
		
		self.AppConfig.Save_Config(self.AppConfig.Auto_Tool_Config_Path, 'AUTO_TOOL', 'scan_lang', select_value)
		
		self.Write_Debug(self.LanguagePack.ToolTips['SetScanLanguage'] + str(select_value))
	
	

###########################################################################################
# General functions
###########################################################################################

	def CorrectPath(self, path):
		if sys.platform.startswith('win'):
			return str(path).replace('/', '\\')
		else:
			return str(path).replace('\\', '//')
	
	def CorrectExt(self, path, ext):
		if path != None and ext != None:
			Outputdir = os.path.dirname(path)
			baseName = os.path.basename(path)
			sourcename = os.path.splitext(baseName)[0]
			newPath = self.CorrectPath(Outputdir + '/'+ sourcename + '.' + ext)
			return newPath

	def Write_Debug(self, text):
		'''
		Function write the text to debugger box and move to the end of the box
		'''
		
		self.Debugger.insert("end", "\n")
		self.Debugger.insert("end", str(datetime.now()) + ': ' + str(text))

		self.Debugger.yview(END)		

	def entry_next(self, event):
		event.widget.tk_focusNext().focus()
		return("break")

	def Get_Serial(self):
		try:
			client = AdbClient(host="127.0.0.1", port=5037)
			devices = client.devices()
			Serrial = []
			for device in devices:
				if device.serial != '':
					Serrial.append(device.serial)
			if len(Serrial) == 0:
				messagebox.showinfo('Error...', 'No device found')	
				return
			self.TextSerial.set_completion_list(Serrial)
			self.TextSerial.set(Serrial[0])
			self.AutoTester.Update_Serial_Number(Serrial[0])
		except:	
			pass
		

	def ADB_Connect(self, event):
		self.AutoTester.Update_Serial_Number(self.TextSerial.get())
		
	def Connect_Device(self):
		IP = self.Device_IP.get("1.0", END).replace('\n', '')
		os.popen( ADBPATH + ' tcpip 5555')
		os.popen( ADBPATH + ' connect ' + str(IP) + ':5555')
		self.Get_Serial()	

	def Btn_Send_Tab(self):
		os.popen( ADBPATH + ' shell input keyevent \'61\'')
	
	def Btn_Send_Enter(self):
		os.popen( ADBPATH + ' shell input keyevent \'66\'')	

	def Btn_Send_4_Touch(self):
		os.system( ADBPATH + ' forward tcp:9889 tcp:9889')
		#Four_Touch()

	def Btn_Send_Backkey(self):
		os.popen( ADBPATH + ' shell input keyevent \'4\'')	

	def Btn_Send_Home(self):
		os.popen( ADBPATH + ' shell input keyevent \'3\'')		

	def Btn_Tap_Template(self):
		if self.Template_Path != None:
			self.Get_Serial()
			Serial = self.TextSerial.get().replace('\n','')
			AutoTester = Tester(self.Status_Queue, Serial, self.DB_Path)
			AutoTester.Tap_Template(self.Template_Path)	

	def Stop(self):
		try:
			if self.Automation_Processor.is_alive():
				self.Automation_Processor.terminate()
		except:
			pass
		self.Show_Error_Message('Test is terminated')

	def Pause(self):
		
		_text = self.Btn_Pause.cget("text")
		#print('Text', _text)
		if _text == self.LanguagePack.Button['Pause']:
			self.Result_Queue.put('Pause')
			self.Btn_Pause.configure(text=self.LanguagePack.Button['Resume'])
			self.Btn_Pause.configure(fg='green')
			
		else:
			#print('Set button to Pause')
			self.Result_Queue.put('Resume')
			self.Btn_Pause.configure(text=self.LanguagePack.Button['Pause'])
			self.Btn_Pause.configure(fg='red')

	# Other function
	def Btn_OCR_Update_Working_Language(self):
		_data_ = str(self.TesseractDataPath.get())
		_exe_ = str(self.TesseractPath.get())
		_tessdata_dir_config = '--tessdata-dir ' + "\"" + _data_ + "\""
		pytesseract.pytesseract.tesseract_cmd = _exe_
		#self.language_list = pytesseract.get_languages(config=_tessdata_dir_config)
		try:
			self.language_list = pytesseract.get_languages(config=_tessdata_dir_config)
			self.Write_Debug(self.LanguagePack.ToolTips['TessLangUpdated'])

		except Exception as e:
			#self.Write_Debug('Tess path: ' + str(_exe_))
			#self.Write_Debug('Data path: ' + str(_data_))
			self.Write_Debug(self.LanguagePack.ToolTips['TessLangLoadFail'] + str(e))
			self.language_list = ['']

		self.option_working_language.set_completion_list(self.language_list)



	def ClearLog(self):
		self.Debugger.delete('1.0', END)
		return

	

###########################################################################################
# Treeview FUNCTION
###########################################################################################

	def delete_treeview_line(self, event):
		'''
		Function activate when select an entry from a Treeview and press Delete btn
		'''
		selected = self.Treeview.selection()
		to_remove = []
		for child_obj in selected:
			child = self.Treeview.item(child_obj)
			tm_index = child['values'][0]
			to_remove.append(tm_index)
			self.Treeview.delete(child_obj)

	# Obsoleted.
	def double_right_click_treeview(self, event):
		'''
		Function activate when double click an entry from Treeview
		'''
		focused = self.Treeview.focus()
		child = self.Treeview.item(focused)
		
		self.Debugger.insert("end", "\n")
		self.Debugger.insert("end", 'Korean: ' + str(child["text"]))
		self.Debugger.insert("end", "\n")
		self.Debugger.insert("end", 'English: ' + str(child["values"][0]))
		self.Debugger.yview(END)


	# Nam will check
	def load_tm_list(self):
		"""
		When clicking the [Load] button in TM Manager tab
		Display the pair languages in the text box.
		"""
		self.remove_treeview()
		
		_area_list = []

		for location in _area_list:	
			try:
				self.Treeview.insert('', 'end', text= '', values=( str(location['index']), str(location['x']), str(location['y']), str(location['h']), str(location['w'])))
			except:
				pass

	def add_treeview_row(self):
		'''
		Add a row to the current Treeview
		'''
		selected = self.Treeview.selection()
	
		index_list = []
		for child_obj in selected:
			#child = self.Treeview.item(child_obj)
			index = self.Treeview.index(child_obj)
			index_list.append(index)
		index_list.sort()	

		if len(selected) == 0:
			start = 'end'
			end = 'end'
		else:
			start = index_list[-1]+1
			end = index_list[-1]+2

		this_type = self.current_action_type.get()
		#print('this_type', this_type)
		this_action = self.current_action_name.get()
		#print('this_action', this_action)
		if self.Current_Arg_Type == None:
			this_tag = ''
		else:	
			this_tag = self.Current_Arg_Type.lower()

		if self.Current_Arg_Type == None:	
			if this_type != '' and this_action != '':
				self.Treeview.insert('', start, text= '', values = [this_type, this_action], tags= this_tag)
				return
			else:	 
				messagebox.showwarning('Warning', 'Please input the value before adding an action.')
				return

		#self.Treeview.insert('', start, text= '', values = [this_type, this_action], tags= this_tag)
		self.Treeview.insert('', start, text= '', values = [*self.Current_Arg_Value], tags= this_tag)
		if self.Current_Arg_Type == 'Loop':
			self.Treeview.insert('', end, text= '', values=['Loop','End Loop'], tags= this_tag)
		elif self.Current_Arg_Type == 'Condition':
			self.Treeview.insert('', end, text= '', values= ['Condition', 'End If'], tags= this_tag)
		else:
			print('Type:', self.Current_Arg_Type)	
		
		#index = self.Treeview.index(self.Treeview.focus()) + 1
		#self.Generate_Arg_Input_Window(index_list[-1]+1, index_list[-1]+2)	

		

	def insert_treeview_above(self):
		'''
		Add a row above the current Treeview
		'''
		selected = self.Treeview.selection()
		
		
		index_list = []
		for child_obj in selected:
			#child = self.Treeview.item(child_obj)
			index = self.Treeview.index(child_obj)
			index_list.append(index)
		index_list.sort()
		if len(selected) == 0:
			start = 0
			end = 1
		else:
			start = index_list[0]
			if len(index_list) == 1:
				end = start+1+len(index_list)
			else:
				end = index_list[-1]+1+len(index_list)
		if self.Current_Arg_Type == None:
			this_type = self.current_action_type.get()
			this_action = self.current_action_name.get()
			if this_type != '' and this_action != '':
				if this_type == 'Comment':
					self.Treeview.insert('', start, text= '', values = [this_type, this_action], tags=('comment',))
				else:	
					self.Treeview.insert('', start, text= '', values = [this_type, this_action])
				return
			else:	
				messagebox.showwarning('Warning', 'Please input the value before adding an action.')
				return
		self.Treeview.insert('', start, text= '', values = self.Current_Arg_Value)
		
		if self.Current_Arg_Type == 'Loop':
			self.Treeview.insert('', end, text= '', values=['Loop','End Loop'])
		elif self.Current_Arg_Type == 'Condition':
			self.Treeview.insert('', end, text= '', values= ['Condition', 'End If'])
		else:
			print('Type:', self.Current_Arg_Type)	

	def insert_treeview_below(self):
		'''
		Add a row below the current Treeview
		'''
		selected = self.Treeview.selection()
		index_list = []
		for child_obj in selected:
			#child = self.Treeview.item(child_obj)
			index = self.Treeview.index(child_obj)
			index_list.append(index)
		index_list.sort()	
		#print(index_list)
		#index = self.Treeview.index(self.Treeview.focus()) + 1
		if len() == 1:
			start = index_list[0]
			end = index_list[0]+1
		else:
			start = index_list[0]+1
			end = index_list[-1]+1
		
		if self.Current_Arg_Type == 'Comment':
			self.Treeview.insert('', start, text= '', values = self.Current_Arg_Value, tags=('comment',))
		else:	
			self.Treeview.insert('', start, text= '', values = self.Current_Arg_Value)

		if self.Current_Arg_Type == 'Loop':
			self.Treeview.insert('', end, text= '', values=['Loop','End Loop'])
		elif self.Current_Arg_Type == 'Condition':
			self.Treeview.insert('', end, text= '', values= ['Condition', 'End If'])
		else:
			print('Type:', self.Current_Arg_Type)		
		
		
	def move_treeview_up(self):
		leaves = self.Treeview.selection()
		#counter = 0
		for i in leaves:
			#counter+=1
			self.Treeview.move(i, self.Treeview.parent(i), self.Treeview.index(i)-1)
		
	def move_treeview_down(self):
		leaves = self.Treeview.selection()
		#counter = 0
		for i in reversed(leaves):
			#counter+=1
			self.Treeview.move(i, self.Treeview.parent(i), self.Treeview.index(i)+1)

	def remove_treeview(self):

		for i in self.Treeview.get_children():
			self.Treeview.delete(i)

	def copy_treeview(self, event = None):
		self.Write_Debug('Update copy list')
		selected = self.Treeview.selection()
		
		self.coppied_list = selected
		self.is_cut = False

	def cut_treeview(self, event = None):
		self.Write_Debug('Update cut list')
		selected = self.Treeview.selection()
		self.coppied_list = selected
		self.is_cut = True


	def paste_treeview(self, event = None):
		if self.coppied_list == None:
			messagebox.showwarning('None row is selected.')
			return
		selected = self.Treeview.selection()
		index_list = []
		for child_obj in selected:
			index = self.Treeview.index(child_obj)
			index_list.append(index)
		index_list.sort()	

		if len(selected) == 0:
			_end_index = 'end'
		else:
			_end_index = index_list[-1]+1
		self.Write_Debug('Add item to row ' + str(_end_index))
		for i in reversed(self.coppied_list):
			child = self.Treeview.item(i)
			values = child['values']
			self.Treeview.insert('', _end_index, text= '', values=values)

			if self.is_cut == True:
				self.Treeview.delete(i)
		self.coppied_list = None
		self.is_cut = False		
		'''
		if len() == 1:
			start = self.coppied_list[0]
			end = self.coppied_list[0]+1
		else:
			start = self.coppied_list[0]+1
			end = self.coppied_list[-1]+1

		for i in reversed(leaves):
			self.Treeview.insert('', start, text= '', values = self.Current_Arg_Value)
		
		for i in self.Treeview.get_children():
			self.Treeview.delete(i)	
		'''
		

	def Treeview_Deselect_Row(self, event):
		'''
		Function activate when double click an entry from Treeview
		'''
		if len(self.Treeview.selection()) > 0:
			self.Treeview.selection_remove(self.Treeview.selection()[0])

	def Treeview_Edit_Row(self, event):
		'''
		Function activate when double click an entry from Treeview
		'''
		self.Focused_Item = self.Treeview.focus()
		child = self.Treeview.item(self.Focused_Item)
		index = self.Treeview.index(self.Focused_Item)
		if len(child["values"]) > 2:
			_type = child["values"][0]
			_name = child["values"][1]
			self.Generate_Arg_Edit_Window(self.Focused_Item, _type, _name)

			

	def Treeview_Show_Right_Menu(self, event):
		item = self.Treeview.identify_row(event.y)
		try:
			self.popup.tk_popup(event.x_root, event.y_root, 0)
		finally:
			self.popup.grab_release()

	
	def createWindow(self):
		self.popup = Menu(self.parent, tearoff=0)
		self.popup.add_command(label="Move Up", command = self.move_treeview_up) # , command=next) etc...
		self.popup.add_command(label="Move Down", command = self.move_treeview_down)
		self.popup.add_separator()
		self.popup.add_command(label="Cut", command = self.cut_treeview)
		self.popup.add_command(label="Copy", command = self.copy_treeview)
		self.popup.add_command(label="Paste", command = self.paste_treeview)
		#self.popup.add_command(label="Exit", command=lambda: self.closeWindow())

	
	def Update_Execute_List(self):
		print('Update execution list.')
		try:
			print(self.TestCase['Data'])
			self.ExecuteList.set_completion_list(self.TestCase['Data'])
			self.ExecuteList.current(0)
		except:
			pass	

	def Update_Action_Arg(self, action_name=None):
		this_type = self.current_action_type.get()
		this_action = self.current_action_name.set(action_name)
		if action_name == 'Loop List':
			print(self.ExecuteType.get())
			if self.ExecuteType.get() == '':
				messagebox.showwarning('Warning', ' Loop list only available when you select a execution list.')
				this_action = self.current_action_name.set('')
				return
		this_arg = self.action_dict[this_type][action_name]
		self.btn_add_step.configure(state=DISABLED)
		print('this_arg', action_name, this_arg)
		
	# This is temporary work, we need to add the value of the arg instead of the arg name.
	# Will be updated later.
	def generate_step_data(self):
		this_type = self.current_action_type.get()
		this_action = self.current_action_name.get()
		input_value = ''
		values = [this_type] + [this_action] + input_value

		return values

	def Generate_Input_Window(self, this_type = None, this_action = None):
		if this_type == None:
			this_type = self.current_action_type.get()
		if this_type == '':
			messagebox.showwarning('Warning', 'Please select an action type')	
			return
		if this_action == None:
			this_action = self.current_action_name.get()
		if this_action == '':
			messagebox.showwarning('Warning', 'Please select an action name')	
			return
		arg_dict = self.action_dict[this_type][this_action]
		self.Current_Arg_Value = []
		self.Current_Arg_Type = None
		if arg_dict == None:
			self.btn_add_step.configure(state=NORMAL)
			return
		
		child_windows = Toplevel(self.parent)
		#child_windows.geometry("200x200")  # Size of the window 
		child_windows.resizable(False, False)
		child_windows.title("Input the action argument")
		StringVar()
		
		arg_name = [*arg_dict]
		row = 1
		value_array = []
		button_array = []
		arg_data_list = []
		if len(arg_name) > 0:
			for arg in arg_name:
				arg_data = {}
				arg_data['variable_type'] = None
				arg_data['value_variable'] = None
				arg_data['widget'] = None
				arg_data['button'] = None

				Label(child_windows, text= arg, anchor=S, justify=LEFT).grid(row=row,column=1, padx=10, pady=10, sticky=S)
				#value_array[row] = StringVar()
				temp_Variable_type = arg_dict[arg]
				if temp_Variable_type == None:
					arg_data['variable_type'] = None
					Label(child_windows, text= 'N/A').grid(row=row,column=2, padx=10, pady=10, sticky=S)
				elif temp_Variable_type.startswith('_'):
					# If variable type has the '_' before the type name, that arg cannot leave as empty.
					arg_data['variable_type'] = temp_Variable_type[1:]
					arg_data['input_require'] = True
				else:
					arg_data['variable_type'] = temp_Variable_type
					arg_data['input_require'] = False
				
				if arg_data['variable_type'] in ['string', 'int', 'float']:
					arg_data['widget'] = Text(child_windows, height = 1, width=30)
					arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=S)
				elif arg_data['variable_type'] == 'string_id':
					arg_data['value_variable'] = StringVar()
					value_list = [*self.AutoTester.UI]
					value_list.sort()
					if len(value_list) > 0:
						arg_data['value_variable'].set(value_list[0])
						# creating widget
						arg_data['widget'] = AutocompleteCombobox(child_windows)
						arg_data['widget'].set_completion_list(value_list)
						arg_data['widget'].Set_Entry_Width(35)
						arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=W+S)
					else:
						continue	
				elif arg_data['variable_type'] == 'user_list':
					#self.Execute_List_values
					arg_data['value_variable'] = StringVar()
					value_list = ['Varlue 1', 'Varlue 2', 'Varlue 3', 'Varlue 4', 'Varlue 5', 'Varlue 6']
					arg_data['value_variable'].set(value_list[0])
					# creating widget
					arg_data['widget'] = AutocompleteCombobox(child_windows)
					arg_data['widget'].set_completion_list(value_list)
					arg_data['widget'].Set_Entry_Width(35)
					arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=W+S)
				elif arg_data['variable_type'] == 'point':
					arg_data['widget'] = Text(child_windows, height = 1, width=30)
					arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=W)
					if self.AutoTester.Device != None:
						btn_status = NORMAL
					else:
						btn_status = DISABLED
					arg_data['button'] = Button(child_windows, text = 'Select', command = lambda val=arg_data['widget']: self.Btn_Select_Point(val), state=btn_status)
					arg_data['button'].grid(row=row,column=3, padx=10, pady=10, sticky=E)

				elif arg_data['variable_type'] == 'area':
					arg_data['widget'] = Text(child_windows, height = 1, width=30)
					arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=S)
					if self.AutoTester.Device != None:
						btn_status = NORMAL
					else:
						btn_status = DISABLED
					arg_data['button'] = Button(child_windows, text = 'Select', command = lambda val=arg_data['widget']: self.Btn_Select_Area(val), state=btn_status)
					arg_data['button'].grid(row=row,column=3, padx=10, pady=10, sticky=E)
				
				elif arg_data['variable_type'] == 'current_area':
					arg_data['widget'] = Text(child_windows, height = 1, width=30)
					arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=S)
					if self.AutoTester.Device != None:
						btn_status = NORMAL
					else:
						btn_status = DISABLED
					arg_data['button'] = Button(child_windows, text = 'Select', command = lambda val=arg_data['widget']: self.Btn_Crop_Area(val), state=btn_status)
					arg_data['button'].grid(row=row,column=3, padx=10, pady=10, sticky=E)
				
				else:
					Label(child_windows, text= 'N/A').grid(row=row,column=2, padx=10, pady=10, sticky=S)
				arg_data_list.append(arg_data)
				row +=1
		child_windows.protocol("WM_DELETE_WINDOW", lambda c=child_windows, a=arg_data_list: self.Store_Input_Value_On_Closing(c,a))		
		Button(child_windows, text = 'Set value', command = lambda c=child_windows, a=arg_data_list: self.Store_Input_Value_On_Closing(c,a)).grid(row=row,column=2, padx=10, pady=10, sticky=E)


	def Generate_Arg_Edit_Window(self, treeview_node, this_type = None, this_action = None):
		if this_type == None:
			this_type = self.current_action_type.get()
		if this_type == '':
			messagebox.showwarning('Warning', 'Please select an action type')	
			return
		if this_action == None:
			this_action = self.current_action_name.get()
		if this_action == '':
			messagebox.showinfo('Error...', 'Please select an action name')	
			return
		arg_dict = self.action_dict[this_type][this_action]
		if arg_dict == None:
			self.Modify_Input_Value_On_Closing(None,[], this_type, this_action, treeview_node)
			return

		child_windows = Toplevel(self.parent)
		#child_windows.geometry("200x200")  # Size of the window 
		child_windows.resizable(False, False)
		child_windows.title("Input the action argument")
		
		arg_name = [*arg_dict]
		row = 1
		value_array = []
		button_array = []
		arg_data_list = []
		child = self.Treeview.item(treeview_node)
		value = child['values']
		if len(arg_name) > 0:
			value_index=  2
			for arg in arg_name:
				arg_data = {}
				arg_data['variable_type'] = None
				arg_data['value_variable'] = None
				arg_data['widget'] = None
				arg_data['button'] = None

				Label(child_windows, text= arg, anchor=S, justify=LEFT).grid(row=row,column=1, padx=10, pady=10, sticky=S)
				#value_array[row] = StringVar()
				temp_Variable_type = arg_dict[arg]
				if temp_Variable_type == None:
					arg_data['variable_type'] = None
					Label(child_windows, text= 'N/A').grid(row=row,column=2, padx=10, pady=10, sticky=S)
				elif temp_Variable_type.startswith('_'):
					# If variable type has the '_' before the type name, that arg cannot leave as empty.
					arg_data['variable_type'] = temp_Variable_type[1:]
					arg_data['input_require'] = True
				else:
					arg_data['variable_type'] = temp_Variable_type
					arg_data['input_require'] = False
				
				if arg_data['variable_type'] in ['string', 'int', 'float']:
					arg_data['widget'] = Text(child_windows, height = 1, width=30)
					arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=S)
					try:	
						arg_data['widget'].insert("end", value[value_index])
					except Exception as e:
						print('Error while adding string value:', e)
						pass	

				elif arg_data['variable_type'] == 'string_id':
					arg_data['value_variable'] = StringVar()
					value_list = [*self.AutoTester.UI]
					value_list.sort()
					arg_data['value_variable'].set(value_list[0])
					# creating widget
					arg_data['widget'] = AutocompleteCombobox(child_windows)
					arg_data['widget'].set_completion_list(value_list)
					arg_data['widget'].Set_Entry_Width(35)
					arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=W+S)
					try:	
						arg_data['widget'].set(value[value_index])
					except Exception as e:
						print('Error while adding list value:', e)
						pass
					
				elif arg_data['variable_type'] == 'user_list':
					#self.Execute_List_values
					arg_data['value_variable'] = StringVar()
					value_list = ['Varlue 1', 'Varlue 2', 'Varlue 3', 'Varlue 4', 'Varlue 5', 'Varlue 6']
					arg_data['value_variable'].set(value_list[0])
					# creating widget
					arg_data['widget'] = AutocompleteCombobox(child_windows)
					arg_data['widget'].set_completion_list(value_list)
					arg_data['widget'].Set_Entry_Width(35)
					arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=S)
					try:
						
						arg_data['widget'].set(value_list[value[value_index]])
					except Exception as e:
						print('Error while adding user list value:', e)
						pass
				elif arg_data['variable_type'] == 'point':
					arg_data['widget'] = Text(child_windows, height = 1, width=30)
					arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=W+S)
					if self.AutoTester.Device != None:
						btn_status = NORMAL
					else:
						btn_status = DISABLED
					arg_data['button'] = Button(child_windows, text = 'Select', command = lambda val=arg_data['widget']: self.Btn_Select_Point(val), state=btn_status)
					arg_data['button'].grid(row=row,column=3, padx=10, pady=10, sticky=E)
					try:
						arg_data['widget'].insert("end", value[value_index])
					except Exception as e:
						print('Error while adding point value:', e)
						pass
				elif arg_data['variable_type'] == 'area':
					arg_data['widget'] = Text(child_windows, height = 1, width=30)
					arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=S)
					if self.AutoTester.Device != None:
						btn_status = NORMAL
					else:
						btn_status = DISABLED
					arg_data['button'] = Button(child_windows, text = 'Select', command = lambda val=arg_data['widget']: self.Btn_Select_Area(val), state=btn_status)
					arg_data['button'].grid(row=row,column=3, padx=10, pady=10, sticky=E)
					try:
						arg_data['widget'].insert("end", value[value_index])
					except Exception as e:
						print('Error while adding area value:', e)
						pass

				elif arg_data['variable_type'] == 'current_area':
					arg_data['widget'] = Text(child_windows, height = 1, width=30)
					arg_data['widget'].grid(row=row,column=2, padx=10, pady=10, sticky=S)
					if self.AutoTester.Device != None:
						btn_status = NORMAL
					else:
						btn_status = DISABLED
					arg_data['button'] = Button(child_windows, text = 'Select', command = lambda val=arg_data['widget']: self.Btn_Crop_Area(val), state=btn_status)
					arg_data['button'].grid(row=row,column=3, padx=10, pady=10, sticky=E)
					try:
						arg_data['widget'].insert("end", value[value_index])
					except Exception as e:
						print('Error while adding custom area value:', e)
						pass

				else:
					Label(child_windows, text= 'N/A').grid(row=row,column=2, padx=10, pady=10, sticky=S)
				arg_data_list.append(arg_data)
				row +=1
				value_index+=1
		child_windows.protocol("WM_DELETE_WINDOW", lambda c=child_windows,x= this_type, y= this_action,i=treeview_node, a=arg_data_list: self.Modify_Input_Value_On_Closing(c,a,x, y, i))		
		Button(child_windows, text = 'Set value', command = lambda c=child_windows,x= this_type, y= this_action,i=treeview_node, a=arg_data_list: self.Modify_Input_Value_On_Closing(c,a,x, y, i)).grid(row=row,column=2, padx=10, pady=10, sticky=E)

	def Btn_Select_Area(self, text_widget):
		im = self.AutoTester.Get_Screenshot_In_Working_Resolution(self.WorkingResolution)
		show_im, ratio = self.Resize_Image_by_ratio(im)
		#image = cv2.imdecode(np.frombuffer(im, np.uint8), cv2.IMREAD_COLOR)
		location = cv2.selectROI("Sekect scan area", show_im, showCrosshair=False,fromCenter=False)
		area = {}
		area['x'] = int(location[0]/ratio)
		area['y'] = int(location[1]/ratio)
		area['w'] = int(location[2]/ratio)
		area['h'] = int(location[3]/ratio)
		cv2.destroyAllWindows()
		text_widget.delete('1.0', END)	
		text_widget.insert("end", json.dumps(area))

	def Btn_Crop_Area(self, text_widget):
		im = self.AutoTester.Get_Screenshot_In_Working_Resolution(self.WorkingResolution)
		show_im, ratio = self.Resize_Image_by_ratio(im)
		#im = cv2.cvtColor(im , cv2.COLOR_BGR2RGB)
		#im = cv2.imdecode(np.frombuffer(im, np.uint8), cv2.IMREAD_COLOR)
		area = cv2.selectROI("Sekect scan area", show_im, showCrosshair=False,fromCenter=False)
		imCrop = im[int(area[1]/ratio):int(area[1]/ratio + area[3]/ratio), int(area[0]/ratio):int(area[0]/ratio + area[2]/ratio)]
		
		cv2.destroyAllWindows()
		#cv2.imshow(' ', imCrop)
		#cv2.waitKey(0)
		out_dir = self.DB_Folder + '\\temp_img'
		Init_Folder(out_dir)
		_name = out_dir   + '\\img_' + str(Function_Get_TimeStamp()) + '.png'	
		cv2.imwrite(_name, imCrop)
		text_widget.delete('1.0', END)	
		text_widget.insert("end", _name)	

	def Btn_Select_Point(self, text_widget):
		self.temp_widget = text_widget
		im = self.AutoTester.Get_Screenshot_In_Working_Resolution(self.WorkingResolution)
		show_im, ratio = self.Resize_Image_by_ratio(im)
		#image = cv2.imdecode(np.frombuffer(im, np.uint8), cv2.IMREAD_COLOR)
		cv2.imshow('Screen', show_im)
		cv2.setMouseCallback('Screen', self.click_event)
		cv2.waitKey(0)
		cv2.destroyAllWindows()
		
		point = dict()
		point['x'] = int(self.temp_x / ratio)
		point['y'] = int(self.temp_y / ratio)
		self.temp_x = None
		self.temp_y = None
		self.temp_widget.delete('1.0', END)	
		self.temp_widget.insert("end", json.dumps(point)) 
		self.temp_widget = None

	def click_event(self, event, x, y, flags, params):
		
		if self.temp_widget != None:
			# checking for left mouse clicks
			if event == cv2.EVENT_LBUTTONDOWN:
				# displaying the coordinates
				# on the Shell
				self.temp_widget.delete('1.0', END)	
				self.temp_x = x
				self.temp_y = y
				cv2.destroyAllWindows()
				

	def Resize_Image_by_ratio(self, img):
		global WIDTH, HEIGHT
		(_h, _w) = img.shape[:2]
		_ratio = 1
		while True:
			temp_w = _w * _ratio
			temp_h = _h * _ratio
			if temp_h > HEIGHT*0.75 or temp_w > WIDTH*0.75:
				_ratio = int(_ratio * 90)/100
			else:
				break
		
		if _ratio != 1:
			width = int(img.shape[1] * _ratio)
			height = int(img.shape[0] * _ratio)
			dim = (width, height)
			img = cv2.resize(img, dim, interpolation = cv2.INTER_AREA)
			#(_h, _w) = img.shape[:2]
		'''
		_resolution_index = self.Resolution.get()
		if _resolution_index == 1:
			_WorkingResolution = 720
		else:
			_WorkingResolution = 1080
		
		print(_h, _w, _WorkingResolution)

		if _w > _h:
			actual_ratio = _WorkingResolution / _h
		else:
			actual_ratio = _WorkingResolution / _w
		print(actual_ratio)	
		actual_ratio = actual_ratio * _ratio
		'''
		
		actual_ratio = _ratio
		print(actual_ratio)
		return img, actual_ratio

	def Store_Input_Value_On_Closing(self,child_windows, arg_data_list):
		this_type = self.current_action_type.get()
		this_action = self.current_action_name.get()
	
		input_value = []
		for arg_data in arg_data_list:
			temp_value = None
			if arg_data['variable_type'] == None:
				temp_value = ''
			elif arg_data['variable_type'] in ['string', 'int', 'float']:
				temp_value = arg_data['widget'].get("1.0", END).replace('\n', '')
				
			elif arg_data['variable_type'] == 'string_id':
				temp_value = arg_data['widget'].get()
			elif arg_data['variable_type'] == 'point':
				temp_value = arg_data['widget'].get("1.0", END).replace('\n', '')
			elif arg_data['variable_type'] == 'user_list':
					value_list = ['Varlue 1', 'Varlue 2', 'Varlue 3', 'Varlue 4', 'Varlue 5', 'Varlue 6']
					selected = arg_data['widget'].get()
					temp_value = value_list.index(selected)
					
					
			elif arg_data['variable_type'] == 'area':
				temp_value = arg_data['widget'].get("1.0", END).replace('\n', '')
			elif arg_data['variable_type'] == 'current_area':
				temp_value = arg_data['widget'].get("1.0", END).replace('\n', '')	
			elif arg_data['variable_type'] == 'list':
				temp_value = []
				
			else:
				input_value.append("")
			input_value.append(temp_value)
		values = [this_type] + [this_action] + input_value

		self.Current_Arg_Value = values
		self.Current_Arg_Type = this_type
		self.btn_add_step.configure(state=NORMAL)
		if child_windows != None:
			child_windows.destroy()
		#self.btn_add_step.configure(state=NORMAL)

	def Modify_Input_Value_On_Closing(self,child_windows, arg_data_list, this_type= None, this_action= None, treeview_node= None):
		if this_type == None:
			this_type = self.current_action_type.get()
		if this_action == None:	
			this_action = self.current_action_name.get()
	
		input_value = []
		for arg_data in arg_data_list:
			temp_value = None
			if arg_data['variable_type'] == None:
				temp_value = ''
			elif arg_data['variable_type'] in ['string', 'int', 'float']:
				temp_value = arg_data['widget'].get("1.0", END).replace('\n', '')
				
			elif arg_data['variable_type'] == 'string_id':
				temp_value = arg_data['widget'].get()
			elif arg_data['variable_type'] == 'point':
				temp_value = arg_data['widget'].get("1.0", END).replace('\n', '')
			elif arg_data['variable_type'] == 'user_list':
					value_list = ['Varlue 1', 'Varlue 2', 'Varlue 3', 'Varlue 4', 'Varlue 5', 'Varlue 6']
					selected = arg_data['widget'].get()
					temp_value = value_list.index(selected)
					
					
			elif arg_data['variable_type'] == 'area':
				temp_value = arg_data['widget'].get("1.0", END).replace('\n', '')
			elif arg_data['variable_type'] == 'current_area':
				temp_value = arg_data['widget'].get("1.0", END).replace('\n', '')
			elif arg_data['variable_type'] == 'list':
				temp_value = []
				
			else:
				input_value.append("")
			input_value.append(temp_value)
		values = [this_type] + [this_action] + input_value

		self.Treeview.item(treeview_node, text= '', values=values)

		if child_windows != None:
			child_windows.destroy()	

	def get_location_on_click(self):
		img = self.AutoTester.Get_Current_Screenshot()
		cv2.imshow('Game screen', img)
		# setting mouse handler for the image
		# and calling the click_event() function
		cv2.setMouseCallback('image', self.click_event)
	
		# wait for a key to be pressed to exit
		cv2.waitKey(0)
		# close the window
		cv2.destroyAllWindows()

	# Other function
	
	def Update_Action_List(self):

		print('Update action list to drop list')
		action_type = ['Action', 'Loop', 'Condition', 'Get_Result', 'Update_Variable', 'Comment']
		action_type.sort()
		self.action_dict = {}
		for type in action_type:
			self.action_dict[type] = {}
		for action in self.AutoTester.action_list:
			if action['type'] in action_type:
				self.action_dict[action['type']][action['name']] = action['arg']

	def ismodulefunction(self, module, member):
			module_member = getattr(module, member)
			member_type = type(module_member).__name__
			return member_type == "function" or member_type == "builtin_function_or_method"

	def Browse_Module(self):
		#need to run again in main Automation lib
		#from inspect import getmembers, isfunction
		print('Enter lib path:')
		path = r'C:\Users\Evan\OneDrive\Documents\GitHub\Evn-Automation-tool\libs\New Python File.py'
		
		loader = importlib.machinery.SourceFileLoader( 'mymodule', path)
		spec = importlib.util.spec_from_loader( 'mymodule', loader )
		mymodule = importlib.util.module_from_spec( spec )
		loader.exec_module( mymodule)
		all_f = dir(mymodule)
		new_f = []
		for mymodule_name in all_f:
			if not mymodule_name.startswith('__'):
				new_f.append(mymodule_name)
		for function in new_f:
			real_f = getattr(mymodule, function)
			a = real_f()
			setattr(self.AutoTester, real_f.__name__, getattr(a, 'func'))
			self.AutoTester.append_action_list(type = 'Action', name = real_f.__name__ , argument = a.kwarg, description= '')

		print('Update action list to drop list')
		action_type = ['Action', 'Loop', 'Condition', 'Get_Result', 'Update_Variable']
		action_type.sort()
		self.action_dict = {}
		for type in action_type:
			self.action_dict[type] = {}
		for action in self.AutoTester.action_list:
			if action['type'] in action_type:
				self.action_dict[action['type']][action['name']] = action['arg']	
				
			
	def Update_Action_Name(self, event=None):
		menu = self.action_name["menu"]
		menu.delete(0, 'end')
		this_type = self.current_action_type.get()
		action_list = []
		if this_type != '':
			temp_action_list = self.action_dict[this_type]
			for action in temp_action_list:
				action_list.append(action)
		
		action_list.sort()
		
		for value in action_list:
			menu.add_command(label=value, command=lambda _value=value: [lambda v=_value: self.current_action_name.set(_value), self.Update_Action_Arg(_value)])
		if len(action_list)> 0:
			self.current_action_name.set(action_list[0])
		else:	
			self.current_action_name.set("")	
		self.btn_add_step.configure(state=DISABLED)
		
	def ImportTestCase(self, Test_Case_File_Path):
		print('Loading My Dictionary')
		if Test_Case_File_Path != None:
			if (os.path.isfile(Test_Case_File_Path)):
				xlsx = load_workbook(Test_Case_File_Path)
				DictList = []
				Dict = []
				for sheet in xlsx:
					sheetname = sheet.title.lower()
					if sheetname not in self.SpecialSheets:	
						EN_Coll = ""
						KR_Coll = ""
						database = None
						ws = xlsx[sheet.title]
						for row in ws.iter_rows():
							for cell in row:
								if cell.value == "KO":
									KR_Coll = cell.column_letter
									KR_Row = cell.row
									database = ws
								elif cell.value == "EN":
									EN_Coll = cell.column_letter
								if KR_Coll != "" and EN_Coll != "":
									DictList .append(sheet.title)
									break	
							if database!=  None:
								break		
						if database != None:
							for i in range(KR_Row, database.max_row): 
								KRAddress = KR_Coll + str(i+1)
								ENAddress = EN_Coll + str(i+1)
								#print('KRAddress', KRAddress)
								#print('ENAddress', ENAddress)
								KRCell = database[KRAddress]
								KRValue = KRCell.value
								ENCell = database[ENAddress]
								ENValue = ENCell.value
								if KRValue == None or ENValue == None or KRValue == 'KO' or ENValue == 'EN':
									continue
								elif KRValue != None and ENValue != None:
									Dict.append([KRValue, ENValue.lower()])
				print("Successfully load dictionary from: ", DictList)
				return Dict
			else:
				return([])	
		else:
			return([])	

	# Menu Function
	def Menu_Function_About(self):
		messagebox.showinfo("About....", "Creator: Evan")

	def Show_Error_Message(self, ErrorText):
		messagebox.showinfo('Error...', ErrorText)	

	def SaveAppLanguage(self, language):
		self.Write_Debug(self.LanguagePack.ToolTips['AppLanuageUpdate'] + " "+ language) 
		self.AppConfig.Save_Config(self.AppConfig.Auto_Tool_Config_Path, 'AUTO_TOOL', 'app_lang', language)

	def SetLanguageKorean(self):
		self.AppLanguage = 'kr'
		self.SaveAppLanguage(self.AppLanguage)
		#self.initUI()
	
	def SetLanguageEnglish(self):
		self.AppLanguage = 'en'
		self.SaveAppLanguage(self.AppLanguage)
		#self.initUI()

	def Menu_Function_Open_Main_Guideline(self):
		#webbrowser.open_new(r"https://confluence.nexon.com/pages/viewpage.action?pageId=298119695")
		print('Done')

	def Menu_Function_Save_TC_Template(self):
		filename = filedialog.asksaveasfilename(title = "Select file", filetypes = (("Scan Config", "*.csv"),),)
		filename = self.CorrectExt(filename, "csv")
		if filename == "":
			messagebox.showinfo('Error...', 'Please input file name to create the template.')	
			return
		else:
			try:
				with open(filename, 'w', newline='') as csvfile:

					fieldnames = ['Type', 'Action', 'Arg1', 'Arg2', 'Arg3', 'Arg4', 'Arg5', 'Arg6']
				
					writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
					writer.writeheader()

			except Exception as e:
				messagebox.showerror("Error", str(e))			
			
	
	def Menu_Function_Save_DB_Template(self):
		filename = filedialog.asksaveasfilename(title = "Select file", filetypes = (("Scan Config", "*.csv"),),)
		filename = self.CorrectExt(filename, "csv")
		if filename == "":
			messagebox.showinfo('Error...', 'Please input file name to create the template.')	
			return
		else:
			try:
				with open(filename, 'w', newline='') as csvfile:

					fieldnames = ['StringID', 'String_EN', 'String_KO', 'Path']
				
					writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
					writer.writeheader()
			except Exception as e:
				messagebox.showerror("Error", str(e))		
	
	def Menu_Function_Save_Execute_List_Template(self):
		filename = filedialog.asksaveasfilename(title = "Select file", filetypes = (("Scan Config", "*.csv"),),)
		filename = self.CorrectExt(filename, "csv")
		if filename == "":
			messagebox.showinfo('Error...', 'Please input file name to create the template.')	
			return
		else:
			try:
				with open(filename, 'w', newline='') as csvfile:

					fieldnames = self.Execute_List_header
				
					writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
					writer.writeheader()		
					writer.writerow({'Name': 'All supported type', 'Value_1': '1', 'Type_1': 'int', 'Value_2': 'Any text', 'Type_2': "string", 'Value_3': '0.01', 'Type_3': 'float', 'Value_4': '{"x":100, "y": 200}', 'Type_4': 'point','Value_5': '{"x":100, "y": 200, "w": 50, "h":75}', 'Type_5': 'area','Value_6': 'String'})

			except Exception as e:
				messagebox.showerror("Error", str(e))		

	def ErrorMsg(self, ErrorText):
		messagebox.showinfo('Error...', ErrorText)	

	def OpenOutput(self):
		Source = self.ListFile[0]
		Outputdir = os.path.dirname(Source)
		BasePath = str(os.path.abspath(Outputdir))
		subprocess.Popen('explorer ' + BasePath)

	
	def BtnSelectColour(self):
		colorStr, self.BackgroundColor = colorchooser.askcolor(parent=self, title='Select Colour')
		if self.BackgroundColor == None:
			self.ErrorMsg('Set colour as defalt colour (Yellow)')
			self.BackgroundColor = 'ffff00'
		else:
			self.BackgroundColor = self.BackgroundColor.replace('#', '')

		return

	def on_closing(self):
		self.parent.destroy()
		try:
			if self.Automation_Processor.is_alive():	
				self.Automation_Processor.terminate()
		except:
			pass
		self.quit()


	def init_App_Setting(self):
		
		self.Text_DB_Path = StringVar()
		self.Str_Test_Case_Path = StringVar()

		self.TesseractPath = StringVar()
		self.TesseractDataPath = StringVar()
		self.WorkingLanguage = StringVar()
		self.Text_CustomAction_Path = StringVar()

		self.language_list = ['']
		self.Current_Arg_Type = []
		self.Current_Arg_Type = None
		self.Resolution = IntVar()

		self.DBPath = StringVar()

		self.CurrentDataSource = StringVar()

		self.ScanType = StringVar()

		self.Notice = StringVar()

		self.current_action_type = StringVar()
		self.current_action_name = StringVar()
	

		self.AppConfig = ConfigLoader('AUTO_TOOL')

		self.Configuration = self.AppConfig.Config
		self.AppLanguage  = self.Configuration['AUTO_TOOL']['app_lang']

		_tesseract_path = self.Configuration['AUTO_TOOL']['tess_path']
		pytesseract.pytesseract.tesseract_cmd = str(_tesseract_path)
		self.TesseractPath.set(_tesseract_path)

		_tesseract_data_path = self.Configuration['AUTO_TOOL']['tess_data']
		self.TesseractDataPath.set(_tesseract_data_path)

		_db_path = self.Configuration['AUTO_TOOL']['db_path']
		
		_custom_action_path = self.Configuration['AUTO_TOOL']['custom_action']
		self.Text_CustomAction_Path.set(_custom_action_path)

		self.DB_Folder = os.path.dirname(_db_path)
		print('DB folder:', self.DB_Folder)
		Init_Folder(self.DB_Folder)
		self.DB_Path = _db_path
		if not os.path.isfile(_db_path):
			with open(_db_path, 'w', newline='') as csvfile:
				fieldnames = ['StringID', 'String_EN', 'String_KO', 'Path']
				writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
				writer.writeheader()
			self.AppConfig.Save_Config(self.AppConfig.Auto_Tool_Config_Path, 'AUTO_TOOL', 'db_path', _db_path, True)
			self.Text_DB_Path.set(_db_path)
			self.DB_Path = _db_path
		
		self.Text_DB_Path.set(_db_path)
		self.Execute_List_header = ['Name', 'Value_1', 'Type_1', 'Value_2', 'Type_2', 'Value_3', 'Type_3', 'Value_4', 'Type_4', 'Value_5', 'Type_5', 'Value_6', 'Type_6']
		self.Execute_List_values = None

	def init_UI_Data(self):
		
		self.Btn_OCR_Update_Working_Language()
		_working_language = self.Configuration['AUTO_TOOL']['scan_lang']
		self.option_working_language.set(_working_language)

		self.AutoTester.Update_DB_Path(self.DB_Path)

		if self.TesseractPath != None and self.TesseractDataPath != None and _working_language != None:
			print('Enable OCR')
			self.AutoTester.Update_Tesseract(self.TesseractPath, self.TesseractDataPath, _working_language)

		_resolution = self.Configuration['AUTO_TOOL']['resolution']
		self.Resolution.set(_resolution)
		if _resolution == 1:
			self.WorkingResolution = 720
		else:
			self.WorkingResolution = 1080 
		self.AutoTester.Update_Resolution(self.WorkingResolution)

		self.Update_Action_List()
		self.Get_Serial()
	
	def Btn_Select_Font_Colour(self):
		colorStr, self.Font_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.Font_Color == None:
			self.Error('Set colour as defalt colour (Yellow)')
			self.Font_Color = 'FF0000'
		else:
			self.Font_Color = self.Font_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return

	def Btn_Browse_Module_File(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Template files", "*.jpg *.png"), ), multiple = False)	
		if filename != "":
			self.Template_Path = self.CorrectPath(filename)
			self.Str_Template_Path.set(filename)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_Browse_Template_File(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Template files", "*.jpg *.png"), ), multiple = False)	
		if filename != "":
			self.Template_Path = self.CorrectPath(filename)
			self.Str_Template_Path.set(filename)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_Browse_Test_Case_File(self,event=None):
		config_file = filedialog.askopenfilename(initialdir = CWD + "//Testcase", title =  self.LanguagePack.ToolTips['SelectSource'], filetypes = (("Config files", "*.csv"), ), multiple = False)	
		
		if os.path.isfile(config_file):
			self.Str_Test_Case_Path.set(config_file)
			self.remove_treeview()

			all_col = ['Type', 'Action', 'Arg1', 'Arg2', 'Arg3', 'Arg4', 'Arg5', 'Arg6']
			
			with open(config_file, newline='', encoding='utf-8-sig') as csvfile:
				reader = csv.DictReader(csvfile)
				
				for location in reader:
					values = []
					for col in all_col:
						if col in location:
							value = str(location[col])
						else:
							value = ""
						values.append(value)

					if len(values) > 0:
						this_tag = values[0].lower()
						self.Treeview.insert('', 'end', text= '', values=values, tags= this_tag)
					
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])

	def Btn_Browse_Execution_List(self):
			
		execution_list_file = filedialog.askopenfilename(initialdir = CWD + "//Testcase", title =  self.LanguagePack.ToolTips['SelectSource'], filetypes = (("Config files", "*.csv"), ), multiple = False)	
		
		if os.path.isfile(execution_list_file):
			all_col = self.Execute_List_header
			self.Display_List = []
			self.Execute_List_values = []

			with open(execution_list_file, newline='', encoding='utf-8-sig') as csvfile:
				reader = csv.DictReader(csvfile)
				_indexer = 0
				for execution_item in reader:
					_indexer+=1
					item_list = []
					if all_col[0] in execution_item:
						self.Display_List.append('[' + str(_indexer) + '] ' + execution_item[all_col[0]])
					for i in range(1, 7):
						_col_val = 'Value_'+ str(i)
						_col_val_type = 'Type_'+ str(i)
						
						if _col_val in execution_item and _col_val_type in execution_item:
							_val = execution_item[_col_val]
							_val_type = execution_item[_col_val_type]
							if _val == '':
								continue
							if _val_type == '':
								_val_type = 'string'
							parsed_value = self.AutoTester.Function_Parse_Data(_val_type, _val)
							item_list.append(parsed_value)	

					self.Execute_List_values.append(item_list)	
			#print('self.Execute_List_values',self.Execute_List_values)
			self.ExecuteList.set_completion_list(self.Display_List)
			self.ExecuteList.current(0)
			self.ExecuteType.configure(state=NORMAL)
			self.AutoTester.Update_LoopList()
			self.Update_Action_List()
		else:
			self.Write_Debug(self.LanguagePack.ToolTips['SourceDocumentEmpty'])

	def Btn_Save_Test_Case_File(self, event=None):
		'''
		Save all added scan areas to csv file.
		'''
		filename = filedialog.asksaveasfilename(title = "Select file", filetypes = (("Scan Config", "*.csv"),),)
		filename = self.CorrectExt(filename, "csv")
		if filename == "":
			return
		else:
			with open(filename, 'w', newline='') as csvfile:

				fieldnames = ['Type', 'Action', 'Arg1', 'Arg2', 'Arg3', 'Arg4', 'Arg5', 'Arg6']
			
				writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
				writer.writeheader()
				for row in self.Treeview.get_children():
					child = self.Treeview.item(row)
					temp_values = child["values"]
					values = []
					for i in range(len(fieldnames)):
						if len(temp_values) < i+1:
							values.append('')
						else:
							values.append(temp_values[i])

					writer.writerow({'Type': values[0], 'Action': values[1], 'Arg1': values[2], 'Arg2': values[3], 'Arg3': values[4], 'Arg4': values[5], 'Arg5': values[6], 'Arg6': values[7]})

	def Create_Test_Object_List(self):
		Test_Obect_List = []
		for row in self.Treeview.get_children():
			_test_object = {}
			child = self.Treeview.item(row)
			_action_type = child['values'][0]
			_test_object['type'] = _action_type
			_action_name = child['values'][1]
			_test_object['name'] = _action_name
			'''
			if _action_name in ['End Loop', 'End If']:
				_action_name = _action_name.replace('End ', '')
			'''
			if _action_name in ['End Loop', 'End If']:
				_arg = []
			else:		
				_arg_info = self.action_dict[_action_type][_action_name]
				_arg = []
				if _arg_info != None:
					
					_index = 2
					for arg in _arg_info:
						print('arg', arg)
						_current_arg = {}
						if len(child['values']) > _index:
							_val = child['values'][_index]
							if isinstance(_val, str):
								_val = _val.replace('\n', '')
							if _val != '':
								_current_arg['name'] = arg
								_current_arg['type'] = _arg_info[arg]
								_current_arg['value'] = _val
								_arg.append(_current_arg)
						_index+=1		
			_test_object['arg'] = _arg
			print('_test_object', _test_object)	
			Test_Obect_List.append(_test_object)	
		return Test_Obect_List

	def Btn_Execute_Script(self):
	
		#self.Btn_Generate_TestCase()

		DB_Path = self.Text_DB_Path.get()
		_Execute_Type = self.ExecuteType.get()
		if _Execute_Type != "":
			if self.Execute_List_values == None:
				messagebox.showinfo('Error...', 'No device found')
				return
		if _Execute_Type == 'Current Value':
			_Execute_Name = self.ExecuteList.get()
			_indexer = self.Display_List.index(_Execute_Name)
			_Execute_Value = [self.Execute_List_values[_indexer]]
		else:
			# Loop all list
			
			_Execute_Value = self.Execute_List_values

		
		Serial = self.TextSerial.get().replace('\n','')
		#MyDB = self.Function_Import_DB(self.DB_Path)
		try:
			self.Automation_Processor.terminate()
		except Exception as e:
			pass
		
		Test_Case_Path = self.Str_Test_Case_Path.get()
		Name = 'General_Test_Case'
		if Test_Case_Path == "":
			Result_Folder_Path = CWD + '\\Result' + '_' + 'General_Test_Case' + '_' + Function_Get_TimeStamp()
		else:
			Dir, Name, Ext = Split_Path(Test_Case_Path)
			Result_Folder_Path = Dir + '\\Result' + '_' + Name + '_' + Function_Get_TimeStamp()
		

		Init_Folder(Result_Folder_Path)
		
		Result_File_Path = Result_Folder_Path + '\\' + Name + '_' + Function_Get_TimeStamp()
		
		_serial = self.TextSerial.get()
		Test_Object =  self.Create_Test_Object_List()

		_tess_language = self.option_working_language.get()
		self.Auto_Setting_Set_Working_Language(_tess_language)
	
		self.Auto_Setting_Set_Working_Resolution()
		
		if len(Test_Object) < 1:
			messagebox.showinfo('Error...', 'Please add at least 1 action before running the application')
			return
	 	
		_kwargs= {	'Status_Queue' : self.Status_Queue, 
					'Result_Queue' : self.Result_Queue, 
					'Progress_Queue': self.Process_Queue,
					'Serial' : _serial, 
					'Resolution': self.WorkingResolution,
					'Language': _tess_language,
					'Tess_Path' : self.TesseractPath.get(), 
					'Tess_Data' : self.TesseractDataPath.get(),
					'DB_Path' : DB_Path,
					'Test_Case_Path' : Test_Case_Path, 
					'Result_Folder_Path' : Result_Folder_Path,
					'Result_File_Path': Result_File_Path,
					'TestCaseObject' : Test_Object,
					'Execute_List' : _Execute_Value, 
				}	
		self.Automation_Processor = Process(target = Function_Execute_Script, kwargs= _kwargs,)
		self.Automation_Processor.start()
		self.after(DELAY1, self.Wait_For_Automation_Processor)	

	def status_listening(self):
		while True:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Debugger.insert("end", "\n\r")
					ct = datetime.now()
					self.Debugger.insert("end", str(ct) + ": " + Status)
					self.Debugger.yview(END)
			except queue.Empty:
				break
		device_status = self.AutoTester.Check_Connectivity()
		if device_status != False:
			device_status = True
			self.after(DELAY2, self.status_listening)
		else:
			self.after(DELAY3, self.status_listening)
		#print('Device status:', device_status, time.time()- Start)
		

	def Wait_For_Automation_Processor(self):
		if (self.Automation_Processor.is_alive()):	
			self.after(DELAY1, self.Wait_For_Automation_Processor)
		else:
			self.Automation_Processor.terminate()
			#self.Show_Error_Message('Test is completed')	

###########################################################################################

	def check_device_connection(self):
		process = subprocess.Popen(ADBPATH + ' devices', stdout=subprocess.PIPE, stderr=None, shell=True)
		device = process.communicate()[0].decode("utf-8") 
		print(device)
		return device
	

###########################################################################################

def Function_Execute_Script(
		Status_Queue, Progress_Queue, Result_Queue, Serial, DB_Path, Tess_Path, Tess_Data, Resolution, Language, Result_Folder_Path, Result_File_Path, TestCaseObject = [], Execute_List = None, **kwargs
):
	
	#print("All variable:", locals())
	Status_Queue.put("Importing test case config")
	Start = time.time()
	#_kwargs = 
	AutoTester = Tester(Status_Queue = Status_Queue, Serial = Serial, DB_Path= DB_Path, Result_Folder_Path= Result_Folder_Path, Resolution = Resolution, 
				Language = Language, Tess_Path = Tess_Path, Tess_Data = Tess_Data)

	#os.system( ADBPATH + ' forward tcp:9889 tcp:9889')

	if TestCaseObject == None or len(TestCaseObject) == 0:
		Status_Queue.put('No action is added')
		return

	result, total = AutoTester.Function_Generate_TestCase(TestCaseObject, Execute_List)
	print('Generate testcase:')
	Status_Queue.put('Len of test cases: ' + str(len(TestCaseObject)))
	Status_Queue.put('Leng of real test case: '+  str(len(result)))

	#index = 0
	#for block in result:
		#index+=1
		#print('Block:', index, str(block))
	
	
	Connect_Status = AutoTester.Check_Connectivity()
	if Connect_Status == False:
		Status_Queue.put('Device is not connected.')
		return

	AutoTester.Function_Execute_Block(Status_Queue, Progress_Queue, Result_Queue, result)
		
	End = time.time()
	Status_Queue.put('Total testing time: ' + str(int(End-Start)) + " seconds.")	


###########################################################################################
def fixed_map(style, option):
	# Fix for setting text colour for Tkinter 8.6.9
	# From: https://core.tcl.tk/tk/info/509cafafae
	#
	# Returns the style map for 'option' with any styles starting with
	# ('!disabled', '!selected', ...) filtered out.
	# style.map() returns an empty list for missing options, so this
	# should be future-safe.
	return [elm for elm in style.map('Treeview', query_opt=option) if
	  elm[:2] != ('!disabled', '!selected')]

def main():
	global WIDTH, HEIGHT
	Process_Queue = Queue()
	Result_Queue = Queue()
	Status_Queue = Queue()
	Debug_Queue = Queue()
	
	MyManager = Manager()
	Default_Manager = MyManager.list()
	
	root = Tk()
	WIDTH = root.winfo_screenwidth()
	HEIGHT = root.winfo_screenheight()
	style = Style(root)
	style.map('Treeview', foreground=fixed_map(style, 'foreground'), background=fixed_map(style, 'background'))

	My_Queue = {}
	My_Queue['Process_Queue'] = Process_Queue
	My_Queue['Result_Queue'] = Result_Queue
	My_Queue['Status_Queue'] = Status_Queue
	My_Queue['Debug_Queue'] = Debug_Queue

	My_Manager = {}
	My_Manager['Default_Manager'] = Default_Manager
	Windows = None
	
	try:
		Windows = Automation_Execuser(root, Queue = My_Queue, Manager = My_Manager,)
		root.mainloop() 
	except Exception as e:
		root.withdraw()
		messagebox.showerror(title='Critical error', message=e)
		if Windows != None:
			Windows.destroy()	
			
			if Windows.Automation_Processor != None:
				try:
					Windows.Automation_Processor.terminate()
				except:
					pass	


if __name__ == '__main__':
	if sys.platform.startswith('win'):
		multiprocessing.freeze_support()

	main()
