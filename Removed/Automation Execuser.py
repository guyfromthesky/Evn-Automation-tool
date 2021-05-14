#System variable and io handling
import sys, getopt
import os


from multiprocessing import Process , Queue, Manager
import queue 
import subprocess
#Get timestamp
import time
from datetime import datetime

#GUI
from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from tkinter import colorchooser

from openpyxl import load_workbook, worksheet, Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Color
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
#MyTranslatorAgent = 'google'
Tool = "Automation Execuser"
VerNum = '0.0.1a'
version = Tool  + " " +  VerNum
DELAY1 = 20

 


#**********************************************************************************
# UI handle ***********************************************************************
#**********************************************************************************

class Automation_Execuser(Frame):
	def __init__(self, Root, Queue = None, Manager = None,):
		
		Frame.__init__(self, Root) 
		#super().__init__()
		self.parent = Root 

		# Queue
		self.Process_Queue = Queue['Process_Queue']
		self.Result_Queue = Queue['Result_Queue']
		self.Status_Queue = Queue['Status_Queue']
		self.Debug_Queue = Queue['Debug_Queue']

		self.Manager = Manager['Default_Manager']

		self.Config_Init()

		self.Options = {}

		# XLSX Optmizer
		self.Optimize_Folder = ""
		self.Optimize_FileList = ""
		# XLSX Comparision
		self.Compare_Folder_Old = ""
		self.Compare_File_List_Old = ""
		self.Compare_Folder_New = ""
		self.Compare_File_List_New = ""

		# UI Variable
		self.Button_Width_Full = 20
		self.Button_Width_Half = 15
		
		self.PadX_Half = 5
		self.PadX_Full = 10
		self.PadY_Half = 5
		self.PadY_Full = 10
		self.StatusLength = 120
		self.AppLanguage = 'en'

		
		self.App_LanguagePack = {}
		self.initGeneralSetting()

		if self.AppLanguage != 'kr':
			from languagepack import LanguagePackEN as LanguagePack
		else:
			from languagepack import LanguagePackKR as LanguagePack

		self.LanguagePack = LanguagePack

		# Init function

		self.parent.resizable(False, False)
		self.parent.title(version)
		# Creating Menubar 
		
		#**************New row#**************#
		self.Notice = StringVar()
		self.Debug = StringVar()
		self.Progress = StringVar()
	
		#Generate UI
		self.Generate_Menu_UI()
		self.Generate_Tab_UI()
		self.init_UI()
		self.init_UI_Configuration()

	def Config_Init(self):
		self.Roaming = os.environ['APPDATA'] + '\\Document_Utility'
		self.AppConfig = self.Roaming + '\\config.ini'
	
		if not os.path.isdir(self.Roaming):
			try:
				os.mkdir(self.Roaming)
			except OSError:
				print ("Creation of the directory %s failed" % self.Roaming)
		else:
			print('Roaming folder exist.')



	# UI init
	def init_UI(self):
	
		self.Generate_Data_Comparision_UI(self.DataCompare)
		#self.Generate_File_Comparision_UI(self.FileComparison)
		self.Generate_Folder_Comparision_UI(self.FolderComparison)
		self.Generate_Optimizer_UI(self.Optimizer)
		self.Generate_Debugger_UI(self.Process)
		
		'''	
		Row = 1
		Label(self.DataLookup, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1

		self.DataLookUpSource = StringVar()
		Label(self.DataLookup, text=  self.LanguagePack.Label['DataSource']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)

		self.TextDataLookupSource = Entry(self.DataLookup,width = 120, state="readonly", textvariable=self.DataLookUpSource)
		self.TextDataLookupSource.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=W)
		Button(self.DataLookup, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Browse'], command= self.BtnLoadDataLookupSource).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		

		Row += 1
		Label(self.DataLookup, text="Lookup value: ").grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.LookupValue = Text(self.DataLookup, width = 90, height=1) #
		self.LookupValue.grid(row=Row, column=3, columnspan=4, padx=5, pady=5, sticky=W)
		Button(self.DataLookup, width = self.Button_Width_Full, text=  self.LanguagePack.Button['Execute'], command= self.BtnCompareLookupData).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		
		'''
		
		# Debugger

	def Generate_Menu_UI(self):
		menubar = Menu(self.parent) 
		# Adding File Menu and commands 
		file = Menu(menubar, tearoff = 0)
		'''
		# Adding Load Menu 
		menubar.add_cascade(label =  self.LanguagePack.Menu['File'], menu = file) 
		file.add_command(label =  self.LanguagePack.Menu['LoadTM'], command = self.Menu_Function_Select_TM) 
		file.add_separator() 
		file.add_command(label =  self.LanguagePack.Menu['CreateTM'], command = self.Menu_Function_Create_TM)
		file.add_separator() 
		file.add_command(label =  self.LanguagePack.Menu['Exit'], command = self.parent.destroy) 
		'''
		# Adding Help Menu
		help_ = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Help'], menu = help_) 
		help_.add_command(label =  self.LanguagePack.Menu['GuideLine'], command = self.Menu_Function_Open_Main_Guideline) 
		help_.add_separator()
		help_.add_command(label =  self.LanguagePack.Menu['About'], command = self.Menu_Function_About) 
		self.parent.config(menu = menubar)

		# Adding Help Menu
		language = Menu(menubar, tearoff = 0) 
		menubar.add_cascade(label =  self.LanguagePack.Menu['Language'], menu = language) 
		language.add_command(label =  self.LanguagePack.Menu['Hangul'], command = self.SetLanguageKorean) 
		language.add_command(label =  self.LanguagePack.Menu['English'], command = self.SetLanguageEnglish) 
		self.parent.config(menu = menubar) 	

	def Generate_Tab_UI(self):
		self.TAB_CONTROL = ttk.Notebook(self.parent)
		#Tab
		
		self.DataCompare = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.DataCompare, text= self.LanguagePack.Tab['StructuredCompare'])
		'''
		#Tab
		self.FileComparison = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.FileComparison, text= self.LanguagePack.Tab['FileComparison'])
		#Tab
		'''
		self.FolderComparison = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.FolderComparison, text= self.LanguagePack.Tab['FolderComparison'])
		#Tab
		self.Optimizer = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.Optimizer, text= self.LanguagePack.Tab['Optimizer'])


		#Tab
		'''
		self.DataLookup = ttk.Frame(TAB_CONTROL)
		TAB_CONTROL.add(self.DataLookup, text=  'Data Lookup')
		
		'''	
		#Tab
		self.Process = ttk.Frame(self.TAB_CONTROL)
		self.TAB_CONTROL.add(self.Process, text= self.LanguagePack.Tab['Debug'])
		
		self.TAB_CONTROL.pack(expand=1, fill="both")
		return

	def Generate_Data_Comparision_UI(self, Tab):
		
		Row = 1
		Label(Tab, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1

		self.Str_Old_File_Path = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['MainDB']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_Old_File_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_Old_Data_File).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		#Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectBGColor'], command= self.Btn_Select_Background_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

		Row += 1
		self.Str_New_File_Path = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['NewDataTable']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_New_File_Path = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_New_File_Path)
		self.Entry_New_File_Path.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_New_Data_File).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		#Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectFontColor'], command= self.Btn_Select_Font_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		
	
		Row += 1
		Label(Tab, text=self.LanguagePack.Label['Main_Data_Sheet']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Str_Data_Sheet_Name = Text(Tab, width = 110, height=1) #
		self.Str_Data_Sheet_Name.grid(row=Row, column=2, columnspan=8, padx=5, pady=5, sticky=E)
		self.Str_Data_Sheet_Name.insert("end", 'Data')
		
		Row += 1
		Label(Tab, text=self.LanguagePack.Label['ID_Col']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Str_Data_Col_Name = Text(Tab, width = 110, height=1) #
		self.Str_Data_Col_Name.grid(row=Row, column=2, columnspan=8, padx=5, pady=5, sticky=E)
		self.Str_Data_Col_Name.insert("end", 'StringId')
		
		Row += 1

		#Label(Tab, text=  self.LanguagePack.Label['CompareOptions']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)

		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.Btn_Compare_Data).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		'''
		Row += 1
		self.CompareFontBold = IntVar()
		FontBold = Checkbutton(Tab, text=  self.LanguagePack.Option['Bold'], variable = self.CompareFontBold, command= None)
		FontBold.grid(row=Row, column=1,padx=5, pady=5, sticky=W)
		FontBold.bind("<Enter>", lambda event : self.Notice.set(self.LanguagePack.ToolTips['Bold']))

		self.CompareFontUnderLine = IntVar()
		FontUnderLine = Checkbutton(Tab, text=  self.LanguagePack.Option['UnderLine'], variable = self.CompareFontUnderLine, command= None)
		FontUnderLine.grid(row=Row, column=3,padx=5, pady=5, sticky=W)
		FontUnderLine.bind("<Enter>", lambda event : self.Notice.set(self.LanguagePack.ToolTips['UnderLine']))
		Row += 1

		self.AllData = IntVar()
		AllDataOption = Checkbutton(Tab, text=  self.LanguagePack.Option['CompareAll'], variable = self.AllData, command= None)
		AllDataOption.grid(row=Row, column=1,padx=5, pady=5, sticky=W)
		AllDataOption.bind("<Enter>", lambda event : self.Notice.set(self.LanguagePack.ToolTips['CompareAll']))
		self.AllData.set(0)
		
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row +=1

		self.CompareProgressbar = Progressbar(Tab, orient=HORIZONTAL, length=1000,  mode='determinate')
		self.CompareProgressbar["maximum"] = 1000
		self.CompareProgressbar.grid(row=Row, column=1, columnspan=9, padx=5, pady=5, sticky=W)		
		'''

	def Generate_File_Comparision_UI(self, Tab):
		
		Row = 1
		Label(Tab, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1

		self.Str_Old_File_Path = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['OldDataTable']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_Old_File_Path = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_Old_File_Path)
		self.Entry_Old_File_Path.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_Old_Data_File).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectBGColor'], command= self.Btn_Select_Background_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

		Row += 1
		self.Str_New_File_Path = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['NewDataTable']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.Entry_New_File_Path = Entry(Tab,width = 110, state="readonly", textvariable=self.Str_New_File_Path)
		self.Entry_New_File_Path.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_New_Data_File).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectFontColor'], command= self.Btn_Select_Font_Colour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		
		Row += 1
		Label(Tab, text=self.LanguagePack.Label['Main_Data_Sheet']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Str_Data_Sheet_Name = Text(Tab, width = 110, height=1) #
		self.Str_Data_Sheet_Name.grid(row=Row, column=2, columnspan=8, padx=5, pady=5, sticky=E)
		self.Str_Data_Sheet_Name.insert("end", 'Data')
		
		Row += 1
		Label(Tab, text=self.LanguagePack.Label['Main_Data_Sheet']).grid(row=Row, column=1, padx=5, pady=5, sticky=W)
		self.Str_Data_Col_Name = Text(Tab, width = 110, height=1) #
		self.Str_Data_Col_Name.grid(row=Row, column=2, columnspan=8, padx=5, pady=5, sticky=E)
		self.Str_Data_Col_Name.insert("end", 'Id')
		
		Row += 1

		Label(Tab, text=  self.LanguagePack.Label['CompareOptions']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)

		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.Btn_Compare_Data).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		'''
		Row += 1
		self.CompareFontBold = IntVar()
		FontBold = Checkbutton(Tab, text=  self.LanguagePack.Option['Bold'], variable = self.CompareFontBold, command= None)
		FontBold.grid(row=Row, column=1,padx=5, pady=5, sticky=W)
		FontBold.bind("<Enter>", lambda event : self.Notice.set(self.LanguagePack.ToolTips['Bold']))

		self.CompareFontUnderLine = IntVar()
		FontUnderLine = Checkbutton(Tab, text=  self.LanguagePack.Option['UnderLine'], variable = self.CompareFontUnderLine, command= None)
		FontUnderLine.grid(row=Row, column=3,padx=5, pady=5, sticky=W)
		FontUnderLine.bind("<Enter>", lambda event : self.Notice.set(self.LanguagePack.ToolTips['UnderLine']))
		Row += 1

		self.AllData = IntVar()
		AllDataOption = Checkbutton(Tab, text=  self.LanguagePack.Option['CompareAll'], variable = self.AllData, command= None)
		AllDataOption.grid(row=Row, column=1,padx=5, pady=5, sticky=W)
		AllDataOption.bind("<Enter>", lambda event : self.Notice.set(self.LanguagePack.ToolTips['CompareAll']))
		self.AllData.set(0)
		
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row +=1

		self.CompareProgressbar = Progressbar(Tab, orient=HORIZONTAL, length=1000,  mode='determinate')
		self.CompareProgressbar["maximum"] = 1000
		self.CompareProgressbar.grid(row=Row, column=1, columnspan=9, padx=5, pady=5, sticky=W)		
		'''

	def Generate_Folder_Comparision_UI(self, Tab):
		
		Row = 1
		Label(Tab, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1

		self.OldDataString = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['OldDataTable']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.TextRawTMPath = Entry(Tab,width = 110, state="readonly", textvariable=self.OldDataString)
		self.TextRawTMPath.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_Old_Data_Folder).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['SelectColor'], command= self.BtnSelectColour).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		
		Row += 1
		self.NewDataString = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['NewDataTable']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
		self.TextRawTMPath = Entry(Tab,width = 110, state="readonly", textvariable=self.NewDataString)
		self.TextRawTMPath.grid(row=Row, column=3, columnspan=4, padx=4, pady=5, sticky=E)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.Btn_Browse_New_Data_Folder).grid(row=Row, column=7, columnspan=2, padx=5, pady=5, sticky=E)
		
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.BtnCompareDataTable).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
	
		Row += 1
		Label(Tab, text=  self.LanguagePack.Label['CompareOptions']).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)

		Row += 1
		self.CompareFontBold = IntVar()
		FontBold = Checkbutton(Tab, text=  self.LanguagePack.Option['Bold'], variable = self.CompareFontBold, command= None)
		FontBold.grid(row=Row, column=1,padx=5, pady=5, sticky=W)
		FontBold.bind("<Enter>", lambda event : self.Notice.set(self.LanguagePack.ToolTips['Bold']))

		self.CompareFontUnderLine = IntVar()
		FontUnderLine = Checkbutton(Tab, text=  self.LanguagePack.Option['UnderLine'], variable = self.CompareFontUnderLine, command= None)
		FontUnderLine.grid(row=Row, column=3,padx=5, pady=5, sticky=W)
		FontUnderLine.bind("<Enter>", lambda event : self.Notice.set(self.LanguagePack.ToolTips['UnderLine']))
		Row += 1

		self.AllData = IntVar()
		AllDataOption = Checkbutton(Tab, text=  self.LanguagePack.Option['CompareAll'], variable = self.AllData, command= None)
		AllDataOption.grid(row=Row, column=1,padx=5, pady=5, sticky=W)
		AllDataOption.bind("<Enter>", lambda event : self.Notice.set(self.LanguagePack.ToolTips['CompareAll']))
		self.AllData.set(0)

		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	

		Row +=1

		self.CompareProgressbar = Progressbar(Tab, orient=HORIZONTAL, length=1000,  mode='determinate')
		self.CompareProgressbar["maximum"] = 1000
		self.CompareProgressbar.grid(row=Row, column=1, columnspan=9, padx=5, pady=5, sticky=W)
		
	def Generate_Optimizer_UI(self, Tab):
		
		Row = 1
		Label(self.Optimizer, textvariable=self.Notice).grid(row=Row, column=1, columnspan = 10, padx=5, pady=5, sticky= W)
		Row += 1
		self.RawSource = StringVar()
		Label(Tab, text=  self.LanguagePack.Label['OptimizeDatafile'],  width = self.Button_Width_Half).grid(row=Row, column=1, columnspan=2, padx=5, pady=5, sticky= W)
	
		self.TextRawSourcePath = Entry(Tab,width = 130, state="readonly", textvariable=self.RawSource)
		self.TextRawSourcePath.grid(row=Row, column=3, columnspan=6, padx=5, pady=5, sticky=W)
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Browse'], command= self.BtnLoadRawSource).grid(row=Row, column=9, columnspan=2, padx=5, pady=5, sticky=E)
		Row+=1
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['Execute'], command= self.BtnOptimizeXLSX).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)
		Row+=1
		Button(Tab, width = self.Button_Width_Half, text=  self.LanguagePack.Button['OpenOutput'], command= self.OpenOptimizerFolder).grid(row=Row, column=9, columnspan=2,padx=5, pady=5, sticky=W)

		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	
		Row+=1
		Label(Tab, text=  None).grid(row=Row, column=1)	

		
		self.Optimize_Progressbar = Progressbar(Tab, orient=HORIZONTAL, length=1000,  mode='determinate')
		self.Optimize_Progressbar["maximum"] = 1000
		self.Optimize_Progressbar.grid(row=10, column=1, columnspan=9, padx=5, pady=5, sticky=W)

	def Generate_Debugger_UI(self,Tab):
		Row = 1
		self.Debugger = Text(Tab, width=125, height=15, undo=True, wrap=WORD, )
		self.Debugger.grid(row=Row, column=1, columnspan=10, padx=5, pady=5, sticky=W+E+N+S)



	# Menu Function
	def Menu_Function_About(self):
		messagebox.showinfo("About....", "Creator: Evan")

	def Show_Error_Message(self, ErrorText):
		messagebox.showinfo('Error...', ErrorText)	

	def SaveAppLanguage(self, language):

		self.Notice.set('Update app language...') 

		config = configparser.ConfigParser()
		config.read(self.AppConfig)
		if not config.has_section('DocumentToolkit'):
			config.add_section('DocumentToolkit')
			cfg = config['DocumentToolkit']	
		else:
			cfg = config['DocumentToolkit']

		cfg['applang']= language
		with open(self.AppConfig, 'w') as configfile:
			config.write(configfile)
		self.Notice.set('Config saved...')
		return

	def SetLanguageKorean(self):
		self.AppLanguage = 'kr'
		self.SaveAppLanguage(self.AppLanguage)
		#self.initUI()
	
	def SetLanguageEnglish(self):
		self.AppLanguage = 'en'
		self.SaveAppLanguage(self.AppLanguage)
		#self.initUI()

	def Menu_Function_Open_Main_Guideline(self):
		webbrowser.open_new(r"https://confluence.nexon.com/display/NWMQA/Document+Toolkit")

	def Function_Correct_Path(self, path):
		return str(path).replace('/', '\\')
	
	def Function_Correct_EXT(self, path, ext):
		if path != None and ext != None:
			Outputdir = os.path.dirname(path)
			baseName = os.path.basename(path)
			sourcename, Obs_ext = os.path.splitext(baseName)
			newPath = self.CorrectPath(Outputdir + '/'+ sourcename + '.' + ext)
			return newPath

	def ErrorMsg(self, ErrorText):
		messagebox.showinfo('Error...', ErrorText)	

	def OpenOutput(self):
		Source = self.ListFile[0]

		Outputdir = os.path.dirname(Source)
		BasePath = str(os.path.abspath(Outputdir))
		subprocess.Popen('explorer ' + BasePath)
		
	def BtnLoadDocument(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("All type files","*.docx *.xlsx *.pptx *.msg"), ("Workbook files","*.xlsx *.xlsm"), ("Document files","*.docx"), ("Presentation files","*.pptx"), ("Email files","*.msg")), multiple = True)	
		if filename != "":
			self.ListFile = list(filename)
			self.CurrentSourceFile.set(str(self.ListFile[0]))
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def BtnLoadRawSource(self):
		#filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files","*.xlsx *.xlsm *xlsb"), ("Document files","*.docx")), multiple = True)	
		FolderName = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'])	
		if FolderName != "":
			self.RawFile = FolderName
			self.RawSource.set(str(FolderName))
			
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return
		'''
		if filename != "":
			self.RawFile = list(filename)
			self.RawSource.set(self.RawFile)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return
		'''

	def BtnLoadTrackingSource(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("TM file","*.pkl"),),)	
		if filename != "":
			self.TrackingFile = self.CorrectPath(filename)
			print(self.TrackingFile)
			self.TrackingSource.set(self.CorrectPath(self.TrackingFile))
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def BtnLoadRawTM(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("TM file","*.pkl"),), multiple = True)	
		if filename != "":
			self.RawTMFile = list(filename)
			Display = self.CorrectPath(self.RawTMFile[0])
			self.RawTMSource.set(Display)
			
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return
	
	
	def BtnSelectColour(self):
		colorStr, self.BackgroundColor = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.BackgroundColor == None:
			self.ErrorMsg('Set colour as defalt colour (Yellow)')
			self.BackgroundColor = 'ffff00'
		else:
			self.BackgroundColor = self.BackgroundColor.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return

	def Btn_Browse_Old_Data_Folder(self):
		FolderName = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'])	
		if FolderName != "":
			self.OldDataTable = FolderName
			self.OldDataString.set(str(FolderName))
			
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_Browse_New_Data_Folder(self):
		FolderName = filedialog.askdirectory(title =  self.LanguagePack.ToolTips['SelectSource'])	
		if FolderName != "":
			self.NewDataTable = FolderName
			
			self.NewDataString.set(str(FolderName))
			
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return	

	def BtnLoadDataLookupSource(self):
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files *.xlsx"), ), multiple = True)	
		if filename != "":
			self.RawFile = list(filename)
			self.RawSource.set(self.RawFile)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def onExit(self):
		self.quit()


	def initGeneralSetting(self):
		
		config = configparser.ConfigParser()
		if os.path.isfile(self.AppConfig):
			config.read(self.AppConfig)
			if config.has_section('DocumentToolkit'):
				cfg = config['DocumentToolkit']
			else:
				config['DocumentToolkit'] = {}
				cfg = config['DocumentToolkit']

			if config.has_option('DocumentToolkit', 'applang'):	
				self.AppLanguage = config['DocumentToolkit']['applang']
				print('Setting saved: ', self.AppLanguage)
			else:
				self.AppLanguage = 'en'
				print('Setting not saved: ', self.AppLanguage)

			#if config.has_option('Translator', 'Subscription'):
			#	self.Subscription = config['Translator']['Subscription']
			#else:
			#	self.Subscription = ''

		else:

			self.AppLanguage = 'en'

	def init_UI_Configuration(self):
		
		config = configparser.ConfigParser()
		if os.path.isfile(self.AppConfig):
			config.read(self.AppConfig)
			if config.has_section('Document_Utility'):
				cfg = config['Document_Utility']
			else:
				config['Document_Utility'] = {}
				cfg = config['Document_Utility']

			if config.has_section('Comparision'):
				cfg = config['Comparision']
			else:
				config.add_section('Comparision')
				cfg = config['Comparision']

			
			'''
			if config.has_option('DocumentTranslator', 'turbo'):
				self.TurboTranslate.set(int(config['DocumentTranslator']['turbo']))
			else:
				self.TurboTranslate.set(0)
			'''
			
		else:
			self.Language = 'en'
			

		return	

	def SaveSetting(self):

		config = configparser.ConfigParser()
		config.read(self.AppConfig)
		if not config.has_section('DocumentTranslator'):
			config.add_section('DocumentTranslator')
			cfg = config['DocumentTranslator']	
		else:
			cfg = config['DocumentTranslator']

		cfg['lang']= str(self.Language.get())
		cfg['turbo']=  str(self.TurboTranslate.get())
		cfg['value']= str(self.DataOnly.get())
		cfg['filename']= str(self.TranslateFileName.get())
		cfg['sheetname']= str(self.TranslateSheetName.get())
		cfg['tmenable']= str(self.TMTranslate.get())
		cfg['tmupdate']= str(self.TMUpdate.get())
		cfg['sheetremoval']= str(self.SheetRemoval.get())
		#config['Document Translator'] = {'overwrite': self.OverWrite.get()}
		#config['Document Translator'] = {'skiprow': self.SkipEmptyRow.get()}
		#config['Document Translator'] = {'sheetremoval': self.SheetRemoval.get()}
		with open(appconfig, 'w') as configfile:
			config.write(self.AppConfig)
		self.Notice.set('Config saved...')
		return

	def GenerateTranslatorEngine(self):
		self.Notice.set(self.LanguagePack.ToolTips['AppInit'])
		if self.Language.get() == 1:
			to_language = 'ko'
			from_language = 'en'
		else:
			to_language = 'en'
			from_language = 'ko'

		self.p1 = Process(target=GenerateTranslator, args=(self.MyTranslator_Queue, self.TMManager, self.TrackingManager, self.MyTranslatorAgent, from_language, to_language, self.SubscriptionKey,))
		self.p1.start()
		self.after(DELAY1, self.GetMyTranslator)
		return

	def GenerateTranslatorEngineWithAlternativeDB(self):
		self.Notice.set(self.LanguagePack.ToolTips['AppInit'])
		if self.Language.get() == 1:
			to_language = 'ko'
			from_language = 'en'
		else:
			to_language = 'en'	
			from_language = 'ko'

		self.p1 = Process(target=GenerateTranslatorWithAlternativeDB, args=(self.MyTranslator_Queue, self.TMManager, self.TrackingManager, self.MyTranslatorAgent, from_language, to_language, self.SubscriptionKey, ))
		self.p1.start()
		self.after(DELAY1, self.GetMyTranslator)
		return
		


	def GetMyTranslator(self):
		try:
			self.MyTranslator = self.MyTranslator_Queue.get_nowait()
		except queue.Empty:
			self.after(DELAY1, self.GetMyTranslator)

		#print("self.MyTranslator: ", self.MyTranslator)	
		if self.MyTranslator != None:	
			self.TranslateBtn.configure(state=NORMAL)
			self.DictionaryStatus.set(str(len(self.MyTranslator.Dictionary)))
			self.ExceptionStatus.set(str(len(self.MyTranslator.Exception)))	
			self.TMStatus.set(str(len(self.MyTranslator.TranslationMemory)))



			self.DictionaryPath = self.MyTranslator.DictionaryPath
			self.AlternativeDictionaryPath = self.MyTranslator.AlternativeDictionaryPath
			self.ExceptionPath = self.MyTranslator.ExceptionPath
			self.TMPath = self.MyTranslator.TMPath
			
			self.CurrentAlternativeDB.set(self.AlternativeDictionaryPath)
			self.CurrentDB.set(self.DictionaryPath)
			self.Notice.set(self.LanguagePack.ToolTips['AppInitDone'])
			self.p1.join()
		else:
			self.Notice.set(self.LanguagePack.ToolTips['AppInit']) 
		return
	
	def TMTranslateModeToggle(self):
		if self.TMTranslate.get() == 1:
			self.MyTranslator.TMModeEnable(True)
		else:
			self.MyTranslator.TMModeEnable(False)

	def GetOptions(self):
		#Get and set language
		if self.Language.get() == 1:
			self.Options['to_language'] = 'ko'
			self.Options['LanguageName'] = self.LanguagePack.Option['Hangul']
			self.Options['from_language'] = 'en'
		else:
			self.Options['to_language'] = 'en'
			self.Options['LanguageName'] = self.LanguagePack.Option['English']
			self.Options['from_language'] = 'ko'
		
		self.MyTranslator.SetTargetLanguage(self.Options['to_language'])
		self.MyTranslator.SetSourceLanguage(self.Options['from_language'])
		self.Notice.set(self.LanguagePack.ToolTips['SetLanguage'] + self.Options['LanguageName'])
		#Set translator engine
		self.MyTranslator.SetTranslatorAgent(self.MyTranslatorAgent)

		
		#Add Subscription key
		self.MyTranslator.SetSubscriptionKey(self.SubscriptionKey)	

		#Set TM Update Mode
		if self.TMUpdate.get() == 1:
			self.MyTranslator.TMUpdateModeEnable(True)
		else:
			self.MyTranslator.TMUpdateModeEnable(False)

		#Set Predict mode 
		if self.TurboTranslate.get() == 1:
			self.MyTranslator.PredictModeEnable(True)
		else:
			self.MyTranslator.PredictModeEnable(False)

		#Set Data Mode
		if self.DataOnly.get() == 1:
			self.Options['DataMode']  = True
		else:
			self.Options['DataMode'] = False	

		#Set Skip Mode
		if self.SkipEmptyRow.get() == 1:
			self.Options['SkipMode'] = True
		else:
			self.Options['SkipMode'] = False

		#Set Sheet removal mode
		if self.SheetRemoval.get() == 1:
			self.Options['SheetRemovalMode'] = True
		else:
			self.Options['SheetRemovalMode'] = False

		# Set Update Mode
		if self.TMUpdate.get() == 1:
			self.Options['TMUpdateMode'] = True
		else:
			self.Options['TMUpdateMode'] = False



		#Set TM Translate Mode
		print('self.TMTranslate', self.TMTranslate.get())
		if self.TMTranslate.get() == 1:
			self.Options['MemoryTranslate']  = True
		else:
			self.Options['MemoryTranslate'] = False

		#Set Translate File name option
		if self.TranslateFileName.get() == 1:
			self.Options['TranslateFileName'] = True
		else:
			self.Options['TranslateFileName'] = False
		
		#Set Translate Sheet name option
		if self.TranslateSheetName.get() == 1:
			self.Options['TranslateSheetName'] = True
		else:
			self.Options['TranslateSheetName'] = False

		#Set Name fixer mode
		if self.FixCorruptFileName.get() == 1:
			self.Options['FixCorruptFileName'] = True
		else:
			self.Options['FixCorruptFileName'] = False	

		#Get sheet list
		try:
			Raw = self.SheetList.get("1.0", END).replace("\n", "").replace(" ", "")	
			self.Options['Sheet'] = Raw.split(",")
		except:
			self.Options['Sheet'] = []

		#Calculate Multiple value
		try:
			self.Options['Multiple'] = int(self.Multiple.get("1.0", END).replace("\n", "").replace(" ", ""))
		except:
			self.Options['Multiple'] = 8

		# Get Documents list
		self.Options['SourceDocument'] = self.ListFile
		if self.Options['SourceDocument'] == "":
			self.Notice.set(self.LanguagePack.ToolTips['SourceNotSelected']) 
			return
		else:
			self.Notice.set(self.LanguagePack.ToolTips['DocumentLoad']) 


	def BtnExportTrackingData(self):
		SourceDocument = self.TrackingFile

		self.processTracking = Process(target=ExportData, args=(SourceDocument, self.StatusQueue,))
		self.processTracking.start()
		self.after(DELAY1, self.GetTrackingStatus)	

	def GetTrackingStatus(self):
		if (self.processTracking.is_alive()):
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.GetTrackingStatus)
		else:
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.processTracking.terminate()

	def BtnBrowseTMSource(self):
		filename = filedialog.askopenfilename(title = "Select TM file",filetypes = (("Translation Memory","*.pkl"), ), multiple = True)	
		if filename != "":
			self.TMSourceList = list(filename)
			for Path in self.TMSourceList:
				Path =  self.CorrectPath(Path)
			
			Display = str(self.CorrectPath(self.TMSourceList[0]))

			self.TMSource.set(Display)
		else:
			self.Notice.set("No document is selected")
			return

	def BtnMergeTM(self):
		if self.TMSource == None:
			self.Notice.set("Please select TM file before starting")
		#elif not os.path.isfile(self.DictionarySource):
		#	self.Notice.set("TM files has been removed")
		else:	
			filename = filedialog.asksaveasfilename(title = "Save file to", filetypes = (("Translation Memory", "*.pkl"),),)
			newPath = self.CorrectExt(filename, "pkl")
			print("Save TM to: ", newPath)
			'''
			Outputdir = os.path.dirname(filename)
			baseName = os.path.basename(filename)
			sourcename, ext = os.path.splitext(baseName)

			newPath = Outputdir + '/'+ sourcename + '.pkl'
			'''
			if filename == "":
				self.Notice.set("Please enter a file name.")
				self.ErrorMsg("Please enter a file name.")
			else:
				
				SourceDocument = self.TMSourceList

				self.p4 = Process(target=MergeTMSource, args=(SourceDocument, newPath, self.StatusQueue,))
				self.p4.start()
				self.after(DELAY1, self.GetMergeTMStatus)	

	def GetMergeTMStatus(self):
		if (self.p4.is_alive()):
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.GetOptimizeStatus)
		else:
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.p4.terminate()

	def OpenOptimizerFolder(self):
		try:
			SourceDocument = self.RawFile
		except AttributeError:
			self.ErrorMsg('Please select source folder.')
			return	
		BasePath = str(os.path.abspath(self.RawFile))
		subprocess.Popen('explorer ' + BasePath)
	
	def BtnOptimizeXLSX(self):
		try:
			SourceDocument = self.RawFile
		except AttributeError:
			self.ErrorMsg('Please select source folder.')

		try:
			while True:
				percent = self.Process_Queue.get_nowait()
				#print("Remain percent: ", percent)
		except queue.Empty:
			pass
		self.Optimize_Process = Process(target=Function_Optimize_XLSX, args=(SourceDocument, self.Status_Queue,self.Process_Queue,))
		self.Optimize_Process.start()
		self.after(DELAY1, self.GetOptimizeStatus)	

	def GetOptimizeStatus(self):
		if (self.Optimize_Process.is_alive()):
			try:
				percent = self.Process_Queue.get(0)
				self.Optimize_Progressbar["value"] = percent
				self.Optimize_Progressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass
			
			
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			
			self.after(DELAY1, self.GetOptimizeStatus)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.Optimize_Process.terminate()

	def BtnOptimizeTM(self):
		SourceDocument = self.RawTMFile

		self.p4 = Process(target=OptimizeTM, args=(SourceDocument, self.Status_Queue,))
		self.p4.start()
		self.after(DELAY1, self.GetOptimizeTMStatus)	

	def GetOptimizeTMStatus(self):
		if (self.p4.is_alive()):
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.GetOptimizeStatus)
		else:
			try:
				Status = self.StatusQueue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.p4.terminate()

	def BtnCompareLookupData(self):
		#OldDocument = self.OldDataTable
		LookupData = self.DataLookUpSource
		if self.AllData.get() == 0:
			DataOnly = True
		else:
			DataOnly = False

		self.p4 = Process(target=CompareTableData, args=(self.StatusQueue, LookupData, LookupValue, Mode, DataOnly,))
		self.p4.start()
		self.after(DELAY1, self.GetCompareStatus)

###########################################################################################
	def BtnCompareDataTable(self):

		try:
			OldDocument = self.OldDataTable
			NewDocument = self.NewDataTable
		except AttributeError:
			self.ErrorMsg('Please select source folder.')
			return

		try:
			self.BackgroundColor
		except:
			self.BackgroundColor = 'ffff00'	
		if self.BackgroundColor == False or self.BackgroundColor == None:
			self.BackgroundColor = 'ffff00'
		print('self.BackgroundColor: ', self.BackgroundColor)

		try:
			self.BackgroundColor
		except:
			self.BackgroundColor = 'ffff00'	
		if self.BackgroundColor == False or self.BackgroundColor == None:
			self.BackgroundColor = 'ffff00'
		print('self.BackgroundColor: ', self.BackgroundColor)

		if self.AllData.get() == 0:
			DataOnly = True
		else:
			DataOnly = False

		self.p4 = Process(target=Compare_Folder_Data, args=(self.Status_Queue, self.Process_Queue, OldDocument, NewDocument, self.BackgroundColor, DataOnly,))
		self.p4.start()
		self.after(DELAY1, self.GetCompareStatus)	

	def GetCompareStatus(self):
		if (self.p4.is_alive()):
			try:
				percent = self.Process_Queue.get(0)
				self.CompareProgressbar["value"] = percent
				self.CompareProgressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass	
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.GetCompareStatus)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set(Status)
					print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.p4.terminate()

###########################################################################################

###########################################################################################
	def Btn_Select_Background_Colour(self):
		colorStr, self.Background_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.Background_Color == None:
			self.Error('Set colour as defalt colour (Yellow)')
			self.Background_Color = 'ffff00'
		else:
			self.Background_Color = self.Background_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return
	
	def Btn_Select_Font_Colour(self):
		colorStr, self.Font_Color = colorchooser.askcolor(parent=self, title='Select Colour')
		
		
		if self.Font_Color == None:
			self.Error('Set colour as defalt colour (Yellow)')
			self.Font_Color = 'FF0000'
		else:
			self.Font_Color = self.Font_Color.replace('#', '')
		#print(colorStr)
		#print(self.BackgroundColor)
		return

	def Btn_Browse_Old_Data_File(self):
			
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = False)	
		if filename != "":
			self.Old_File_Path = self.Function_Correct_Path(filename)
			self.Str_Old_File_Path.set(self.Old_File_Path)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_Browse_New_Data_File(self):
		
		filename = filedialog.askopenfilename(title =  self.LanguagePack.ToolTips['SelectSource'],filetypes = (("Workbook files", "*.xlsx *.xlsm"), ), multiple = False)
		
		if filename != "":
			self.New_File_Path = self.Function_Correct_Path(filename)
			self.Str_New_File_Path.set(self.New_File_Path)
			self.Notice.set(self.LanguagePack.ToolTips['SourceSelected'])
		else:
			self.Notice.set(self.LanguagePack.ToolTips['SourceDocumentEmpty'])
		return

	def Btn_Compare_Data(self):
		Old_File = self.Old_File_Path
		New_File = self.New_File_Path

		Sheet_Name = "Data"
		
		try:
			Sheet_Name = self.Str_Data_Sheet_Name.get("1.0", END).replace('\n', '')
		except Exception as e:
			ErrorMsg = ('Error message: ' + str(e))
			print(ErrorMsg)

		Index_Col = "Id"
		try:
			Index_Col = self.Str_Data_Col_Name.get("1.0", END).replace('\n', '')
		except Exception as e:
			ErrorMsg = ('Error message: ' + str(e))
			print(ErrorMsg)
		

		#Sheet_Name
		'''
		if self.AllData.get() == 0:
			DataOnly = True
		else:
			DataOnly = False
		'''
		try:
			self.Background_Color
		except:
			self.Background_Color = 'ffff00'	
		if self.Background_Color == False or self.Background_Color == None:
			self.Background_Color = 'ffff00'
		print('Background_Color: ', self.Background_Color)
		
		try:
			self.Font_Color
		except:
			self.Font_Color = 'FF0000'	
		if self.Font_Color == False or self.Font_Color == None:
			self.Font_Color = 'FF0000'
		print('Font_Color: ', self.Font_Color)

		timestamp = Function_Get_TimeStamp()			
		path, filename = os.path.split(New_File)
		Output_Result =path + '/' + 'Compare_Result_' + str(timestamp) + '.xlsx'

		self.Data_Compare_Process = Process(target=Function_Compare_Data, args=(self.Status_Queue, self.Process_Queue, Old_File, New_File, Output_Result, Sheet_Name, Index_Col, self.Background_Color, self.Font_Color,))
		#self.Data_Compare_Process = Process(target=Old_Function_Compare_Excel, args=(self.Status_Queue, self.Process_Queue, Old_File, New_File, Output_Result, Sheet_Name, Index_Col, self.Background_Color, self.Font_Color,))
		self.Data_Compare_Process.start()
		self.after(DELAY1, self.Wait_For_Data_Compare_Process)	

	def Wait_For_Data_Compare_Process(self):
		if (self.Data_Compare_Process.is_alive()):
			'''
			try:
				percent = self.ProcessQueue.get(0)
				self.CompareProgressbar["value"] = percent
				self.progressbar.update()
				#self.Progress.set("Progress: " + str(percent/10) + '%')
			except queue.Empty:
				pass	
			'''
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:
					self.Notice.set(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass	
			self.after(DELAY1, self.Wait_For_Data_Compare_Process)
		else:
			try:
				Status = self.Status_Queue.get(0)
				if Status != None:	
					self.Notice.set('Compare complete')
					print(Status)
					self.Debugger.insert("end", "\n\r")
					self.Debugger.insert("end", Status)
			except queue.Empty:
				pass
			self.Data_Compare_Process.terminate()

###########################################################################################
	
	

###########################################################################################

def Function_Optimize_XLSX(SourceDocument, StatusQueue, ProgressQueue):
	from openpyxl import load_workbook, worksheet, Workbook
	from openpyxl.styles import Font
	#print(SourceDocument)
	FileList = os.listdir(SourceDocument)
	TotalFile = len(FileList)
	Complete = 0
	#print(FileList)
	#Outputdir = os.path.dirname(File) + '/Optimized/'
	FolderName = os.path.basename(SourceDocument)
	Outputdir = SourceDocument + '//' + FolderName + '_Optimized'
	if not os.path.isdir(Outputdir):
		try:
			os.mkdir(Outputdir)
		except OSError:
			print ("Creation of the directory %s failed" % Outputdir)
			Outputdir = SourceDocument + '//Optmized'

	for FileName in FileList:

		if FileName != None:
			File = SourceDocument + '//' + FileName
			try:
				xlsx = load_workbook(File, data_only=True)
			except:
				continue
			
			baseName = os.path.basename(File)
			sourcename, ext = os.path.splitext(baseName)
			
			StatusQueue.put('Optimizing file: ' + sourcename)
			
			output_file = Outputdir+ '//' + sourcename + ext
			#print('output_file:',output_file)
			
			try:
				xlsx.save(output_file)
				#StatusQueue.put('Optimized done.')
			except Exception as e:
				StatusQueue.put('Failed to save the result: ' + str(e))
			
			Complete+=1
			
			percent = ShowProgress(Complete, TotalFile)
			ProgressQueue.put(percent)				

	StatusQueue.put('Optimized done.')	

###########################################################################################

# Functions used for processor
def CheckList(item, list):
	if list == [] or list == [""]:
		return True
	for text in list:
		if item == text:
			return True
	return False

# Functions used for processor
def GenerateSheetList(SheetList, TranslateList):
	if TranslateList == [] or TranslateList == [""]:
		return SheetList
	TotalSheet = len(SheetList)
	IndexList = []
	for sheet in TranslateList:
		print("Sheet: ", sheet)
		if sheet.isnumeric():
			IndexList.append(int(sheet))
		else:	
			tempList = sheet.split('-')
			#print('tempList: ', tempList)
			#print('Len tempList: ', len(tempList))
			invalid = False
			if len(tempList) == 2:
				for tempIndex in tempList:
					#print("Item: ",  tempIndex)
					if tempIndex.isnumeric() != True:
						invalid = True
			else:
				invalid = True
			if invalid == False:
				
				A = int(tempList[0])
				B = int(tempList[1])
				
				if A < B:
					Min = A
					Max = B
				else:
					Min = B
					Max = A	
				#print('Min max: ', Min, Max)
				i = Min
				while i >= Min and i <= Max:
					IndexList.append(i)
					i+=1
			else:
				i = 0
				for source in SheetList:
					if source == sheet:
						IndexList.append(i)		
					i+=1

	toReturn = []

	i = 1
	for toTranslate in SheetList:
		for num in IndexList:
			if i == num:
				toReturn.append(toTranslate)
		i+=1
	return toReturn

def ShowProgress(Counter, TotalProcess):
	#os.system('CLS') 
	percent = int(1000 * Counter / TotalProcess)
	#print("Current progress: " +  str(Counter) + '/ ' + str(TotalProcess))
	return percent


# xlsx object
# Add a cell as a task with an address
class CellData:
	def __init__(self, SheetName, CellAddress, Text):
		self.Sheet = SheetName
		self.Address = CellAddress
		self.Text = Text
		self.Fail = 0

# pptx object
# Add a Paragraph as a task with an address
class TextFrameData:
	def __init__(self, ParagraphsNum, RunsNum, String):
		self.Paragraphs = ParagraphsNum
		self.Runs = RunsNum
		self.Text = String

# docx object
class ParagraphData:
	def __init__(self, ParagraphsNum, RunList):
		self.Paragraphs = ParagraphsNum
		self.RunList = RunList

class TableData:
	def __init__(self, ParagraphsNum, RunsNum, String):
		self.Paragraphs = ParagraphsNum
		self.Runs = RunsNum
		self.Text = String

class RunData:
	def __init__(self, RunsNum, String):
		self.Runs = RunsNum
		self.Text = String

###########################################################################################

def Compare_Folder_Data(StatusQueue, ProgressQueue, OldData, NewData, BackgroundColor, DataOnly, ):

	print('BackgroundColor: ', BackgroundColor)
	my_color = Color(rgb=BackgroundColor)
	my_fill = PatternFill(patternType='solid', fgColor=my_color)
	print('DataOnly: ', DataOnly)
	NewAdded = []
	Removed = []
	Changed = []
	NoChange = []
	NotData = []
	Broken = []
	OldFiles = os.listdir(OldData)
	TotalOldFiles = len(OldFiles)
	NewFiles = os.listdir(NewData)
	TotalNewFiles = len(NewFiles)
	Checked = 0
	#print(OldFile)
	Result = []
	ChangedFolder = NewData + '/Compare Result'
	if not os.path.isdir(ChangedFolder):
			try:
				os.mkdir(ChangedFolder)
			except OSError:
				print ("Creation of the directory %s failed" % ChangedFolder)
				ChangedFolder = NewData

	for File in NewFiles:
		OldFile = OldData + '//' + File
		NewFile = NewData + '//' + File
		if os.path.isfile(NewFile): 
			StatusQueue.put('Checking file: ' + File)
			
			if not os.path.exists(NewFile):
				#NewAdded.append(File)
				Result.append({"File Name": File, 'Type': 'Newly added'})
				continue
			else:
				try:
					Old = load_workbook(OldFile, data_only=True)
					New = load_workbook(NewFile, data_only=True)
				except:
					Result.append({"File Name": File, 'Type': 'Broken'})
					continue
				
				OldWS = None
				NewWS = None
				if DataOnly == True:
					for OldSheet in Old:
						CurrentSheetName = OldSheet.title
						if CurrentSheetName.lower() == 'data':
							OldWS = OldSheet
					for NewSheet in New:
						CurrentSheetName = NewSheet.title
						if CurrentSheetName.lower() == 'data':
							NewWS = NewSheet		

					if OldWS == None or NewWS == None:
						NotData.append(File)
						Result.append({"File Name": File, 'Type': 'Not a Data file'})
						#print('Not a data file: ', File)
						continue
					else:
						try:
							ChangeCount = 0
							Old_Col = OldWS.max_row
							New_Col = NewWS.max_row
							if Old_Col >= New_Col:
								for row in OldWS.iter_rows():
									for cell in row:
										OldCellData = cell.value
										if OldCellData != "" and OldCellData != None:
											CellAddress = cell.column_letter + str(cell.row)
											NewCellData  = NewWS[CellAddress].value
											if OldCellData != NewCellData:
												NewWS[CellAddress].fill = my_fill
												NewWS[CellAddress].font = Font(color="FF0000", italic=True)
												ChangeCount+= 1
							else:
								for row in NewWS.iter_rows():
									for cell in row:
										NewCellData = cell.value
										if NewCellData != "" and NewCellData != None:
											CellAddress = cell.column_letter + str(cell.row)
											OldCellData  = OldWS[CellAddress].value
											if OldCellData != NewCellData:
												NewWS[CellAddress].fill = my_fill
												NewWS[CellAddress].font = Font(color="FF0000", italic=True)
												ChangeCount+= 1
						except:
							Result.append({"File Name": File, 'Type': 'Broken'})
							continue					
					if ChangeCount > 0:
						#Changed.append(File)
						Result.append({"File Name": File, 'Type': 'Change', "Amount": ChangeCount})
						
						NewName = Function_Add_Surflix(NewFile, 'Draft_Compare', 'Draft_Compare')
						
						try:			
							New.save(NewName)	
							StatusQueue.put('Compare Done, please check result file at: ' + ChangedFolder)
						except:
							StatusQueue.put("Permission denied, fail to save result file!")
						
					else:
						#NoChange.append(File)
						Result.append({"File Name": File, 'Type': 'No change'})		

				else:
					FileChange = False
					for OldSheet in Old:
						#OldWS = OldSheet.title
						CurrentSheetName = OldSheet.title
						try:
							NewWS = New.get_sheet_by_name(CurrentSheetName)
						except:		
							Result.append({"File Name": File + '-' + CurrentSheetName, 'Type': 'Not Found'})
							continue
						
						ChangeCount = 0
						try:
							for row in OldSheet.iter_rows():
								for cell in row:
									OldCellData = cell.value
									if OldCellData != "" and OldCellData != None:
										CellAddress = cell.column_letter + str(cell.row)
										NewCellData  = NewWS[CellAddress].value
										if OldCellData != NewCellData:
											NewWS[CellAddress].value = OldCellData + '>' + NewCellData
											NewWS[CellAddress].fill = my_fill
											ChangeCount+= 1	
						except:
							ChangeCount = 0
								
						if ChangeCount > 0:
							Result.append({"File Name": File + '-' + CurrentSheetName, 'Type': 'Change', "Amount": ChangeCount})
							FileChange = True

					if FileChange == True:
						NewName = Function_Add_Surflix(NewFile, 'Draft_Compare', 'Draft_Compare')
						
						try:			
							New.save(NewName)	
							StatusQueue.put('Compare Done, please check result file at: ' + ChangedFolder)
						except:
							StatusQueue.put("Permission denied, fail to save result file!")
						
					else:
						Result.append({"File Name": File, 'Type': 'No change'})					

				#if OldWS == None:
			
		else:
			print('Not a file, ', File)	
		Checked +=1
		percent = ShowProgress(Checked, TotalNewFiles)
		ProgressQueue.put(percent)				
		#ProgressQueue.put()

	for File in OldFiles:
		if os.path.isfile(File): 
			try:
				index = NewFiles.index(File)
			except:
				index = -1

			if index == -1:
				#Removed.append(File)
				Result.append({"File Name": File, 'Type': 'Removed'})

	if len (Result) > 0:
		wb = Workbook()
		ws =  wb.active
		Row = 2
		ws.cell(row=1, column=2).value = "File Name"
		ws.cell(row=1, column=3).value = "Type"
		ws.cell(row=1, column=4).value = "Value"
		for details in Result:
			if details['Type'] == 'Change':
				ws.cell(row=Row, column=2).value = details["File Name"]
				ws.cell(row=Row, column=3).value = details['Type']
				ws.cell(row=Row, column=4).value = details['Amount']
				ws.cell(row=Row, column=4).fill = my_fill
			else:
				ws.cell(row=Row, column=2).value = details["File Name"]
				#ws.cell(row=Row, column=2).fill = PatternFill(bgColor=BackgroundColor, fill_type = "solid")
				ws.cell(row=Row, column=3).value = details['Type']
			
			Row+=1

		Tab = Table(displayName="Change", ref="B1:" + 'D' + str(Row-1))
		style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
		Tab.tableStyleInfo = style
		ws.add_table(Tab)

		NewName = Function_Add_Surflix(NewData + '/Summary_Result.xlsx', 'Draft_Compare', None)
		try:			
			wb.save(NewName)	
			StatusQueue.put('Compare Done, please check result file at: ' + NewData + '/Summary_Result.xlsx' + '//Summary_Result.xlsx')
		except:
			StatusQueue.put("Permission denied, fail to save result file!")
		return
		'''
		wb.save(ChangedFolder + '//Result.xlsx')
		StatusQueue.put('Compare Done, please check result file at: ' + ChangedFolder + '//Result.xlsx')
		return
		'''
	StatusQueue.put('Compare Done, no changed found.')	
	return

###########################################################################################

###########################################################################################

# Deep Compare
def report_diff(x):
	"""Function to use with groupby.apply to highlihgt value changes."""
	return x[0] if x[0] == x[1] or pd.isna(x).all() else f'{x[0]} ---> {x[1]}'


def strip(x):
	"""Function to use with applymap to strip whitespaces in a dataframe."""
	return x.strip() if isinstance(x, str) else x


def diff_pd(old_df, new_df, idx_col):
	"""Identify differences between two pandas DataFrames using a key column.
	Key column is assumed to have only unique data
	(like a database unique id index)
	Args:
		old_df (pd.DataFrame): first dataframe
		new_df (pd.DataFrame): second dataframe
		idx_col (str|list(str)): column name(s) of the index,
		  needs to be present in both DataFrames
	"""
	# setting the column name as index for fast operations
	old_df = old_df.set_index(idx_col)
	new_df = new_df.set_index(idx_col)
	# get the added and removed rows
	old_keys = old_df.index
	new_keys = new_df.index
	if isinstance(old_keys, pd.MultiIndex):
		removed_keys = old_keys.difference(new_keys)
		added_keys = new_keys.difference(old_keys)
	else:
		removed_keys = np.setdiff1d(old_keys, new_keys)
		added_keys = np.setdiff1d(new_keys, old_keys)
	out_data = {
		'removed': old_df.loc[removed_keys],
		'added': new_df.loc[added_keys]
	}
	# focusing on common data of both dataframes
	common_keys = np.intersect1d(old_keys, new_keys, assume_unique=True)
	common_columns = np.intersect1d(
		old_df.columns, new_df.columns, assume_unique=True
	)
	new_common = new_df.loc[common_keys, common_columns].applymap(strip)
	old_common = old_df.loc[common_keys, common_columns].applymap(strip)
	# get the changed rows keys by dropping identical rows
	# (indexes are ignored, so we'll reset them)
	common_data = pd.concat(
		[old_common.reset_index(), new_common.reset_index()], sort=True
	)
	changed_keys = common_data.drop_duplicates(keep=False)[idx_col]
	if isinstance(changed_keys, pd.Series):
		changed_keys = changed_keys.unique()
	else:
		changed_keys = changed_keys.drop_duplicates().set_index(idx_col).index
	# combining the changed rows via multi level columns
	df_all_changes = pd.concat(
		[old_common.loc[changed_keys], new_common.loc[changed_keys]],
		axis='columns',
		keys=['old', 'new']
	).swaplevel(axis='columns')
	# using report_diff to merge the changes in a single cell with "-->"
	df_changed = df_all_changes.groupby(level=0, axis=1).apply(
		lambda frame: frame.apply(report_diff, axis=1))
	out_data['changed'] = df_changed

	return out_data

def Function_Create_Data(Data_Workbook, Sheet_Name, Index_Col, Case_Sensitive = True, Start_Row = 0):

	Data_Dict = {}
	xlsx = Data_Workbook
	Col_List = {}
	Col_Count = {}
	Param_List = []
	ID_Col = None
	ID_Row = None
	database = None
	Start = time.time()
	for sheet in xlsx:
		sheetname = sheet.title
	
		if sheetname == Sheet_Name:
			if not Case_Sensitive:	
				sheetname = sheetname.lower()
				Sheet_Name = Sheet_Name.lower().lower()

			ws = xlsx[sheet.title]

			for row in ws.iter_rows():
				for cell in row:
					if cell.value == Index_Col:
						ID_Col = cell.column_letter
						ID_Row = cell.row
						#print('ID row:',ID_Row)
						database = ws
						break

				if database!=  None:
					break

			if ID_Col == None:
				return None

			for cell in ws[ID_Row]:
				if cell.value not in ['', Index_Col]:
					Col_Val = cell.value
					if Col_Val == None:
						Col_Val = 'None'
					try:
						Count = Col_Count[Col_Val]
					except:
						Col_Count[Col_Val] = 1
						Count = 1
					
	
					if Col_Val not in Param_List:
						Col_List[cell.column_letter] = Col_Val
						Param_List.append(Col_Val)
					else:
						Col_Count[Col_Val] +=1

						Col_List[cell.column_letter] = str(Col_Val) + '[' + str(Count) + ']'
					'''
					param = {}
					param['Label'] = cell.value
					param['Col'] = cell.column_letter
					Col_List.append(param)
					'''

	Key_List = Col_List.keys()
	print(Col_List)
	if database == None:
		return {}
	else:
		EmptyRow = 0	
		Data_Frame = {}
		Data_Sheet = xlsx[Sheet_Name]
		MaxRow = Data_Sheet.max_row
		for Row in range(ID_Row+1, MaxRow+1):
			if EmptyRow <= 20:
				Row_Data =  Data_Sheet[Row]
				Current_ID =  Data_Sheet[ID_Col+str(Row)].value
				if Current_ID != None:
					#print('Adding ID:', Current_ID)
					Entry = {}
					#Entry['ID'] = Current_ID
					for Pair in Key_List:
						Value = Data_Sheet[Pair+str(Row)].value
						if Value in ["", None]:
							Value = '#N/A'
						Entry[Col_List[Pair]] = Value

					Data_Frame[Current_ID] = Entry
					EmptyRow = 0
				else:
					EmptyRow+=1
			else:
				break		
		'''
		
		'''
		End = time.time()
		Total = End - Start
		#print('Time spent:', Total)
		#print('Col_List', Col_List)
		#print('Data_Frame', Data_Frame)
		return 	{'Label': Col_List, 'Data':  Data_Frame}

	
	'''

			if database != None:
				for i in range(KR_Row, database.max_row): 
					KRAddress = KR_Coll + str(i+1)
					ENAddress = EN_Coll + str(i+1)
					print('KRAddress', KRAddress)
					print('ENAddress', ENAddress)
					KRCell = database[KRAddress]
					KRValue = KRCell.value
					ENCell = database[ENAddress]
					ENValue = ENCell.value
					if KRValue == None or ENValue == None or KRValue == 'KO' or ENValue == 'EN':
						continue
					elif KRValue != None and ENValue != None:
						Dict.append([KRValue, ENValue.lower()])
	print("Successfully load dictionary from: ", DictList)
	return Dict
		
	'''

def Deep_Compare_Data(Old_Data, New_Data):

	Old_Set = Old_Data['Data']
	New_Set = New_Data['Data']

	Removed = []
	Removed_Data =  {}
	Added = []
	Added_Data =  {}
	ToCheck = []
	Changed = []
	Changed_Data = {}
	Old_List = list(Old_Set.keys())

	New_List = list(New_Set.keys())



	for ID in Old_List:
		if ID not in New_List:
			#print('Add', ID, 'to Removed')
			Removed_Data[ID] = Old_Set[ID]
		else:
			#print('Add', ID, 'to toCheck')
			ToCheck.append(ID)
	

	for ID in New_List:
		if ID not in Old_List:
			#print('Add', ID, 'to Added')
			Added_Data[ID] = New_Set[ID]
	
	for ID in ToCheck:
		ChangeFlag = False
		Old_ID_Data = Old_Set[ID]

		New_ID_Data = New_Set[ID]
		
		Old_Param_List = list(Old_ID_Data.keys())

		New_Param_List = list(New_ID_Data.keys())
	
		for Par in Old_Param_List:
			#print(Par)
			if Par not in New_Param_List:
				New_ID_Data[Par] = 'Removed value --> ' + str(Old_ID_Data[Par])
				ChangeFlag = True
				
			else:
				if New_ID_Data[Par] != Old_ID_Data[Par]:
					#print('Old: ', str(Old_ID_Data[Par]))
					#print('New: ', str(New_ID_Data[Par]))
					New_ID_Data[Par] = str(Old_ID_Data[Par]) + ' --> ' + str(New_ID_Data[Par])
					ChangeFlag = True
		
		for Par in New_Param_List:
			#print(Par)
			if Par not in Old_Param_List:
				#Added.append(ID)
				New_ID_Data[Par] = 'New added --> ' + str(New_ID_Data[Par])
				ChangeFlag = True

		if ChangeFlag:
			Changed_Data[ID] = New_Set[ID]
			

	#for ID in Changed:
	#	Changed_Data[ID] = New_Set[ID]
		#Changed_Data.append({'ID':New_Set[ID]})
	#print(Changed_Data)
	Diff = {
		'Removed': Removed_Data,
		'Added': Added_Data,
		'Changed': Changed_Data
	}
	return Diff

def Function_Print_Excel(Data_Frame):

	for x in Data_Frame:
		print(x)
		print(Data_Frame[x])
		print('\n')
		#for ID in Data_Frame[x]:
		#	print(ID)
		#	print('\n')

	return


def Function_Compare_Data(
		Status_Queue, Process_Queue, Old_File, New_File, out_path, Sheet_Name, Index_Col_Name, Background_Colour, Font_Colour, **kwargs
):





	print('BackgroundColor: ', Background_Colour)
	my_color = Color(rgb=Background_Colour)
	my_fill = PatternFill(patternType='solid', fgColor=my_color)

	my_font = Font(color=Font_Colour)

	Status_Queue.put("Loading old data")
	Old_DB = load_workbook(Old_File, data_only=True)

	Status_Queue.put("Loading new data")
	New_DB = load_workbook(New_File, data_only=True)

	Status_Queue.put("Import new data")
	New_Data = Function_Create_Data(New_DB, Sheet_Name, Index_Col_Name)
	if New_Data == None:
		Status_Queue.put("Invalid New Data's structure.")
	#Function_Print_Excel(New_Data)
	
	Status_Queue.put("Import old data")
	Old_Data = Function_Create_Data(Old_DB, Sheet_Name, Index_Col_Name)
	if Old_Data == None:
		Status_Queue.put("Invalid Old Data's structure.")

	#Function_Print_Excel(Old_Data)
	Status_Queue.put("Comparing data")
	Diff_Data = Deep_Compare_Data(Old_Data, New_Data)
	Status_Queue.put("Generate compare result")

	Changed = Diff_Data['Changed']
	Removed = Diff_Data['Removed']
	Added = Diff_Data['Added']


	
	wb = Workbook()
	ws =  wb.active
	ws.title = 'Changed'
	Row = 2
	Changed_List = list(Changed.keys())
	Header = []
	for Sample in Changed:
		Header = list(Changed[Sample].keys())
		break
	Header = [Index_Col_Name] + Header
	Row = 2
	if Changed != {}:
		ws.sheet_properties.tabColor = Background_Colour
		Col = 1
		for Par in Header:
			ws.cell(row=1, column=Col).value = Par
			Col +=1


		for item in Changed:
			ws.cell(row=Row, column=1).value = item
			Col = 2	
			for Par in Changed[item]:
				Value = Changed[item][Par]
				ws.cell(row=Row, column=Col).value = Value = Changed[item][Par]
				if '-->' in str(Value) :
					ws.cell(row=Row, column=Col).fill = my_fill
					ws.cell(row=Row, column=Col).font = my_font
					#Font_Colour
				LastCell = ws.cell(row=Row, column=Col).column_letter
				Col+=1
			Row +=1

		Tab = Table(displayName="Change", ref="A1:" + LastCell + str(Row-1))
		style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
		Tab.tableStyleInfo = style
		ws.add_table(Tab)

	
	
	if Added != {}:

		wb.create_sheet('Added')
		ws = wb['Added']
		#Added_List = list(Added.keys())
		Header = []
		for Sample in Added:
			Header = list(Added[Sample].keys())
			break
		Header = [Index_Col_Name] + Header
		Row = 2

		ws.sheet_properties.tabColor = Background_Colour
		Col = 1
		for Par in Header:
			ws.cell(row=1, column=Col).value = Par
			Col +=1


		for item in Added:
			ws.cell(row=Row, column=1).value = item
			Col = 2	
			for Par in Added[item]:
				ws.cell(row=Row, column=Col).value = Added[item][Par]
				Col+=1
				LastCell = ws.cell(row=Row, column=Col).column_letter
			Row +=1
	
		Tab = Table(displayName="Change", ref="A1:" + LastCell + str(Row-1))
		style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
		Tab.tableStyleInfo = style
		ws.add_table(Tab)


	
	if Removed != {}:
		wb.create_sheet('Removed')
		#Removed_List = list(Removed.keys())
		ws = wb['Removed']
		Header = []
		for Sample in Removed:
			Header = list(Removed[Sample].keys())
			break
		Header = [Index_Col_Name] + Header
		Row = 2
		ws.sheet_properties.tabColor = Background_Colour
		Col = 1
		for Par in Header:
			ws.cell(row=1, column=Col).value = Par
			Col +=1


		for item in Removed:
			ws.cell(row=Row, column=1).value = item
			Col = 2	
			for Par in Removed[item]:
				ws.cell(row=Row, column=Col).value = Removed[item][Par]
				Col+=1
				LastCell = ws.cell(row=Row, column=Col).column_letter
			Row +=1
		
		Tab = Table(displayName="Change", ref="A1:" + LastCell + str(Row-1))
		style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
		Tab.tableStyleInfo = style
		ws.add_table(Tab)


	NewName = Function_Add_Surflix(New_File, 'Compare Result', "Deep_Compare")
	try:			
		wb.save(NewName)	
		Status_Queue.put("Done!")

	except:
		Status_Queue.put("Permission denied, fail to save result file!")
	'''
	Outputdir = os.path.dirname(New_File)
	baseName = os.path.basename(New_File)
	sourcename, ext = os.path.splitext(baseName)
	timestamp = Function_Get_TimeStamp()	
	ResultFolder = Outputdir + '/Compare Result'
	if not os.path.isdir(ResultFolder):
			try:
				os.mkdir(ResultFolder)
			except OSError:
				print ("Creation of the directory %s failed" % ChangedFolder)
				ResultFolder = Outputdir
	try:			
		wb.save(ResultFolder + '//' + sourcename + '_Result_' + str(timestamp) +'.xlsx')	
		Status_Queue.put("Done!")

	except:
		Status_Queue.put("Permission denied, fail to save result file!")
	
	'''
	'''
	#print(diff)
	with pd.ExcelWriter(out_path) as writer:
		workbook  = writer.book
		for sname, data in diff.items():
			data.to_excel(writer, sheet_name=sname)
			worksheet = writer.sheets[sname]
			#worksheet.hide_gridlines(2)
			highlight_fmt = workbook.add_format({'font_color': '#' + Font_Colour, 'bg_color':'#' + Background_Colour})
			worksheet.conditional_format('A1:ZZ1000', {'type': 'text',
													'criteria': 'containing',
													'value':'--->',
													'format': highlight_fmt})
	
	Status_Queue.put(f"Differences saved in {out_path}")
	print(f"Differences saved in {out_path}")
	'''

def Add_Sheet():
	return

def Old_Function_Compare_Excel(
		Status_Queue, Process_Queue, path1, path2, out_path, sheet_name, index_col_name, Background_Colour, Font_Colour, **kwargs
):
	Status_Queue.put("Loading old data")
	old_df = pd.read_excel(path1, sheet_name=sheet_name, **kwargs)
	Status_Queue.put("Loading new data")
	new_df = pd.read_excel(path2, sheet_name=sheet_name, **kwargs)
	Status_Queue.put("Comparing data")
	diff = diff_pd(old_df, new_df, index_col_name)
	#print(diff)
	with pd.ExcelWriter(out_path) as writer:
		workbook  = writer.book
		for sname, data in diff.items():
			data.to_excel(writer, sheet_name=sname)
			worksheet = writer.sheets[sname]
			#worksheet.hide_gridlines(2)
			highlight_fmt = workbook.add_format({'font_color': '#' + Font_Colour, 'bg_color':'#' + Background_Colour})
			worksheet.conditional_format('A1:ZZ1000', {'type': 'text',
													'criteria': 'containing',
													'value':'--->',
													'format': highlight_fmt})
	
	Status_Queue.put(f"Differences saved in {out_path}")
	print(f"Differences saved in {out_path}")

def Function_Get_TimeStamp():		
	now = datetime.now()
	timestamp = str(int(datetime.timestamp(now)))			
	return timestamp

def Function_Add_Surflix(File, SubFolder = None, Surflix = None):
	
	#print('Source:', File )
	Outputdir = os.path.dirname(File)
	#print('Outputdir:', Outputdir)
	baseName = os.path.basename(File)

	sourcename, ext = os.path.splitext(baseName)
	timestamp = Function_Get_TimeStamp()	
	if SubFolder != None:
		ResultFolder = Outputdir +'/' + SubFolder
	else:
		ResultFolder = Outputdir
	if not os.path.isdir(ResultFolder):
		try:
			os.mkdir(ResultFolder)
		except OSError:
			print ("Creation of the directory %s failed" % ChangedFolder)
			ResultFolder = Outputdir
	if Surflix != None:
		Name = ResultFolder + '/' + sourcename + "_" + Surflix +  "_" + str(timestamp) + ext	
	else:
		Name = ResultFolder + '/' + sourcename + '_' + str(timestamp) + ext
	print('Generated name: ', Name)
	return Name
###########################################################################################

def main():
	Process_Queue = Queue()
	Result_Queue = Queue()
	Status_Queue = Queue()
	Debug_Queue = Queue()
	
	MyManager = Manager()
	Default_Manager = MyManager.list()
	
	root = Tk()
	My_Queue = {}
	My_Queue['Process_Queue'] = Process_Queue
	My_Queue['Result_Queue'] = Result_Queue
	My_Queue['Status_Queue'] = Status_Queue
	My_Queue['Debug_Queue'] = Debug_Queue

	My_Manager = {}
	My_Manager['Default_Manager'] = Default_Manager

	Document_Utility(root, Queue = My_Queue, Manager = My_Manager,)
	root.mainloop()  


if __name__ == '__main__':
	if sys.platform.startswith('win'):
		multiprocessing.freeze_support()

	main()
