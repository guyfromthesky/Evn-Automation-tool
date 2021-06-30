from ppadb.client import Client as AdbClient
import os
import cv2
import numpy as np

Root = os.getcwd()

from openpyxl import load_workbook, worksheet, Workbook








Project = "V4"
Features = ['Gacha - Companion', 'Call cheat']

	
def Function_Import_DB(DB_Path, List_Sheet = []):
	
	if DB_Path != None:
		if (os.path.isfile(DB_Path)):
			xlsx = load_workbook(DB_Path, data_only=True)
			MyDB = {}
			#Entry = {}

			for sheet in xlsx:
				print('Checking sheet: ', sheet)
				sheetname = sheet.title.lower()
				if sheetname in List_Sheet or len(List_Sheet) == 0:	
					DB_Name = sheetname
					
					Col_StringID = ""
					Col_String_EN = ""
					Col_String_KR = ""
					Col_Path = ""

					ws = xlsx[sheet.title]

					database = None
					ListCol = {}

					#Get Col Label and Letter
					for row in ws.iter_rows():
						for cell in row:
							if cell.value == "StringID":
								Col = cell.column_letter
								Row_ColID = cell.row
								Col_StringID = Col
								ListCol['StringID'] = Col_StringID

							if Col_StringID != "":
								database = ws
								lastChar = Col_StringID
								
								while True:
									lastChar = chr(ord(lastChar) + 1)
									try:
										ColLabel = database[lastChar + str(Row_ColID)].value
									except:
										break	
									if ColLabel in ["",None] :
										break
									else:
										ListCol[ColLabel] = lastChar
			
						if database!=  None:
							break		
					# Load data 			
					if database != None:
						for i in range(Row_ColID, database.max_row): 
							StringID = database[Col_StringID + str(i+1)].value
							MyEntry = {}
							for Label in ListCol:
								MyEntry[Label] = database[ListCol[Label] + str(i+1)].value
							MyDB[StringID] = MyEntry
			return MyDB
		else:
			return({})	
	else:
		return({})

class CallCheat:
    # Serial: Device's serial
    # DB: Database's Path.
	def __init__(self, Serial, DB):

		self.Project = "V4"
        self.Client = AdbClient(host="127.0.0.1", port=5037)
        self.Device = client.device(Serial)

    def Enable_MH(self):

    def Enable_Console(self):