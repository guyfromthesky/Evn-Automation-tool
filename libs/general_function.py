# Level 3 class
# Upper level is testscript

import os, sys
import cv2
import numpy as np

from openpyxl import load_workbook
from openpyxl.styles import Alignment

import pytesseract

from datetime import datetime
import time
import imutils

CWD = os.path.abspath(os.path.dirname(sys.argv[0]))
import inspect

import easyocr
import cv2
import numpy as np

################################################################################################################


def Function_Import_TestCase(TestCase_Path):

	if TestCase_Path != None:
		if (os.path.isfile(TestCase_Path)):
			xlsx = load_workbook(TestCase_Path, data_only=True)
			Testcase = {}

			for sheet in xlsx:
				
				sheetname = sheet.title.lower()
				if sheetname  == 'testcase':
					FirstInfoRow = None
					EndInfoRow = None

					FirstTestCaseRow = None
					EndTestCaseRow = None

					ws = xlsx[sheet.title]

					Testcase['Info'] = {}
					Testcase['Testcase'] = []
					database = None
					ListCol = {}

					Loop = False
					LoopStep = []

					#Get Col Label and Letter
					for col in ws.iter_cols(min_row=1, max_col=1):
						for cell in col:
							Value = cell.value
							if Value != None:
								Value = Value.lower()
								CurrentRow = cell.row
								
								if Value == 'info':
									if FirstInfoRow == None:
										FirstInfoRow = CurrentRow
									if EndInfoRow == None or EndInfoRow < CurrentRow:
										EndInfoRow = CurrentRow
									if CurrentRow > FirstInfoRow:
										Par = ws['B' + str(CurrentRow)].value
										Val = ws['C' + str(CurrentRow)].value
										Testcase['Info'][Par] = Val
									

								if Value == 'test case':
									if FirstTestCaseRow == None:
										FirstTestCaseRow = CurrentRow
									if EndTestCaseRow == None or EndTestCaseRow < CurrentRow:
										EndTestCaseRow = CurrentRow
									
									
									if CurrentRow > FirstTestCaseRow:
										Step = {}
										Function = ws['C' + str(CurrentRow)].value	
										Step['Name'] = Function
										Step['Argument'] = []
										lastChar = "D"
										while True:		
											try:
												Val = ws[lastChar + str(CurrentRow)].value
												if Val != None:
													Step['Argument'].append(Val)
													lastChar = chr(ord(lastChar) + 1)
												else:
													break
											except:
												break
										
										Type = ws['B' + str(CurrentRow)].value
										#print('Type', Type)
										if Type.find('Loop') > -1:
											LoopStep.append(Step)
											#print('Add loop step', Step)
										else:

											if len(LoopStep) > 0:
												#LoopStep.append(Step)
												LoopSteps = {}
												LoopSteps['Name'] = 'Loop'
												Type = ws['B' + str(CurrentRow-1)].value
												Type = Type.replace('Loop(',"")
												Type = Type.replace(')',"")

												LoopSteps['Amount'] = int(Type)
												LoopSteps['Step'] = LoopStep
												
												Testcase['Testcase'].append(LoopSteps)
												Testcase['Testcase'].append(Step)
												LoopStep = []
												#print('Update loop step', LoopSteps)
											else:
												Testcase['Testcase'].append(Step)
												#print('Add normal step: ', Step)

										CurrentRow +=1
										
						if len(LoopStep) > 0:
							LoopSteps = {}
							LoopSteps['Name'] = 'Loop'
							Type = ws['B' + str(CurrentRow-1)].value
							Type = Type.replace('Loop(',"")
							Type = Type.replace(')',"")

							LoopSteps['Amount'] = int(Type)
							LoopSteps['Step'] = LoopStep
							
							Testcase['Testcase'].append(LoopSteps)

			if 'Type' not in Testcase['Info']:
				Testcase['Info']['Type'] = 'General'
			return Testcase
		else:
			return({})	
	else:
		return({})

def Function_Import_Data(TestCase_Path, Data_ID):

	Data_ID = Data_ID.lower()
	if TestCase_Path != None:
		if (os.path.isfile(TestCase_Path)):
			xlsx = load_workbook(TestCase_Path, data_only=True)
			#Entry = {}
			Data = []
			for sheet in xlsx:
				
				sheetname = sheet.title.lower()
				print('sheetname:', sheetname)
				if sheetname.find('data_')  > -1:
					DataName = sheetname.replace('data_', '')
					if DataName ==  Data_ID:
						FirstDataRow = None
						EndDataRow = None

						ws = xlsx[sheet.title]
						#Get Col Label and Letter
						for col in ws.iter_cols(min_row=1, max_col=1):
							for cell in col:
								Value = cell.value
								if Value != None:
									Value = Value.lower()
									print('Value:', Value)
									CurrentRow = cell.row
									
									if Value == 'stringid':
										FirstInfoRow = CurrentRow
									else:
										Data.append(Value)

						return Data

		else:
			return([])		
	else:
		return([])		

def Print_Result(TestCase_Path, Result_Array, Result_Patch):
	if TestCase_Path != None:
		if (os.path.isfile(TestCase_Path)):
			
			xlsx = load_workbook(TestCase_Path, data_only=True)
			
			Testcase = {}
			#Entry = {}

			for sheet in xlsx:
				
				sheetname = sheet.title.lower()
				if sheetname  == 'result':

					FirstResultRow = None
					EndResultRow = None

					ws = xlsx[sheet.title]

					#Get Col Label and Letter
					for col in ws.iter_cols(min_row=1, max_col=1):
						for cell in col:
							Value = cell.value
							if Value != None:
								Value = Value.lower()
								CurrentRow = cell.row
								
								if FirstResultRow == None:
									FirstResultRow = CurrentRow
					
					Current_ResultRow = FirstResultRow + 1
					for Result in Result_Array:
						List = Result
						if 'Name' in Result:
							Step = str(Result['Name'])
						else:
							continue
						
						if 'Status' in Result:
							TestResult = str(Result['Status'])
						else:
							continue	

						if 'Details' in Result:
							Note = str(Result['Details'])
						else:
							Note = ""
						
						Comment = Result.get('Key', "")
						
						ws['B' + str(Current_ResultRow)].value = Step
						ws['C' + str(Current_ResultRow)].value = TestResult
						ws['D' + str(Current_ResultRow)].value = Note
						ws['D' + str(Current_ResultRow)].alignment = Alignment(wrap_text=True)
						Current_ResultRow+=1
						
	xlsx.save(Result_Patch)
	return


def Function_Import_DB(DB_Path, List_Sheet = [], List_Item = []):

	if DB_Path != None:
		if (os.path.isfile(DB_Path)):
			xlsx = load_workbook(DB_Path, data_only=True)
			MyDB = {}
			#Entry = {}

			for sheet in xlsx:
				
				sheetname = sheet.title
				if sheetname in List_Sheet or len(List_Sheet) == 0:	
					print('Checking sheet: ', sheet)
					DB_Name = sheetname
					Col_StringID = ""
					Col_String_EN = ""
					Col_String_KR = ""
					Col_Path = ""

					ws = xlsx[sheet.title]

					database = None
					ListCol = {}

					#Get Col Label and Letter
					for row in ws.iter_rows():
						for cell in row:
							if cell.value == "StringID":
								Col = cell.column_letter
								Row_ColID = cell.row
								Col_StringID = Col
								ListCol['StringID'] = Col_StringID

							if Col_StringID != "":
								database = ws
								lastChar = Col_StringID
								
								while True:
									lastChar = chr(ord(lastChar) + 1)
									try:
										ColLabel = database[lastChar + str(Row_ColID)].value
									except:
										break	
									if ColLabel in ["",None] :
										break
									else:
										ListCol[ColLabel] = lastChar

								
						if database!=  None:
							break		
					# Load data 			
					if database != None:
						for i in range(Row_ColID, database.max_row): 
							StringID = database[Col_StringID + str(i+1)].value
							if StringID in List_Item or len(List_Item) == 0:
								print('Adding StringID to pool:', StringID)
								MyEntry = {}
								for Label in ListCol:
									if Label == 'Path':
										Path = database[ListCol[Label] + str(i+1)].value
										try:
											Path = Correct_Path(Path)
										except:
											Path = None	
										if Path != None:
											if os.path.isfile(Path):		
												MyEntry['Image'] = read_img(Path)
												MyEntry[Label] = database[ListCol[Label] + str(i+1)].value

										
									else:
										MyEntry[Label] = database[ListCol[Label] + str(i+1)].value
								if 'Image' in MyEntry:
									MyDB[StringID] = MyEntry
					else:
						return({})	
			return MyDB
		else:
			return({})	
	else:
		return({})

################################################################################################################
# Sleep for an amount of time
def Sleep(total_miliseconds):
	time.sleep(total_miliseconds/1000)

def tap(Device, x, y):
	command = "input tap " + str(x) + " " + str(y)
	print('Tap:', command)
	Device.shell(command)
	return

def tap_location(Device, location):
	command = "input tap " + str(location['x']) + " " + str(location['y'])
	Device.shell(command)
	return

#Tap on Location_Object
def tap_object(Device, Location_Object):
	tap(Device, Location_Object['x'], Location_Object['y'])
	return

def read_img(img_path):
	return cv2.imdecode(np.fromfile(img_path, dtype=np.uint8), cv2.IMREAD_UNCHANGED)

'''
def Four_Touch():
	points = [(500,500), (1000, 500), (1000, 1000), (500, 1000)]
	th = TouchActionBuilder()
	th.multifinger_ntap(points, 1, 250).execute_and_reset()
	del th
	return
def Three_Touch():
	points = [(500,500), (1000, 500), (1000, 1000)]
	th = TouchActionBuilder()
	th.multifinger_ntap(points, 1, 250).execute_and_reset()
	del th
	return


'''

def swipe(Device, x1, y1, x2, y2):
	command = "input swipe " + str(x1) + " " + str(y1) + " " + str(x2) + " " + str(y2)
	Device.shell(command)
	return


#Swipe from A -> B
def swipe_object(Device, Location_Object_A, Location_Object_B):
	swipe(Device, Location_Object_A['x'], Location_Object_A['y'], Location_Object_B['x'], Location_Object_B['x'])
	return


#Swipe up
def swipe_up(Device, Location_Object_A, Range):
	Location_Object_B = {'x': Location_Object_A['x'], 'y': Location_Object_A['y'] - Range} 
	command = "input swipe " + str(Location_Object_A['x']) + " " + str(Location_Object_A['y']) + " " + str(Location_Object_B['x']) + " " + str(Location_Object_B['y'])
	Device.shell(command)
	return

def send_text(Device, Text):
	command = "input text \'%s\'" %Text
	Device.shell(command)
	return

def send_key(Device, KeyID):
	command = "input keyevent \'%s\'" %KeyID
	print('Command: ', command)
	Device.shell(command)
	return

def save_screenshot(Device, Name):

	Image = Device.screencap()
	Img_Name = Correct_Path(Name + '_' + Function_Get_TimeStamp() + '.png', 'Screenshot')
	with open(Img_Name, "wb") as fp:
		fp.write(Image)
	return	

def save_image(Img, Name):
	Img_Name = Correct_Path(Name + '_' + Function_Get_TimeStamp() + '.png', 'Image')
	cv2.imwrite(Img_Name, Img) 
	return

################################################################################################################

def resize(Img, scale_percent):

	width = int(Img.shape[1] * scale_percent)
	height = int(Img.shape[0] * scale_percent)
	dim = (width, height)
	
	# resize image
	resized = cv2.resize(Img, dim, interpolation = cv2.INTER_AREA)
	return resized

def Draw_Line(Img, Location_Object_A, Location_Object_B):
	cv2.line(Img,(Location_Object_A['x'],Location_Object_A['y']),(Location_Object_B['x'],Location_Object_B['y']),(0,0,255),2)
	return Img

def Search_Best_Match(Img_Screenshot, Img_Template, Match_Rate=0.9):

	Start = time.time()
	template = cv2.cvtColor(Img_Template, cv2.COLOR_BGR2GRAY)
	template = cv2.Canny(template, 50, 200)
	(tH, tW) = template.shape[:2]
	image = np.asarray(Img_Screenshot)
	gray = cv2.imdecode(image, cv2.COLOR_BGR2GRAY)
	Found = None
	for scale in np.linspace(0.2, 1.0, 20)[::-1]:
		resized = imutils.resize(gray, width = int(gray.shape[1] * scale))
		r = gray.shape[1] / float(resized.shape[1])

		if resized.shape[0] < tH or resized.shape[1] < tW:
			break
		edged = cv2.Canny(resized, 50, 200)
		result = cv2.matchTemplate(edged, template, cv2.TM_CCOEFF)
		(_, maxVal, _, maxLoc) = cv2.minMaxLoc(result)
		if Found is None or maxVal > Found[0]:		
			Found = (maxVal, maxLoc, r)
	(_, maxLoc, r) = Found
	(startX, startY) = (int(maxLoc[0] * r), int(maxLoc[1] * r))
	(endX, endY) = (int((maxLoc[0] + tW) * r), int((maxLoc[1] + tH) * r))
	
	End = time.time()
	Message = 'Total time spend for ' + inspect.stack()[0][3] + ": " + str(End-Start) + ' miliseconds.'
	print(Message)
	
	if Found != None:
		Loc = {"x": int((maxLoc[0] + 0.5 * tW) * r), "y": int((maxLoc[1] + 0.5 * tH) * r), "w": int(abs(startX-endX)), "h": int(abs(startY-endY))}	
		
		return Loc
	else:
		return False

def Search_All_Object(Img_Screenshot, Img_Template, Match_Rate=0.9):

	Start = time.time()
	print('Img_Template', Img_Template)
	template = cv2.imread(Img_Template)
	#template = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
	#template = cv2.Canny(template, 50, 200)
	##cv2.imshow('',template )
	#cv2.waitKey(0)
	#template = cv2.cvtColor(template, cv2.IMREAD_COLOR)
	template = cv2.Canny(template, 50, 200)
	#template  = cv2.imdecode(template , cv2.COLOR_BGR2GRAY)
	(tH, tW) = template.shape[:2]
	
	#image = np.asarray(Img_Screenshot)
	#gray = cv2.imdecode(Img_Screenshot, cv2.IMREAD_COLOR)
	gray = cv2.cvtColor(Img_Screenshot, cv2.COLOR_BGR2GRAY)
	#cv2.imshow('',gray )
	#cv2.waitKey(0)
	Found = None
	Results = []
	for scale in np.linspace(0.2, 1.0, 20)[::-1]:
		resized = imutils.resize(gray, width = int(gray.shape[1] * scale))
		r = gray.shape[1] / float(resized.shape[1])
		if resized.shape[0] < tH or resized.shape[1] < tW:
			break
		edged = cv2.Canny(resized, 50, 200)
		result = cv2.matchTemplate(edged, template, cv2.TM_CCOEFF_NORMED)
		(_, maxVal, _, maxLoc) = cv2.minMaxLoc(result)

		if Found is None or maxVal > Found[0]:
			Found = (maxVal, maxLoc, r)
			if maxVal > Match_Rate:

				(startX, startY) = (int(maxLoc[0] * r), int(maxLoc[1] * r))
				(endX, endY) = (int((maxLoc[0] + tW) * r), int((maxLoc[1] + tH) * r))
				cv2.rectangle(gray, (startX, startY), (endX, endY), (0, 0, 255), 2)
				Loc = {"x": int((maxLoc[0] + 0.5 * tW) * r), "y": int((maxLoc[1] + 0.5 * tH) * r), "w": int(abs(startX-endX)), "h": int(abs(startY-endY))}
				Results.append(Loc)
	
	End = time.time()
	Message = 'Total time spend for ' + inspect.stack()[0][3] + ": " + str(End-Start) + ' miliseconds.'
	print(Message)

	if len(Results) > 0:
		print('Found')
		return True
	else:
		return False	


def Get_Item(Img_Screenshot, Img_Template, Match_Rate=0.50):

	(source_H, _source_W) = Img_Screenshot.shape[:2]
	
	ratio = 1080 / source_H
	if ratio != 1:
		Img_Screenshot = resize(Img_Screenshot, ratio)

	template = cv2.cvtColor(Img_Template, cv2.COLOR_BGR2GRAY)
	(tH, tW) = template.shape[:2]

	gray = cv2.cvtColor(Img_Screenshot, cv2.COLOR_BGR2GRAY)

	Found = None
	Loc = None
	Results = []
	#for scale in np.linspace(0.2, 1.0, 20)[::-1]:
	result = cv2.matchTemplate(gray, template, cv2.TM_CCOEFF_NORMED)
	(_, maxVal, _, maxLoc) = cv2.minMaxLoc(result)

	if Found is None or maxVal > Found[0]:
		Found = (maxVal, maxLoc)
		if maxVal >= Match_Rate:
			
			(startX, startY) = (int(maxLoc[0]), int(maxLoc[1]))
			(endX, endY) = (int((maxLoc[0] + tW)), int((maxLoc[1] + tH)))
			Loc = {"x": int((maxLoc[0] + 0.5 * tW)), "y": int((maxLoc[1] + 0.5 * tH)), "w": int(abs(startX-endX)), "h": int(abs(startY-endY))}

	if Loc != None:
		print('Loc', Loc)
		for key in Loc:
			Loc[key] /= ratio
		print('New Loc', Loc)
		return Loc
	else:
		return False


def Count_Object(Img_Screenshot, Img_Template, Match_Rate=0.50):

	Img_Template = resize(Img_Template, 50)
	#Img_Template = cv2.bitwise_not(Img_Template)

	template = cv2.cvtColor(Img_Template, cv2.COLOR_BGR2GRAY)
	#template = cv2.Canny(template, 50, 200)
	(tH, tW) = template.shape[:2]
	#image = cv2.bitwise_not(Img_Screenshot)
	gray = cv2.cvtColor(Img_Screenshot, cv2.COLOR_BGR2GRAY)
	#cv2.imshow("template", template)
	#cv2.waitKey(0)

	matches = []
	Counter = 0
	for scale in np.linspace(0.2, 1.0,20)[::-1]:
		# resize the image according to the scale, and keep track
		# of the ratio of the resizing
		resized = imutils.resize(gray, width = int(gray.shape[1] * scale))
		r = gray.shape[1] / float(resized.shape[1])
		# if the resized image is smaller than the template, then break
		# from the loop
		if resized.shape[0] < tH or resized.shape[1] < tW:
			break
		# detect edges in the resized, grayscale image and apply template
		# matching to find the template in the image
		#edged = cv2.Canny(resized, 50, 200)
		res = cv2.matchTemplate(resized, template,cv2.TM_CCOEFF_NORMED)
		#print("res", res)
		loc = np.where( res >= Match_Rate)
		
		for pt in zip(*loc[::-1]):
			#sensitivity= 100
			#print(round(pt[0]/sensitivity), round(pt[1]/sensitivity))
			intersection = 0
			for match in matches:

				x1 = int(match[0] + 0.5 * tW)
				y1 = int(match[1] + 0.5 * tH)
				x2 = int(pt[0] + 0.5 * tW)
				y2 = int(pt[1] + 0.5 * tH)
				if Duplicated((x1, y1), (x2, y2), tW, tH):
					intersection = 1
					break

			if intersection == 0:
				matches.append(pt)
				Counter+=1
				cv2.rectangle(Img_Screenshot, pt, (pt[0] + tW, pt[1] + tH), (0, 255, 0), 2)
				print("Adding item: ",pt, (pt[0] + tW, pt[1] + tH))
	#cv2.imshow('Img_Screenshot', Img_Screenshot)
	#cv2.waitKey(0)
	
	if Counter > 0:
		return Counter, Img_Screenshot
	else:
		return False, Img_Screenshot	

def Duplicated(center1, center2, tw, th):
	
	if abs(center1[0] - center2[0]) <= tw:
		if abs(center1[1] - center2[1]) <= th:
			#print('Dup: ',  center1, center2, tw, th)
			return True	
	#print('Not Dup: ',  center1, center2, tw, th)		
	return False

def intersected(bottom_left1, top_right1, bottom_left2, top_right2):
    if top_right1[0] < bottom_left2[0] or bottom_left1[0] > top_right2[0]:
        return 0
    if top_right1[1] < bottom_left2[1] or bottom_left1[1] > top_right2[1]:
        return 0
    return 1

################################################################################################################

def Correct_Path(path, Folder = 'DB'):
	print("Folder", Folder)
	print('Path', path )
	if not os.path.isdir(Folder):
		try:
			os.mkdir(Folder)
		except OSError:
			return False
	return Folder + '//' + path

def Init_Folder(FolderPath):
	if not os.path.isdir(FolderPath):
		try:
			os.mkdir(FolderPath)
			print('Create new folder:', FolderPath)
		except OSError:
			print ("Creation of the directory %s failed" % FolderPath)

def Function_Get_TimeStamp():		
	now = datetime.now()
	timestamp = str(int(datetime.timestamp(now)))			
	return timestamp

def Split_Path(Path):
	Outputdir = os.path.dirname(Path)
	baseName = os.path.basename(Path)
	sourcename, ext = os.path.splitext(baseName)
	return [Outputdir, sourcename, ext]

################################################################################################################

def get_text_from_image_old(tess_path, tess_language, tess_data, input_image):
	pytesseract.pytesseract.tesseract_cmd = tess_path
	advanced_tessdata_dir_config = '--psm 7 --tessdata-dir ' + '"' + tess_data + '"'
	ocr = pytesseract.image_to_string(input_image, lang = tess_language, config=advanced_tessdata_dir_config)
	ocr = ocr.replace("\n", "").replace("\r", "").replace("\x0c", "")
	return ocr

def get_text_from_image(tess_language, input_image, export_path):
	#print('Local:', locals())
	reader = easyocr.Reader([tess_language], gpu=False)
	#image = cv2.imread(input_image)
	results = reader.readtext(input_image, paragraph="False")[-1]
	print('Scan result:', results)
	f = open(export_path, "a")
	f.write(str(results) + '\n')
	f.close()
	return results