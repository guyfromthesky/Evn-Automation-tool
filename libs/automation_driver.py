from ppadb.client import Client as AdbClient
import os, sys
import cv2
import numpy as np
import time
import json

from datetime import datetime
import time
import pytesseract

#from libs.general_function import *

import queue 

#cwd = os.path.dirname(os.path.realpath(__file__))
CWD = os.path.abspath(os.path.dirname(sys.argv[0]))

import csv
from openpyxl import load_workbook

import importlib.machinery
import importlib.util


# Return result format:
# {
# 	"Type": Execute/Result/..,
#	"Status": True/False,
#	"Details": {},
#	"Screenshot": Array,
# }

class Automation:
	# Serial: Device's serial
	# DB: Database's Path.
	def __init__(self, Status_Queue, Resolution = '1080', Language = 'en', Tess_Path = None, Tess_Data = None, Serial = None, DB_Path = None, Result_Folder_Path = None, Module_Path = None):
		self.Debugger = Status_Queue
		
		self.Client = AdbClient(host="127.0.0.1", port=5037)

		

		self.Gacha_Pool = []
		self.Execution_List = []
		self.Current_Value = None
		
		self.action_list = []
		
		self.Item_List = []
		self.Result_Array = []
		
		self.Result_Path = Result_Folder_Path

		self.tess_path = Tess_Path
		self.tess_data = Tess_Data
		self.tess_lang = Language
		if Tess_Path != None and Tess_Data != None and Language != None:
			self.OCR = True
		else:
			self.OCR = False
		self.LoopList = False
		self.Resolution = Resolution
		self.Ratio = 1
		print('Serial', Serial)
		print('Client', self.Client)
		if Serial != None:	
			self.Serial = Serial
			if self.Client != None:	
				self.Device = self.Client.device(Serial)
				if self.Device != None:
					self.Ratio = self.Get_Ratio()
			else:
				self.Device = None
		else:
			self.Device = None

		if DB_Path != None:
			self.Function_Import_DB(DB_Path)

		self.Update_Action_List()

		if Module_Path != None:
			self.Function_Load_Module(Module_Path)

	def append_action_list(self, type = None, name = None, argument = [], description = ''):
	
		_action = {}
		_action['type'] = type
		_action['name'] = name
		_action['arg'] = argument
		_action['description'] = description

		self.action_list.append(_action)
		return _action

	def Function_Parse_Data(self, data_type, data_value):
		print('Locals:', locals())
		if data_type == 'int':
			parsed_value = int(data_value)
		elif data_type in ['point', 'area']:
			parsed_value = json.loads(data_value)
		elif data_type == 'float':
			parsed_value = float(data_value)	
		else:
			parsed_value = data_value
		return parsed_value


	def Function_Execute_Block(self, Status_Queue, Progress_Queue, Pause_Status, Block):

		'''
		{
			'type': 'condition',
			'condition_string': condition_string,
			'execute_block': block,
			'current_list_value': []
			'else_block': else_block
		}
			
		'''
		last_status = 'Running'
		for chain in Block:
			while True:
				Sleep(20)
				try:
					Status = Pause_Status.get(0)
					if Status == 'Pause':
						last_status = 'Pause'
						continue
					elif Status == 'Resume':
						last_status = 'Running'
						break
					else:
						last_status = 'Pause'
						print('Dont know: ', Status)
						continue
				except queue.Empty:
					if last_status == 'Running':
						break
					else:
						continue

			_block_type = chain['type']
			Code_Block = chain['execute_block']
			_current_list_value = chain['current_list_value']
			
			if len(_current_list_value) > 0:
				self.Execution_Value = _current_list_value
			else:
				self.Execution_Value = []
		
			if _block_type == "action":
				
				kwarg = {}
				_test_name = Code_Block['name']
			
				_arg_list = Code_Block['arg']
				function_object = getattr(self, _test_name)
				if len(_arg_list) > 0:	
					for _temp_arg in _arg_list:
						#print('_temp_arg', _temp_arg)
						_value_type = _temp_arg['type']
						_raw_value = _temp_arg['value']
						if _raw_value != '':
							kwarg[_temp_arg['name']] = self.Function_Parse_Data(_value_type,_raw_value)
						
				Status_Queue.put('Execute action: ' + str(_test_name))
				#Status_Queue.put('Execute value: ' + str(self.Execution_Value))
				result = function_object(**kwarg)
				Status_Queue.put('Execute result: '+ _test_name + '->' + str(result))
			elif _block_type == "condition":
				_condition_string = chain['condition_string']
				print('Checking condition:', _condition_string)
				try:
					_check_condition = eval(_condition_string)
				except: 
					_check_condition = False
				if _check_condition == True:
					self.Function_Execute_Block(Status_Queue, Progress_Queue, Pause_Status, Code_Block)
			elif _block_type == "Comment":
				#Status_Queue.put('Execute result: '+ _test_name + '->' + str(result))
				continue

	def Function_Load_Module(self, Module_Path):
		# Need to run again in main Automation lib
		for file_path in Module_Path:
			loader = importlib.machinery.SourceFileLoader( 'mymodule', file_path)
			spec = importlib.util.spec_from_loader( 'mymodule', loader )
			mymodule = importlib.util.module_from_spec( spec )
			
			loader.exec_module( mymodule)
			all_functions = dir(mymodule)
			new_functions = []
			for mymodule_name in all_functions:
				if not mymodule_name.startswith('__'):
					new_functions.append(mymodule_name)
			for function in new_functions:
				real_f = getattr(mymodule, function)
				a = real_f()
				setattr(self, real_f.__name__, getattr(a, 'func'))
				self.append_action_list(type = 'Action', name = real_f.__name__ , argument = a.kwarg, description= '')
			
			self.append_action_list(type = 'Action', name = 'Send_Enter_Key', argument = None, 
				description= 'Press ENTER key on the phone, which will move the focus to the next widget.')



	def Function_Generate_TestCase(self, TestCase_Object, Execution_List, _current_execution_value = None, recursion_type = None, start_index = None):

		test_case_list = []
	
		_index = -1
		for index in range(0, len(TestCase_Object)):
			#print('index', index, '_index', index, 'start_index', start_index)
			if index < _index:
				continue
			else:
				_index = index
			
			
			test_object = TestCase_Object[index]

			if start_index != None:
				if index < start_index:
					continue
			elif index == len(TestCase_Object):
				return test_case_list,index
			else:
				pass
			test_object = TestCase_Object[index]

			_test_type = test_object['type']
			_test_name = test_object['name']
			_arg_list = test_object['arg']
			print('Action:', _test_name)
			if _test_type == "Loop":
				# Loop number/list
				_test_name = test_object['name']
				if _test_name == 'Loop':
					
					arg = _arg_list[0]
					_loop_amount = int(arg['value'])
					print('Enter a normal loop',_loop_amount )
				
					if 	_loop_amount > 0:
						_loop_block_chain = []	
						_current_loop_index = _index
						#print('Index', _index, index)
						while True:
							_current_loop_index+=1
							if _current_loop_index >= len(TestCase_Object):
								break
							_temp_loop_step = TestCase_Object[_current_loop_index]
							_temp_test_type = _temp_loop_step['type']
							_temp_test_name = _temp_loop_step['name']	
							if _temp_test_type == 'Loop':
								if _temp_test_name == 'End Loop':
									print('End of a loop')
									break
									#return _loop_block_chain
								elif _temp_test_name in ['Loop', 'Loop List']:
									# A loop within a loop:
									print('Generate a sub-loop within a loop.', _current_loop_index)
									_loop_block, _current_loop_index = self.Function_Generate_TestCase(TestCase_Object, Execution_List, _current_execution_value, 'loop', _current_loop_index)	
									
									_loop_block_chain = _loop_block_chain + _loop_block
								else:
									continue	
							elif _temp_test_type == 'Condition':
								# Condition within a loop:
							
								print('Generate a condition within a loop.', _current_loop_index)
								_condition_Block,_current_loop_index = self.Function_Generate_TestCase(TestCase_Object, Execution_List, _current_execution_value, 'condition',  _current_loop_index)	
							
								_loop_block_chain = _loop_block_chain + _condition_Block
							else:
								_chain = self.chain_warpped('action', _temp_loop_step)
								_loop_block_chain += _chain
							
							#test_case_list += _loop_block_chain
						print('Leng of the normal loop', len(_loop_block_chain))	
						for i in range(0, _loop_amount):
							test_case_list += _loop_block_chain
						_index = _current_loop_index+1
						if recursion_type == 'loop':
							return test_case_list, _current_loop_index
					else:
						print('Just a blank loop')
						_index = _current_loop_index+1
				
				elif _test_name == 'Loop List':
					print('Enter a list loop' )
					_loop_start_index = 0
					_loop_end_index = len(Execution_List)-1
					for arg in _arg_list:
						
						if arg['name'] == 'start_index':
							#Normal loop
							_loop_start_index = int(arg['value'])
							continue
						elif arg['name'] == 'end_index':
							if _loop_end_index > int(arg['value']):
								_loop_end_index = int(arg['value'])
							continue
						else:
							print('Invalid arg config')

					print('Start:', _loop_start_index, 'End:', _loop_end_index)
					
					_loop_amount = _loop_end_index - _loop_start_index + 1
					print('_loop_amount', _loop_amount)
					if 	_loop_amount > 0:
						
						for loop_indexer in range(0, _loop_amount):
							print('Indexer of loop:', loop_indexer)
							_loop_block_chain = []	
							_current_loop_index = _index
							_current_execution_value = Execution_List[loop_indexer]
							print('Current execute value:', _current_execution_value)
							while True:
								_current_loop_index+=1
								if _current_loop_index >= len(TestCase_Object):
									break
								_temp_loop_step = TestCase_Object[_current_loop_index]
								_temp_test_type = _temp_loop_step['type']
								_temp_test_name = _temp_loop_step['name']	
								if _temp_test_type == 'Loop':
									if _temp_test_name == 'End Loop':
										print('End of a loop', _current_loop_index)
										break
										#return _loop_block_chain
									elif _temp_test_name  in ['Loop', 'Loop List']:
										# A loop within a loop:
										
										print('Generate a sub-loop within a loop list.', _current_loop_index)
										_loop_block, _current_loop_index = self.Function_Generate_TestCase(TestCase_Object, Execution_List,_current_execution_value, 'loop', _current_loop_index)	
										_loop_block_chain = _loop_block_chain + _loop_block
									else:
										continue	
								elif _temp_test_type == 'Condition':
									# Condition within a loop:
								
									print('Generate a sub-condition within a loop list.', _current_loop_index)
									_condition_Block,_current_loop_index = self.Function_Generate_TestCase(TestCase_Object, Execution_List, _current_execution_value, 'condition',  _current_loop_index)	
									
									_loop_block_chain = _loop_block_chain + _condition_Block
								else:
									_chain = self.chain_warpped('action', _temp_loop_step, current_list_value= _current_execution_value)
									_loop_block_chain += _chain
							print('Leng of the list loop', len(_loop_block_chain))		
							test_case_list += _loop_block_chain
							#End of loop
							
						_index = _current_loop_index+1
						if recursion_type == 'loop':
							return test_case_list, _index
					else:
						print('Just a blank loop')
						_index = _current_loop_index

				elif _test_name in ['End Loop', 'End If']:
					print('End of a loop/condition')
					if recursion_type in ['loop', 'condition']:
						return test_case_list, index+1
					break
				else:
					_chain = self.chain_warpped('action', test_object)
					test_case_list += _chain
			elif _test_type == "Condition":
				print('Enter a condition',index)
				# Condition type
				'''
				{
					'type': 'condition',
					'condition_string': condition_string,
					'execute_block': test_case_block,
					'fail_block': fail_test_case_block
				}
				'''
				_condition_block_chain = []
				# Only support If condition for now.
				#print('Condition:', test_object['arg'])
				_condition_string = True
				for arg in test_object['arg']:
					if arg['name'] == 'condition':
						#Normal loop
						_condition_string = arg['value']
						break
					
				_current_condition_index = _index
				while True:
					_current_condition_index+=1
					#print('_current_condition_index', _current_condition_index, len(TestCase_Object))
					if _current_condition_index > len(TestCase_Object):
						break
					_temp_test_object = TestCase_Object[_current_condition_index]
					_temp_test_type = _temp_test_object['type']
					_temp_test_name = _temp_test_object['name']	
					#print('_temp_test_object', _temp_test_object)
					if _temp_test_type == 'Condition':
						if _temp_test_name == 'End If':
							print('End of a condition')
							#_chain = self.chain_warpped('condition', _condition_block_chain)
							#_condition_block_chain += _chain
							break
							#return _chain
						elif _temp_test_name == 'If':
							# A condition within a condition:
							
							print('Generate a sub-condition within a condition.', _current_loop_index+1)
							_condition_block,_current_condition_index = self.Function_Generate_TestCase(TestCase_Object, 'condition', _current_execution_value, _current_condition_index+1)	
							
							_condition_block_chain = _condition_block_chain + _condition_block
						else:
							print('Invalid type')	
					elif _temp_test_type  in ['Loop', 'Loop List']:
						# Loop within a condition:
					
						print('Generate a sub-loop within a condition.', _current_loop_index+1)
						_loop_block,_current_condition_index = self.Function_Generate_TestCase(TestCase_Object, 'loop', _current_execution_value, _current_condition_index+1)	
						
						_condition_block_chain = _condition_block_chain + _loop_block
					else:
						_chain = self.chain_warpped('action', _temp_test_object)
						_condition_block_chain += _chain
					if _current_condition_index == len(TestCase_Object):
						break
				_chain = self.chain_warpped('condition', _condition_block_chain, condition_string = _condition_string)
				test_case_list += _chain
				_index = _current_condition_index+1
				if recursion_type == 'condition':
					return test_case_list,_index
			elif _test_type == "Comment":
				continue
			else:
				# Action type
				_chain = self.chain_warpped('action', test_object)
				test_case_list += _chain
		return test_case_list, _index+1

	def chain_warpped(self, type, block, current_list_value = [], condition_string = ''):
		_type = type.lower()
		if _type == 'action':
			'''
			{
				'type': 'action',
				'current_list_value': []
				'execute_block': test_object,
			}
			'''	
			return [{'type': _type, 'execute_block': block, 'current_list_value': current_list_value}]
		else:
			'''
			{
				'type': 'condition',
				'condition_string': condition_string,
				'current_list_value': [],
				'execute_block': block,
				'else_block': else_block
			}
			
			'''	
			return [{'type': _type, 'condition_string': condition_string, 'execute_block': block, 'current_list_value': current_list_value}]


	def Function_Import_DB(self, DB_Path):

		#self.StringID = []	
		self.UI = {}
		#db_dir = os.path.dirname(DB_Path)
		#print('Base', db_dir)
		if (os.path.isfile(DB_Path)):
			db_dir, Name, Ext = Split_Path(DB_Path)
		
			if Ext == '.csv':
				with open(DB_Path, newline='', encoding='utf-8-sig') as csvfile:
					reader = csv.DictReader(csvfile)
					#print('ALl DB', reader)
					all_cols = ['StringID', 'String_EN', 'String_KO', 'Path']
					#print('Header', all_cols)
					for entry in reader:
						if not 'StringID' in entry:
							continue
						MyEntry = {}
						StringID = entry['StringID']
						for col in all_cols:
							if col in entry:
								if col == 'Path' :
									_relative_path = entry[col]
									if _relative_path not in [None, '']:
										_obsolute_path = db_dir + '\\'  + _relative_path
										if os.path.isfile(_obsolute_path):	
											MyEntry[col] = _obsolute_path	
								else:
									MyEntry[col] = entry[col]
							else:
								MyEntry[col] = ''
						if 'Path' in MyEntry:
							self.UI[StringID] = MyEntry				
				
			elif Ext in ['.xlsx', '.xlsm']:
				xlsx = load_workbook(DB_Path, data_only=True)
				for sheet in xlsx:	
					sheetname = sheet.title			
					print('Adding DB from: ', sheet)
					DB_Name = sheetname
					Col_StringID = ""
					Col_String_EN = ""
					Col_String_KR = ""
					Col_Path = ""

					ws = xlsx[sheet.title]

					database = None
					ListCol = {}

					#Get Col Label and Letter
					for row in ws.iter_rows():
						for cell in row:
							if cell.value == "StringID":
								Col = cell.column_letter
								Row_ColID = cell.row
								Col_StringID = Col
								ListCol['StringID'] = Col_StringID

							if Col_StringID != "":
								database = ws
								lastChar = Col_StringID
								
								while True:
									lastChar = chr(ord(lastChar) + 1)
									try:
										ColLabel = database[lastChar + str(Row_ColID)].value
									except:
										break	
									if ColLabel in ["",None] :
										break
									else:
										ListCol[ColLabel] = lastChar

								
						if database!=  None:
							break		
					# Load data 			
					if database != None:
						for i in range(Row_ColID, database.max_row): 
							StringID = database[Col_StringID + str(i+1)].value
							#self.StringID.append(StringID)
							MyEntry = {}
							for Label in ListCol:
								if Label == 'Path' :
									_relative_path = database[ListCol[Label] + str(i+1)].value
									if _relative_path not in [None, '']:
										_obsolute_path = db_dir + '\\'  + _relative_path
										if os.path.isfile(_obsolute_path):	
											MyEntry[Label] = _obsolute_path
								else:
									MyEntry[Label] = database[ListCol[Label] + str(i+1)].value
							if 'Path' in MyEntry:
								self.UI[StringID] = MyEntry	
			else:
				print('Unsupported format.')					
		#print(self.UI)

	def Function_Merge_Path(path, folder):
		return folder + '\\' + path

	def Function_Import_Data(self, TestCase_Path, Data_ID):

		Data_ID = Data_ID.lower()
		if TestCase_Path != None:
			if (os.path.isfile(TestCase_Path)):
				xlsx = load_workbook(TestCase_Path, data_only=True)
				#Entry = {}
				Data = []
				for sheet in xlsx:
					
					sheetname = sheet.title.lower()
					print('sheetname:', sheetname)
					if sheetname.find('data_')  > -1:
						DataName = sheetname.replace('data_', '')
						if DataName ==  Data_ID:
							FirstDataRow = None
							EndDataRow = None

							ws = xlsx[sheet.title]
							#Get Col Label and Letter
							for col in ws.iter_cols(min_row=1, max_col=1):
								for cell in col:
									Value = cell.value
									if Value != None:
										Value = Value.lower()
										print('Value:', Value)
										CurrentRow = cell.row
										
										if Value == 'stringid':
											FirstInfoRow = CurrentRow
										else:
											Data.append(Value)

							return Data

			else:
				return([])		
		else:
			return([])	


	def Update_Action_List(self):
		self.action_list = []

		self.append_action_list(type = 'Action', name = 'Tap_Item', argument = {'string_id': 'string_id', 'total_attemp': 'int', 'match_rate': 'float'}, 
			description= 'Tap an image that has been registered in the DB.')

		self.append_action_list(type = 'Action', name = 'Tap_Location', argument = {'location': 'point'}, 
			description= 'Tap on a location select on the screen.')
		# 
		self.append_action_list(type = 'Action', name = 'Tap', argument = {'x': 'int', 'y': 'int'}, 
			description= 'Tap on a point with the input coordinates.')

		self.append_action_list(type = 'Action', name = 'Tap_Template', argument = {'image_path': 'current_area', 'total_attemp': 'int', 'match_rate': 'float'}, 
			description= 'Tap on a center of an image select from the screen.')

		self.append_action_list(type = 'Action', name = 'Relative_Tap', argument = {'string_id': 'string_id', 'delta_X': 'int', 'delta_Y': 'int','match_rate': 'float'}, 
			description= 'Tap on an image the has been registered in the DB, but moving away from the center delta value.')

		self.append_action_list(type = 'Action', name = 'Send_Tab_Key', argument = None, 
			description= 'Press TAB key on the phone, which will move the focus to the next widget.')

		self.append_action_list(type = 'Action', name = 'Send_Enter_Key', argument = None, 
			description= 'Press ENTER key on the phone, which will move the focus to the next widget.')

		self.append_action_list(type = 'Action', name = 'Input_Text', argument = {'input_text':'string'}, 
			description= 'Send a text to the input box.')

		#self.append_action_list(type = 'Get_Result', name = 'Count_Object', argument = {'string_id':'string_id'}, 
		#	description= 'Count the number of image in the current screen.')
		#self.append_action_list(type = 'Update_Variable', name = 'Update_Gacha_Pool', argument = {'db_path':'string', 'db_sheet_name': 'string', 'db_sheet_list':'string'}, description= '')
		#self.append_action_list(type = 'Update_Variable', name = 'Update_Execution_List', argument = {'execute_list':'string'}, description= '')
		#self.append_action_list(type = 'Update_Variable', name = 'Update_Execution_Value', argument = {'execute_value':'string'}, description= '')
		#self.append_action_list(type = 'Get_Result', name = 'Analyse_Gacha_Acquired', argument = {'total_item_in_gacha': 'int'}, description= '')
		#self.append_action_list(type = 'Update_Variable', name = 'Analyse_Gacha_Result', argument = {'total_item_in_gacha': 'int'}, description= '')
		self.append_action_list(type = 'Action', name = 'wait_for_item', argument = {'string_id':'string_id', 'match_rate': 'float', 'timeout': 'int'}, 
			description= 'Wait until an image appears on the screen.')

		self.append_action_list(type = 'Action', name = 'wait_for_template', argument = {'template_path':'current_area', 'match_rate': 'float', 'timeout': 'int'}, 
			description= 'Wait until an image (select from the screen) appears on the screen.')

		self.append_action_list(type = 'Action', name = 'wait_and_tap_item', argument = {'string_id':'string_id', 'match_rate': 'float', 'timeout': 'int'}, 
			description= 'Wait until an image appears on the screen and tap it.')
		
		self.append_action_list(type = 'Action', name = 'wait_and_tap_template', argument = {'template_path':'current_area', 'match_rate': 'float', 'timeout': 'int'}, 
			description= 'Wait until an image (select from the screen) appears on the screen and tap it.')

		self.append_action_list(type = 'Action', name = 'Swipe_by_Location', argument = {'point_A': 'point', 'point_B':'point'}	, 
			description= 'Swipe from a point to another, input with json format.')

		self.append_action_list(type = 'Action', name = 'Swipe_by_StringID', argument = {'string_id_A': 'string_id', 'string_id_B':'string_id'}	, 
			description= 'Swipe from a center of an image to another, select with registerd image in the DB.')

		
		
		
		self.append_action_list(type = 'Action', name = 'Get_Screenshot', argument = {'name': 'string'}, 
			description= 'Save the screenshot of the screen to the result folder of the testcase.')

		self.append_action_list(type = 'Action', name = 'Crop_Image', argument = {'scan_area': 'area', 'name': 'string'}, 
			description= 'Save a selected area of the screen to the result folder of the testcase.')

		#self.append_action_list(type = 'Loop', name = 'List_Loop', argument = {'list_name': 'string'}, description= '')

		self.append_action_list(type = 'Loop', name = 'Loop', argument = {'amount': 'int'}, 
			description= 'Repeat the action between the start and end of the loop.')
		
		self.append_action_list(type = 'Condition', name = 'If', argument = {'condition': 'string'}, 
			description= 'Perform the action if the condition is reached.')
		
		

		self.append_action_list(type = 'Action', name = 'Sleep', argument = {'time': 'int'}, 
			description= 'Idle for an amount of time.')

		self.append_action_list(type = 'Comment', name = 'Comment', argument = {'comment': 'string'}, 
			description= 'Comment for the test step.')

		if self.OCR == True:
			self.append_action_list(type = 'Action', name = 'Scan_Text', argument = {'scan_area': 'area'}, 
				description= 'Scan the text in the selected area.')
			#self.append_action_list(type = 'Action', name = 'Test_Scan_Text', argument = {'scan_area': 'area', '2nd_scan_area': 'area'}, description= '')
		
		
		if self.LoopList == True:
			self.append_action_list(type = 'Loop', name = 'Loop List', argument = {'start_index': 'int','end_index': 'int'}, 
				description= 'Repeat the action between the start and end of the loop, during the loop, new value in the execute list are updated and can be used in some function.')
			self.append_action_list(type = 'Action', name = 'Input_Current_Value', argument = {'indexer': 'user_list'}, 
				description= 'Can be used within the Loop List. Input the current value of the execution list')
			self.append_action_list(type = 'Action', name = 'Tap_Current_Item', argument = {'indexer': 'user_list'}, 
				description= 'Can be used within the Loop List. Tap the current item declare in the execution list')
			self.append_action_list(type = 'Action', name = 'Wait_For_Current_Item', argument = {'indexer': 'user_list'}, 
				description= 'Can be used within the Loop List. Wait for an item declare in the execution list, and tap on it.')
			
		#self.append_action_list(type = 'Action', name = 'Test_Tap', argument = {'touch_point': 'point', '2nd_touch_point': 'point'}, description= '')

	def Get_Current_Screenshot(self):
		Img_Screenshot = self.Device.screencap()
		return Img_Screenshot

	def Get_Folder(self, Path):

		return os.path.dirname(Path)
		 
	def Update_Result_Path(self, Path):
		self.Result_Path = Path

	def Update_Resolution(self, res):
		self.Resolution = res

	def Update_DB_Path(self, Path):
		#self.DB_Path = Path
		self.Function_Import_DB(Path)

	def Update_Tesseract(self, tess_path, tess_data, tess_lang):
		self.tess_path = tess_path
		self.tess_data = tess_data
		self.tess_lang = tess_lang
		self.OCR = True
		self.Update_Action_List()

	def Update_LoopList(self):
		self.LoopList = True
		self.Update_Action_List()	

	def Update_Serial_Number(self, Serial):
		print('Connect to device:', Serial)
		self.Serial = Serial
		self.Device = self.Client.device(Serial)


	def Check_Connectivity(self):
		status = None
		try:
			status = self.Client.device(self.Serial)
			#print('Status', status)
		except Exception as e:
			#print('Error while checking connection:', e)
			return False	
		if status == None:
			return False
		return True	

	def Generate_Result(self, Type = None, Status = None, Details = None, Screenshot = []):
		ReturnResult = {}
		if Type != None:
			ReturnResult['Type'] = Type
		else:
			ReturnResult['Type'] = 'Execute'

		if Status != None:
			if Status == True:
				ReturnResult['Status'] = "Pass"
			else:
				ReturnResult['Status'] = "Fail"	
		else:
			ReturnResult['Status'] = "Fail"
		
		if Details != None:
			ReturnResult['Details'] = Details

		if len(Screenshot) != 0:
			ReturnResult['Screenshot'] = Screenshot

		return 	ReturnResult

	def Update_Result_Array(self, Result):
		self.Result_Array.append(Result)

	def Sleep(self, Time):
		STime = int(Time)/ 1000
		time.sleep(STime)
		return self.Generate_Result(Status = True)

	def Tap(self, x, y):
		self._tap(x, y)		
		return self.Generate_Result(Status = True)

	def Tap_Item(self, string_id, total_attemp = 5, match_rate = 0.8):

		Img_Path = self.UI[string_id]['Path']	
		#cv2.imshow('template', Img_Template)
		#cv2.waitKey(0)
		for i in range(total_attemp):
			try:
				result = self._get_item_location(Img_Path, 0.)
			except Exception as e:
				result = False
				#print('Error from Tap_Item:', e)
			if result != False:
				self._tap_location(result)
				return self.Generate_Result(Status = True)
		return self.Generate_Result(Status = False)	

	def Tap_Location(self, location):
		self._tap_location(location)
		return self.Generate_Result(Status = True)
 

	def Tap_Template(self, image_path, total_attemp = 5, match_rate = 0.8):
		if not os.path.isfile(image_path):
			return self.Generate_Result(Status = False)

		#_img_template = read_img(image_path)
		
		for i in range(total_attemp):
			try:
				result = self._get_item_location(image_path, match_rate)
				print('Tap_Template', result)
			except Exception as e:
				result = False
				print('Error from Tap_Item:', e)
			if result:
				self._tap_location(result)
				return self.Generate_Result(Status = True)

		
		return self.Generate_Result(Status = False)

	def Relative_Tap(self, string_id, delta_X=0, delta_Y=0, match_rate = 0.8):
		_img_template = self.UI[string_id]['Image']	

		Loc = self._get_item_location(_img_template, match_rate)
		
		Loc['x'] += int(delta_X)
		Loc['y'] += int(delta_Y)

		if Loc:		
			self._raw_tap(Loc['x'], Loc['y'])
			ResultStatus = True
		else:
			ResultStatus = False
		
		return self.Generate_Result(Status = ResultStatus)	

	#def Four_Touch(self):
	#	Four_Touch()

	#def Three_Touch(self):
	#	Three_Touch()

	def Count_Object(self, string_id):
		img_template_path = self.UI[string_id]['Image']
		count = self._count_object(img_template_path)
		if count == False:
			return self.Generate_Result(Status = False)
		else:
			return self.Generate_Result(Status = True, Details = count)


	def Scan_Text(self, scan_area):
		try:
			_img = self.Get_Screenshot_In_Working_Resolution()
			imCrop = _img[int(scan_area[1]):int(scan_area[1]+scan_area[2]), int(scan_area[0]):int(scan_area[0]+scan_area[3])]
			text = get_text_from_image(self.tess_path, self.tess_lang, self.tess_data, imCrop)
		except:
			return self.Generate_Result(Status = False)
		return self.Generate_Result(Status = text)

	def Sleep(self, time):
		Sleep(time)
		return self.Generate_Result(Status = True)

	def Crop_Image(self, scan_area, name = 'Crop_IMG_'):
		#print(scan_area)
		_img = self.Get_Screenshot_In_Working_Resolution()
		imCrop = _img[int(scan_area['y']):int(scan_area['y']+scan_area['h']), int(scan_area['x']):int(scan_area['x']+scan_area['w'])]
		Img_Name = Correct_Path(name + '_' + Function_Get_TimeStamp() + '.png', self.Result_Path)
		print('Save image to: ', Img_Name)
		ResultStatus = True
		try:
			cv2.imwrite(Img_Name, imCrop)
		except:
			ResultStatus = False
		return self.Generate_Result(Status = ResultStatus)
			


	def Analyse_Gacha_Acquired(self, Gacha_Amount = 11):
		Img_Screenshot = self.Get_Screenshot_In_Working_Resolution()
		#Img_Screenshot = cv2.cvtColor(Img_Screenshot, cv2.COLOR_GRAY2BGR)
	
		Gacha_Pool = self.Gacha_Pool
		Gacha_Result = {}

		for StringID in Gacha_Pool:
			if  'Image' in Gacha_Pool[StringID]:
				result = False
				result, Img_Screenshot = self._count_object(Gacha_Pool[StringID]['Image'], 0.95)
				
				if result != False:
					Gacha_Result[Gacha_Pool[StringID]['StringID']] = result
					
		Acquired = 0
		Details = {}
		Details['Data'] = Gacha_Result
		Details['Image'] = Img_Screenshot
		for item in Gacha_Result:
			Acquired += int(Gacha_Result[item])
			
		if 	Acquired != Gacha_Amount:
			Status_Result = False
			Name = Correct_Path('Gacha result_' + Function_Get_TimeStamp() + '.png', 'Test Result')
			cv2.imwrite(Name, Img_Screenshot) 
		else:
			Status_Result = True	
		

		return self.Generate_Result(Type = 'Result', Status = Status_Result, Screenshot = Img_Screenshot, Details = Gacha_Result)

	def Analyse_Gacha_Result(self, Total_Item_In_Gacha= 11):
		Count = 0
		Gacha_Pool = self.Gacha_Pool
		Gacha_Result = {}

		Item_Not_In_Pool = False

		for StringID in self.Gacha_Pool:
			Gacha_Result[StringID] = 0
		
		for Result in self.Result_Array:
			if Result['Name'] == "Analyse_Gacha_Acquired":
				Count+=1
				Within_Count = 0
				for StringID in Result['Detail']:
					ItemAmount = Result['Detail'][StringID]
					Gacha_Result[StringID] += ItemAmount
					Within_Count += ItemAmount

			if Within_Count != Total_Item_In_Gacha:
				Item_Not_In_Pool = True
			
		Fail_Details = ""
		Item_Missing = []
		for StringID in Gacha_Result:
			print(StringID)
			if Gacha_Result[StringID] == 0:
				Item_Missing.append(StringID)
		
		if len(Item_Missing) >0:
			Status_Result = False
			Fail_Details+= "Item didn't acquire: " + str(Item_Missing) + '.\r\n'
		else:
			Status_Result = True	

		if Item_Not_In_Pool:
			Status_Result = False
			Fail_Details+= "Acquired amount not match.\r\n"

		return self.Generate_Result(Type = 'Execute', Status = Status_Result, Details = Fail_Details)

	def wait_for_item(self, StringID, match_rate = 0.9, timeout=15):
		#print('Local:', locals())
		Start = time.time()
		Wait_Time = timeout * 1000
		Now = Start
		while (Now - Start) < Wait_Time:
			Now = time.time()
			result, Img_Screenshot = self._get_item_location(self.UI[StringID]['Image'], match_rate)
			if result != False:
				ResultStatus = True
				break
			else:
				ResultStatus = False

		return self.Generate_Result(Status = ResultStatus, Screenshot = Img_Screenshot)

	def wait_for_template(self, template_path, match_rate = 0.9, timeout=15):
		#print('Local:', locals())
		Start = time.time()
		Wait_Time = timeout
		Now = Start
		while (Now - Start) < Wait_Time:
			print('Waiting time:', str(Now- Start))
			Now = time.time()
			result = self._get_item_location(template_path, match_rate)
			print('Item location:', result)
			if result != False:
				ResultStatus = True
				break
			else:
				ResultStatus = False

		return self.Generate_Result(Status = ResultStatus)	

	def wait_and_tap_item(self, StringID, match_rate = 0.9, timeout=15):
		#print('Local:', locals())
		Start = time.time()
		Wait_Time = timeout * 1000
		Now = Start
		while (Now - Start) < Wait_Time:
			Now = time.time()
			result = self._get_item_location(self.UI[StringID]['Image'], match_rate)
			if result != False:
				result = self._tap_location(result)
				return self.Generate_Result(Status = result)
			else:
				ResultStatus = False

		return self.Generate_Result(Status = ResultStatus)

	def wait_and_tap_template(self, template_path, match_rate = 0.9, timeout=15):
		#print('Local:', locals())
		Start = time.time()
		Wait_Time = timeout
		Now = Start
		while (Now - Start) < Wait_Time:
			print('Waiting time:', str(Now- Start))
			Now = time.time()
			result = self._get_item_location(template_path, match_rate)
			#print('Count issue:', result)
			if result != False:
				print('wait result:', result)
				result = self._tap_location(result)
				return self.Generate_Result(Status = result)
			else:
				ResultStatus = False

		return self.Generate_Result(Status = ResultStatus)	

	def Swipe_by_Location(self, point_A, point_B):
		
		result = self._swipe_location(point_A, point_B)

		return self.Generate_Result(Status = result)

	def Swipe_by_StringID(self, StringID_A, StringID_B):
	
		Template_A_Path = self.UI[StringID_A]['Path']
		Loc_A = self._get_item_location(Template_A_Path)
		
		Template_B_Path = self.UI[StringID_B]['Path']
		Loc_B= self._get_item_location(Template_B_Path)

		result = self._swipe_location(Loc_A, Loc_B)

		return self.Generate_Result(Status = result)

	def Send_Enter_Key(self):
		result = self._raw_send_key('66')
		ResultStatus = True
		if result != True:
			ResultStatus = False
		return self.Generate_Result(Status = ResultStatus)


	def Send_Tab_Key(self):
		result = self._raw_send_key('61')
		ResultStatus = True
		if result != True:
			ResultStatus = False
		return self.Generate_Result(Status = ResultStatus)

	def Send_BackKey_Key(self):
		result = self._raw_send_key('4')
		ResultStatus = True
		if result != True:
			ResultStatus = False

		return self.Generate_Result(Status = ResultStatus)
		

	def Input_Text(self, input_text):
		self._raw_send_text(input_text)
		ResultStatus = True
		return self.Generate_Result(Status = ResultStatus)


############################################################################
# FIMCTOPM THAT USE CURRENT LIST VALUE
############################################################################
	def Input_Current_Value(self,indexer):
		_current_item = self.Execution_Value[indexer]
		self._raw_send_text(_current_item)
		ResultStatus = True
		return self.Generate_Result(Status = ResultStatus)

	def Tap_Current_Item(self,indexer):
		_current_item = self.Execution_Value[indexer]
		self.Tap_Item(_current_item)
		ResultStatus = True
		return self.Generate_Result(Status = ResultStatus)

	def Wait_For_Current_Item(self,indexer):
		_current_item = self.Execution_Value[indexer]
		self.wait_for_item(_current_item)	
		ResultStatus = True
		return self.Generate_Result(Status = ResultStatus)
	
	def Get_Screenshot(self, name = 'Screenshot_'):
		Image = self.Get_Screenshot_In_Working_Resolution()
		Img_Name = Correct_Path(name + '_' + Function_Get_TimeStamp() + '.png', self.Result_Path)
		print('Save image to: ', Img_Name)
		ResultStatus = True
		try:
			cv2.imwrite(Img_Name, Img_Name)
		except:
			ResultStatus = False
		return self.Generate_Result(Status = ResultStatus)

	def Get_Ratio(self):
		Img_Screenshot = self.Device.screencap()
		Img_Screenshot = np.asarray(Img_Screenshot)
		Img_Screenshot = cv2.imdecode(Img_Screenshot, cv2.IMREAD_COLOR)
		(_h, _w) = Img_Screenshot.shape[:2]
	
		resolution= self.Resolution
		if _w > _h:
			_ratio = resolution / _h
		else:
			_ratio = resolution / _w
		print('Current ratio:', _ratio)
		return _ratio


###################################################################################################
# BASIC FUNCTION
###################################################################################################
	def _raw_tap(self, x, y):
		try:
			command = "input tap " + str(int(x)) + " " + str(int(y))
			print('Tap:', command)
			self.Device.shell(command)
			return True
		except Exception as e:
			return e	

	def _raw_swipe(self, x1, y1, x2, y2):
		try:
			command = "input swipe " + str(x1) + " " + str(y1) + " " + str(x2) + " " + str(y2)
			print('Swipe:', command)
			self.Device.shell(command)
			return True
		except Exception as e:
			return e	

	def _raw_send_key(self, KeyID):
		try:
			command = "input keyevent \'%s\'" %KeyID
			print('Command: ', command)
			self.Device.shell(command)
			return True
		except Exception as e:
			return e	

	def _raw_send_text(self, text):
		try:
			command = "input text \'%s\'" %text
			print('Command: ', command)
			self.Device.shell(command)
			return True
		except Exception as e:
			return e

	
	def _count_object(self, img_template_path, match_rate = 0.80):

		img_template = cv2.imread(img_template_path)
		template = cv2.cvtColor(img_template, cv2.COLOR_BGR2GRAY)
		#template = cv2.Canny(template, 50, 200)
		(tH, tW) = template.shape[:2]
		#image = cv2.bitwise_not(Img_Screenshot)
		_img_screenshot = self.Get_Screenshot_In_Working_Resolution() 
		gray = cv2.cvtColor(_img_screenshot, cv2.COLOR_BGR2GRAY)
		#cv2.imshow("template", template)
		#cv2.waitKey(0)

		matches = []
		Counter = 0
		# detect edges in the resized, grayscale image and apply template
		# matching to find the template in the image
		#edged = cv2.Canny(resized, 50, 200)
		res = cv2.matchTemplate(gray, template,cv2.TM_CCOEFF_NORMED)
		loc = np.where( res >= match_rate)
		for pt in zip(*loc[::-1]):
			intersection = 0
			for match in matches:
				x1 = int(match[0] + 0.5 * tW)
				y1 = int(match[1] + 0.5 * tH)
				x2 = int(pt[0] + 0.5 * tW)
				y2 = int(pt[1] + 0.5 * tH)
				if self._duplicated((x1, y1), (x2, y2), tW, tH):
					intersection = 1
					break

			if intersection == 0:
				matches.append(pt)
				Counter+=1
				#cv2.rectangle(_img_screenshot, pt, (pt[0] + tW, pt[1] + tH), (0, 255, 0), 2)
				#print("Adding item: ",pt, (pt[0] + tW, pt[1] + tH))
	
		if Counter > 0:
			return Counter
		else:
			return False	


	def _swipe(self, x1, y1, x2, y2):
		result = self._raw_swipe(x1/self.Ratio, y1/self.Ratio, x2/self.Ratio, y2/self.Ratio)
		return result

	def _swipe_location(self, point1, point2):
		result = self._raw_swipe(point1['x']/self.Ratio, point1['y']/self.Ratio, point2['x']/self.Ratio, point2['y']/self.Ratio)
		return	result

	def _tap(self, x, y):
		result = self._raw_tap(x/self.Ratio, y/self.Ratio)
		return result

	def _tap_location(self, location):
		result = self._tap(location['x'], location['y'])
		return	result

	def _get_item_location(self, img_template_path, match_rate= 0.80):
		_img_screenshot = self.Get_Screenshot_In_Working_Resolution()
		
		template = cv2.imread(img_template_path)
		#template = cv2.cvtColor(template, cv2.COLOR_RGB2BGR)
		template = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
		#template = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
		(tH, tW) = template.shape[:2]
		_img_screenshot = cv2.cvtColor(_img_screenshot, cv2.COLOR_BGR2GRAY)
		#cv2.imshow('Template',template)
		#cv2.imshow('Scr',_img_screenshot)
		#cv2.waitKey(0)
		Found = None
		Loc = None
		result = cv2.matchTemplate(_img_screenshot, template, cv2.TM_CCOEFF_NORMED)
		(_, maxVal, _, maxLoc) = cv2.minMaxLoc(result)
		#match_rate *=0.9
		if Found is None or maxVal > Found[0]:
			
			Found = (maxVal, maxLoc)
			if maxVal >= match_rate:
				print('maxVal', maxVal, 'match_rate', match_rate)
				(startX, startY) = (int(maxLoc[0]), int(maxLoc[1]))
				(endX, endY) = (int((maxLoc[0] + tW)), int((maxLoc[1] + tH)))
				Loc = {"x": int((maxLoc[0] + 0.5 * tW)), "y": int((maxLoc[1] + 0.5 * tH)), "w": int(abs(startX-endX)), "h": int(abs(startY-endY))}

		if Loc != None:
			# Return the real location on the screen:
			for key in Loc:
				Loc[key]
			return Loc
		else:
			return False	


	def Get_Screenshot_In_Working_Resolution(self, resolution=None):
		
		Img_Screenshot = self.Device.screencap()
		Img_Screenshot = np.asarray(Img_Screenshot)
		Img_Screenshot = cv2.imdecode(Img_Screenshot, cv2.IMREAD_COLOR)
		#Img_Screenshot = cv2.cvtColor(Img_Screenshot , cv2.COLOR_BGR2RGB)
		if resolution != None:
			(_h, _w) = Img_Screenshot.shape[:2]
			if _w > _h:
				_ratio = resolution / _h
			else:
				_ratio = resolution / _w
		else:
			_ratio = self.Ratio
		print('Rescale ratio:', _ratio)
		if _ratio != 1:
			width = int(Img_Screenshot.shape[1] * _ratio)
			height = int(Img_Screenshot.shape[0] * _ratio)
			dim = (width, height)
			Img_Screenshot = cv2.resize(Img_Screenshot, dim, interpolation = cv2.INTER_AREA)
		#cv2.imshow('Test', Img_Screenshot)
		#cv2.waitKey(0)
		return Img_Screenshot		

		
	def _duplicated(self, center1, center2, tw, th):
		
		if abs(center1[0] - center2[0]) <= tw:
			if abs(center1[1] - center2[1]) <= th:
				#print('Dup: ',  center1, center2, tw, th)
				return True	
		#print('Not Dup: ',  center1, center2, tw, th)		
		return False

	def _intersected(self, bottom_left1, top_right1, bottom_left2, top_right2):
		if top_right1[0] < bottom_left2[0] or bottom_left1[0] > top_right2[0]:
			return 0
		if top_right1[1] < bottom_left2[1] or bottom_left1[1] > top_right2[1]:
			return 0
		return 1

def Init_Folder(FolderPath):
	if not os.path.isdir(FolderPath):
		try:
			os.mkdir(FolderPath)
			print('Create new folder:', FolderPath)
			return True
		except OSError:
			print ("Creation of the directory %s failed" % FolderPath)
			return False

def Function_Get_TimeStamp():		
	now = datetime.now()
	timestamp = str(int(datetime.timestamp(now)))			
	return timestamp

def Sleep(total_miliseconds):
	time.sleep(total_miliseconds/1000)

def Split_Path(Path):
	Outputdir = os.path.dirname(Path)
	baseName = os.path.basename(Path)
	sourcename, ext = os.path.splitext(baseName)
	return [Outputdir, sourcename, ext]

def get_text_from_image(tess_path, tess_language, tess_data, input_image):
	pytesseract.pytesseract.tesseract_cmd = tess_path
	advanced_tessdata_dir_config = '--psm 7 --tessdata-dir ' + '"' + tess_data + '"'
	ocr = pytesseract.image_to_string(input_image, lang = tess_language, config=advanced_tessdata_dir_config)
	ocr = ocr.replace("\n", "").replace("\r", "").replace("\x0c", "")
	return ocr

def Correct_Path(path, Folder = 'DB'):
	print("Folder", Folder)
	print('Path', path )
	if not os.path.isdir(Folder):
		try:
			os.mkdir(Folder)
		except OSError:
			return False
	return Folder + '\\' + path