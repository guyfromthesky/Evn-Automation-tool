from ppadb.client import Client as AdbClient
import os
import cv2
import numpy as np

from openpyxl import load_workbook
from datetime import datetime
import time
import imutils
cwd = os.path.dirname(os.path.realpath(__file__))

Project = "V4"
Features = ['Gacha - Companion', 'Call cheat']

################################################################################################################


def Function_Import_DB(DB_Path, List_Sheet = [], List):
	Start = time.time()
	if DB_Path != None:
		if (os.path.isfile(DB_Path)):
			xlsx = load_workbook(DB_Path, data_only=True)
			MyDB = {}
			#Entry = {}

			for sheet in xlsx:
				
				sheetname = sheet.title
				if sheetname in List_Sheet or len(List_Sheet) == 0:	
					print('Checking sheet: ', sheet)
					DB_Name = sheetname
					Col_StringID = ""
					Col_String_EN = ""
					Col_String_KR = ""
					Col_Path = ""

					ws = xlsx[sheet.title]

					database = None
					ListCol = {}

					#Get Col Label and Letter
					for row in ws.iter_rows():
						for cell in row:
							if cell.value == "StringID":
								Col = cell.column_letter
								Row_ColID = cell.row
								Col_StringID = Col
								ListCol['StringID'] = Col_StringID

							if Col_StringID != "":
								database = ws
								lastChar = Col_StringID
								
								while True:
									lastChar = chr(ord(lastChar) + 1)
									try:
										ColLabel = database[lastChar + str(Row_ColID)].value
									except:
										break	
									if ColLabel in ["",None] :
										break
									else:
										ListCol[ColLabel] = lastChar

								
						if database!=  None:
							break		
					# Load data 			
					if database != None:
						for i in range(Row_ColID, database.max_row): 
							StringID = database[Col_StringID + str(i+1)].value
							MyEntry = {}
							for Label in ListCol:
								if Label == 'Path':
									Path = database[ListCol[Label] + str(i+1)].value
									Path = Correct_Path(Path)
									if Path != None:
										if os.path.isfile(Path):		
											MyEntry['Image'] = cv2.imread(Path)
											MyEntry[Label] = database[ListCol[Label] + str(i+1)].value

									
								else:
									MyEntry[Label] = database[ListCol[Label] + str(i+1)].value
							if 'Image' in MyEntry:
								MyDB[StringID] = MyEntry
			End = time.time()
			Message = "Total time spend: " + str(End-Start) + ' miliseconds.'
			print(Message)
			return MyDB
		else:
			return({})	
	else:
		return({})

def Function_Get_TimeStamp():		
	now = datetime.now()
	timestamp = str(int(datetime.timestamp(now)))			
	return timestamp

################################################################################################################

def Sleep(Total_miliseconds):
	Start = int(1000*time.time())
	End = Start + Total_miliseconds
	Now = Start
	while Now < End:
		Now = int(1000*time.time())
	return


def Correct_Path(path):
	return cwd + '//DB//' + path

################################################################################################################
#Tap on Location_Object

def Tap(Device, Location_Object):
	command = "input tap " + str(Location_Object['x']) + " " + str(Location_Object['y'])
	Device.shell(command)
	return

#Swipe from A -> B
def Swipe(Device, Location_Object_A, Location_Object_B):
	command = "input swipe " + str(Location_Object_A['x']) + " " + str(Location_Object_A['y']) + " " + str(Location_Object_B['x']) + " " + str(Location_Object_B['y'])
	print(command)
	Device.shell(command)
	return

#Swipe up
def Swipe_Up(Device, Location_Object_A, Range):

	Location_Object_B = {'x': Location_Object_A['x'], 'y': Location_Object_A['y'] - Range} 
	print('Swipe from', Location_Object_A, 'to', Location_Object_B)
	command = "input swipe " + str(Location_Object_A['x']) + " " + str(Location_Object_A['y']) + " " + str(Location_Object_B['x']) + " " + str(Location_Object_B['y'])
	print(command)
	Device.shell(command)
	return

def Send_Text(Device, Text):
	command = "input text \' %s \'" %Text
	Device.shell(command)
	return


################################################################################################################

def HD_Resize(Img):
	
	scale_width = 1280
	scale_percent = int(scale_width * 100 / Img.shape[1])
	print('scale_percent', scale_percent)
	#width = int(img.shape[1] * scale_percent / 100)
	scale_height = int(Img.shape[0] * scale_percent / 100)
	dim = (scale_width, scale_height)
	# resize image
	resized = cv2.resize(Img, dim, interpolation = cv2.INTER_AREA)
	return resized

def Resize(Img, scale_percent):

	width = int(Img.shape[1] * scale_percent / 100)
	height = int(Img.shape[0] * scale_percent / 100)
	dim = (width, height)
	# resize image
	resized = cv2.resize(Img, dim, interpolation = cv2.INTER_AREA)
	return resized

def Draw_Line(Img, Location_Object_A, Location_Object_B):
	cv2.line(Img,(Location_Object_A['x'],Location_Object_A['y']),(Location_Object_B['x'],Location_Object_B['y']),(0,0,255),2)
	return Img

def Search_Best_Match(Img_Screenshot, Img_Template, Match_Rate=0.9):

	Start = time.time()
	template = cv2.cvtColor(Img_Template, cv2.COLOR_BGR2GRAY)
	template = cv2.Canny(template, 50, 200)
	(tH, tW) = template.shape[:2]
	image = np.asarray(Img_Screenshot)
	gray = cv2.imdecode(image, cv2.COLOR_BGR2GRAY)
	Found = None
	for scale in np.linspace(0.2, 1.0, 20)[::-1]:
		resized = imutils.resize(gray, width = int(gray.shape[1] * scale))
		r = gray.shape[1] / float(resized.shape[1])

		if resized.shape[0] < tH or resized.shape[1] < tW:
			break
		edged = cv2.Canny(resized, 50, 200)
		result = cv2.matchTemplate(edged, template, cv2.TM_CCOEFF)
		(_, maxVal, _, maxLoc) = cv2.minMaxLoc(result)
		if Found is None or maxVal > Found[0]:		
			Found = (maxVal, maxLoc, r)
	(_, maxLoc, r) = Found
	(startX, startY) = (int(maxLoc[0] * r), int(maxLoc[1] * r))
	(endX, endY) = (int((maxLoc[0] + tW) * r), int((maxLoc[1] + tH) * r))
	
	End = time.time()
	Message = "Total time spend: " + str(End-Start) + ' miliseconds.'
	print(Message)
	
	if Found != None:
		Loc = {"x": int((maxLoc[0] + 0.5 * tW) * r), "y": int((maxLoc[1] + 0.5 * tH) * r), "w": int(abs(startX-endX)), "h": int(abs(startY-endY))}	
		
		return Loc
	else:
		return False

def Search_All_Object(Img_Screenshot, Img_Template, Match_Rate=0.9):

	Start = time.time()

	template = cv2.cvtColor(Img_Template, cv2.IMREAD_COLOR)
	template = cv2.Canny(template, 50, 200)
	(tH, tW) = template.shape[:2]
	
	image = np.asarray(Img_Screenshot)
	gray = cv2.imdecode(image, cv2.IMREAD_COLOR)
	
	Found = None
	Results = []
	for scale in np.linspace(0.2, 1.0, 20)[::-1]:
		resized = imutils.resize(gray, width = int(gray.shape[1] * scale))
		r = gray.shape[1] / float(resized.shape[1])
		if resized.shape[0] < tH or resized.shape[1] < tW:
			break
		edged = cv2.Canny(resized, 50, 200)
		result = cv2.matchTemplate(edged, template, cv2.TM_CCOEFF_NORMED)
		(_, maxVal, _, maxLoc) = cv2.minMaxLoc(result)

		if Found is None or maxVal > Found[0]:
			Found = (maxVal, maxLoc, r)
			if maxVal > Match_Rate:

				(startX, startY) = (int(maxLoc[0] * r), int(maxLoc[1] * r))
				(endX, endY) = (int((maxLoc[0] + tW) * r), int((maxLoc[1] + tH) * r))
				cv2.rectangle(gray, (startX, startY), (endX, endY), (0, 0, 255), 2)
				Loc = {"x": int((maxLoc[0] + 0.5 * tW) * r), "y": int((maxLoc[1] + 0.5 * tH) * r), "w": int(abs(startX-endX)), "h": int(abs(startY-endY))}
				Results.append(Loc)
	
	End = time.time()
	Message = "Total time spend: " + str(End-Start) + ' miliseconds.'
	print(Message)

	if len(Results) > 0:
		return Results
	else:
		return False	


def Fast_Search(Img_Screenshot, Img_Template, Match_Rate=0.50):

	Start = time.time()

	#Img_Template = cv2.bitwise_not(Img_Template)

	template = cv2.cvtColor(Img_Template, cv2.COLOR_BGR2GRAY)
	#template = cv2.Canny(template, 50, 200)
	(tH, tW) = template.shape[:2]
	print(template.shape[:2])
	print(Img_Screenshot.shape[:2])

	#image = cv2.bitwise_not(Img_Screenshot)
	gray = cv2.cvtColor(Img_Screenshot, cv2.COLOR_BGR2GRAY)
	
	
	#cv2.imshow("template",template)
	#cv2.waitKey(0)

	#cv2.imshow("gray",gray)
	#cv2.waitKey(0)

	Found = None
	Loc = None
	Results = []
	counter = 0
	for scale in np.linspace(0.2, 1.0, 20)[::-1]:
		counter +=1
		resized = imutils.resize(gray, width = int(gray.shape[1] * scale))
		r = gray.shape[1] / float(resized.shape[1])
		if resized.shape[0] < tH or resized.shape[1] < tW:
			break
		#resized = cv2.Canny(resized, 50, 200)
		result = cv2.matchTemplate(resized, template, cv2.TM_CCOEFF_NORMED)
		(_, maxVal, _, maxLoc) = cv2.minMaxLoc(result)

		if Found is None or maxVal > Found[0]:
			Found = (maxVal, maxLoc, r)
			print('Match rate', int(maxVal*100))
			if maxVal >= Match_Rate:
				
				(startX, startY) = (int(maxLoc[0] * r), int(maxLoc[1] * r))
				(endX, endY) = (int((maxLoc[0] + tW) * r), int((maxLoc[1] + tH) * r))
				Loc = {"x": int((maxLoc[0] + 0.5 * tW) * r), "y": int((maxLoc[1] + 0.5 * tH) * r), "w": int(abs(startX-endX)), "h": int(abs(startY-endY))}
				break

	End = time.time()
	Message = "Total time spend for Fast_Search: " + str(End-Start) + ' miliseconds.'
	print(Message)

	if Loc != None:
		return Loc
	else:
		return False


def Count_Object(Img_Screenshot, Img_Template, Match_Rate=0.50):

	Start = time.time()
	Img_Template = Resize(Img_Template, 50)
	#Img_Template = cv2.bitwise_not(Img_Template)

	template = cv2.cvtColor(Img_Template, cv2.COLOR_BGR2GRAY)
	#template = cv2.Canny(template, 50, 200)
	(tH, tW) = template.shape[:2]

	
	
	#image = cv2.bitwise_not(Img_Screenshot)
	gray = cv2.cvtColor(Img_Screenshot, cv2.COLOR_BGR2GRAY)
	#cv2.imshow("template", template)
	#cv2.waitKey(0)


	Counter = 0
	for scale in np.linspace(0.2, 1.0,20)[::-1]:
		# resize the image according to the scale, and keep track
		# of the ratio of the resizing
		resized = imutils.resize(gray, width = int(gray.shape[1] * scale))
		r = gray.shape[1] / float(resized.shape[1])
		# if the resized image is smaller than the template, then break
		# from the loop
		if resized.shape[0] < tH or resized.shape[1] < tW:
			break
		# detect edges in the resized, grayscale image and apply template
		# matching to find the template in the image
		#edged = cv2.Canny(resized, 50, 200)
		res = cv2.matchTemplate(resized, template,cv2.TM_CCOEFF_NORMED)
		#print("res", res)
		loc = np.where( res >= Match_Rate)
		matches = []
		for pt in zip(*loc[::-1]):
			#sensitivity= 100
			#print(round(pt[0]/sensitivity), round(pt[1]/sensitivity))
			intersection = 0
			for match in matches:

				x1 = match[0] + 0.5 * tW
				y1 = match[1] + 0.5 * tH
				x2 = pt[0] + 0.5 * tW
				y2 = pt[1] + 0.5 * tH
				if Duplicated((x1, y1), (x2, y2), tW, tH):
				#if intersected(match, (match[0] + tW, match[1] + tH), pt, (pt[0] + tW, pt[1] + tH)):
					intersection = 1
					break

			if intersection == 0:
				matches.append(pt)
				Counter+=1
				print("Adding", pt, tW, tH)
				cv2.rectangle(Img_Screenshot, pt, (pt[0] + tW, pt[1] + tH), (0, 0, 255), 2)
				
			
	#cv2.imshow('Img_Screenshot', Img_Screenshot)
	#cv2.waitKey(0)
	
	End = time.time()
	Message = "Total time spend for count: " + str(End-Start) + ' miliseconds.'
	
	if Counter > 0:
		#print('Found: ', Counter)
		return Counter
	else:
		return False	

def  Duplicated(center1, center2, tw, th):
	if abs(center1[0] - center2[0]) <= tw:
		if abs(center1[1] - center2[1]) <= th:
			return True	
	return False

def intersected(bottom_left1, top_right1, bottom_left2, top_right2):
    if top_right1[0] < bottom_left2[0] or bottom_left1[0] > top_right2[0]:
        return 0

    if top_right1[1] < bottom_left2[1] or bottom_left1[1] > top_right2[1]:
        return 0

    return 1